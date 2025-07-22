import openpyxl
from datetime import datetime

# Путь к файлу
file_path = "/mnt/f/Работа-Никита/Autosort_tovar/6_Склад_фурнитуры_Овощная_база_Магазин_продажи_01_07_2024_01_07.xlsx"

print("=== АНАЛИЗ ФАЙЛА ПРОДАЖ ===")
print(f"Файл: {file_path.split('/')[-1]}")
print()

try:
    # Открываем файл
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print("1. ИНФОРМАЦИЯ О ФАЙЛЕ:")
    print(f"   - Количество листов: {len(wb.sheetnames)}")
    print(f"   - Названия листов: {', '.join(wb.sheetnames)}")
    print()
    
    # Работаем с первым листом
    ws = wb.active
    print(f"2. АНАЛИЗ ЛИСТА '{ws.title}':")
    
    # Определяем размеры данных
    max_row = ws.max_row
    max_col = ws.max_column
    print(f"   - Максимальная строка: {max_row}")
    print(f"   - Максимальная колонка: {max_col}")
    
    # Читаем заголовки (первая строка)
    headers = []
    print("\n3. ЗАГОЛОВКИ КОЛОНОК:")
    for col in range(1, max_col + 1):
        header = ws.cell(row=1, column=col).value
        headers.append(header)
        print(f"   {col}. {header}")
    
    print("\n4. ПЕРВЫЕ 15 СТРОК ДАННЫХ:")
    print("-" * 100)
    
    # Выводим заголовки
    header_line = " | ".join([str(h)[:20] if h else "None" for h in headers[:7]])  # Первые 7 колонок
    print(header_line)
    print("-" * 100)
    
    # Выводим данные
    for row in range(2, min(17, max_row + 1)):  # Строки 2-16 (15 строк данных)
        row_data = []
        for col in range(1, min(8, max_col + 1)):  # Первые 7 колонок
            value = ws.cell(row=row, column=col).value
            if isinstance(value, datetime):
                value = value.strftime("%d.%m.%Y")
            row_data.append(str(value)[:20] if value else "")
        print(" | ".join(row_data))
    
    # Анализ типов данных в колонках
    print("\n5. ПРИМЕРЫ ДАННЫХ ПО КОЛОНКАМ:")
    for col_idx, header in enumerate(headers, 1):
        print(f"\n   Колонка {col_idx}: {header}")
        
        # Собираем несколько непустых значений
        sample_values = []
        value_types = set()
        
        for row in range(2, min(50, max_row + 1)):
            value = ws.cell(row=row, column=col_idx).value
            if value is not None:
                value_types.add(type(value).__name__)
                if len(sample_values) < 5 and value not in sample_values:
                    sample_values.append(value)
        
        print(f"   - Типы данных: {', '.join(value_types)}")
        print(f"   - Примеры значений:")
        for val in sample_values:
            if isinstance(val, datetime):
                print(f"     • {val.strftime('%d.%m.%Y %H:%M:%S')}")
            else:
                print(f"     • {str(val)[:50]}")
    
    # Подсчет непустых строк
    non_empty_rows = 0
    for row in range(2, max_row + 1):
        if any(ws.cell(row=row, column=col).value for col in range(1, max_col + 1)):
            non_empty_rows += 1
    
    print(f"\n6. СТАТИСТИКА:")
    print(f"   - Всего строк с данными: {non_empty_rows}")
    
    # Анализ дат
    print("\n7. АНАЛИЗ ПЕРИОДА ДАННЫХ:")
    date_cols = []
    for col_idx, header in enumerate(headers, 1):
        if header and ('дата' in str(header).lower() or 'date' in str(header).lower()):
            date_cols.append((col_idx, header))
    
    if date_cols:
        for col_idx, header in date_cols:
            dates = []
            for row in range(2, max_row + 1):
                value = ws.cell(row=row, column=col_idx).value
                if isinstance(value, datetime):
                    dates.append(value)
            
            if dates:
                print(f"   Колонка '{header}':")
                print(f"   - Минимальная дата: {min(dates).strftime('%d.%m.%Y')}")
                print(f"   - Максимальная дата: {max(dates).strftime('%d.%m.%Y')}")
                print(f"   - Количество уникальных дат: {len(set(dates))}")
    
    wb.close()
    
except Exception as e:
    print(f"Ошибка при чтении файла: {e}")
    import traceback
    traceback.print_exc()