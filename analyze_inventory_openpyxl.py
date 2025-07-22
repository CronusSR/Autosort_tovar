#!/usr/bin/env python3
"""Анализ структуры файла остатков с использованием openpyxl"""

import json
from pathlib import Path
from openpyxl import load_workbook

def analyze_inventory_file(file_path):
    """Детальный анализ файла остатков"""
    
    print(f"Анализ файла: {file_path}")
    print("=" * 80)
    
    try:
        # Загружаем файл
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Получаем все данные
        data = list(ws.values)
        
        if not data:
            print("❌ Файл пустой")
            return
            
        # Предполагаем, что первая строка - заголовки
        headers = data[0]
        rows = data[1:]
        
        # 1. Основная информация
        print("\n1. ОСНОВНАЯ ИНФОРМАЦИЯ:")
        print(f"   - Количество строк с данными: {len(rows)}")
        print(f"   - Количество колонок: {len(headers)}")
        print(f"   - Активный лист: {ws.title}")
        print(f"   - Все листы: {wb.sheetnames}")
        
        # 2. Структура колонок
        print("\n2. СТРУКТУРА КОЛОНОК:")
        for i, col in enumerate(headers):
            print(f"   [{i}] {col}")
            
        # 3. Первые 15 строк
        print("\n3. ПЕРВЫЕ 15 СТРОК:")
        print("-" * 80)
        
        # Показываем первые несколько колонок для читаемости
        max_cols_to_show = min(6, len(headers))
        
        # Заголовки
        header_line = " | ".join([str(headers[i])[:20] for i in range(max_cols_to_show)])
        print(header_line)
        print("-" * len(header_line))
        
        # Данные
        for row_idx, row in enumerate(rows[:15]):
            if row:
                row_line = " | ".join([str(row[i] if i < len(row) and row[i] is not None else '')[:20] 
                                      for i in range(max_cols_to_show)])
                print(f"{row_idx+1:3d}. {row_line}")
                
        if len(headers) > max_cols_to_show:
            print(f"\n... и еще {len(headers) - max_cols_to_show} колонок")
            
        # 4. Анализ складов/филиалов
        print("\n4. АНАЛИЗ СКЛАДОВ/ФИЛИАЛОВ:")
        
        warehouse_cols = []
        warehouse_patterns = ['склад', 'филиал', 'магазин', 'барыс', 'абая', 'айнабулак']
        
        for i, col in enumerate(headers):
            if col and any(pattern in str(col).lower() for pattern in warehouse_patterns):
                warehouse_cols.append((i, col))
                
        if warehouse_cols:
            print(f"   Найдены колонки, связанные со складами:")
            for idx, col in warehouse_cols:
                print(f"     - [{idx}] {col}")
                
        # 5. Поиск колонок с остатками (числовые колонки)
        print("\n5. ПОИСК КОЛОНОК С ОСТАТКАМИ:")
        
        numeric_cols = []
        for i, col in enumerate(headers):
            if col:
                # Проверяем первые несколько значений в колонке
                numeric_count = 0
                for row in rows[:10]:
                    if row and i < len(row) and row[i] is not None:
                        try:
                            float(row[i])
                            numeric_count += 1
                        except (ValueError, TypeError):
                            pass
                            
                if numeric_count > 5:  # Если больше половины значений числовые
                    numeric_cols.append((i, col))
                    
        if numeric_cols:
            print(f"   Найдено {len(numeric_cols)} числовых колонок:")
            for idx, col in numeric_cols[:10]:
                print(f"     - [{idx}] {col}")
                
        # 6. Анализ номенклатуры
        print("\n6. АНАЛИЗ НОМЕНКЛАТУРЫ:")
        
        product_cols = []
        product_patterns = ['товар', 'номенклатура', 'наименование', 'название', 'артикул', 'продукт']
        
        for i, col in enumerate(headers):
            if col and any(pattern in str(col).lower() for pattern in product_patterns):
                product_cols.append((i, col))
                
        if product_cols:
            print(f"   Найдены колонки товаров:")
            for idx, col in product_cols:
                print(f"     - [{idx}] {col}")
                # Показываем примеры
                print("       Примеры:")
                unique_values = set()
                for row in rows[:50]:
                    if row and idx < len(row) and row[idx]:
                        unique_values.add(str(row[idx]))
                for val in list(unique_values)[:5]:
                    print(f"         • {val}")
                    
        # 7. Поиск категорий
        print("\n7. ПОИСК КАТЕГОРИЙ:")
        
        category_cols = []
        category_patterns = ['категория', 'группа', 'раздел', 'тип', 'класс']
        
        for i, col in enumerate(headers):
            if col and any(pattern in str(col).lower() for pattern in category_patterns):
                category_cols.append((i, col))
                
        if category_cols:
            print(f"   Найдены колонки категорий:")
            for idx, col in category_cols:
                print(f"     - [{idx}] {col}")
                # Показываем уникальные категории
                print("       Уникальные значения:")
                unique_values = set()
                for row in rows:
                    if row and idx < len(row) and row[idx]:
                        unique_values.add(str(row[idx]))
                for val in sorted(list(unique_values))[:10]:
                    print(f"         • {val}")
                if len(unique_values) > 10:
                    print(f"         ... и еще {len(unique_values) - 10} категорий")
                    
        # 8. Специальный анализ структуры остатков
        print("\n8. СТРУКТУРА ДАННЫХ ОСТАТКОВ:")
        
        # Ищем колонки, которые могут быть складами с остатками
        possible_warehouse_stock_cols = []
        
        for i, col in enumerate(headers):
            if col:
                col_str = str(col)
                # Проверяем, похоже ли на название склада
                if (any(char.isdigit() for char in col_str) or 
                    any(place in col_str.lower() for place in ['барыс', 'абая', 'айнабулак', 'склад', 'магазин'])):
                    # Исключаем явно не складские колонки
                    exclude_patterns = ['артикул', 'код', 'дата', 'цена', 'категория', 'группа']
                    if not any(pattern in col_str.lower() for pattern in exclude_patterns):
                        # Проверяем, есть ли числовые значения
                        has_numbers = False
                        for row in rows[:10]:
                            if row and i < len(row) and row[i] is not None:
                                try:
                                    float(row[i])
                                    has_numbers = True
                                    break
                                except:
                                    pass
                        if has_numbers:
                            possible_warehouse_stock_cols.append((i, col))
                            
        if possible_warehouse_stock_cols:
            print(f"   Найдено {len(possible_warehouse_stock_cols)} колонок с остатками по складам:")
            for idx, col in possible_warehouse_stock_cols:
                print(f"     - [{idx}] {col}")
                # Показываем примеры значений
                values = []
                for row in rows[:5]:
                    if row and idx < len(row) and row[idx] is not None:
                        values.append(row[idx])
                if values:
                    print(f"       Примеры остатков: {values}")
                    
        # 9. Анализ первой строки данных подробно
        print("\n9. ДЕТАЛЬНЫЙ АНАЛИЗ ПЕРВОЙ СТРОКИ:")
        if rows and rows[0]:
            print("   Значения по колонкам:")
            for i, (header, value) in enumerate(zip(headers, rows[0])):
                if header and value is not None:
                    print(f"     [{i}] {header}: {value}")
                    
        # 10. Сохраняем структуру
        structure = {
            "file_name": str(file_path),
            "sheet_name": ws.title,
            "all_sheets": wb.sheetnames,
            "rows": len(rows),
            "columns": len(headers),
            "column_names": [str(h) if h else f"Column_{i}" for i, h in enumerate(headers)],
            "warehouse_columns": [{"index": idx, "name": str(col)} for idx, col in warehouse_cols],
            "product_columns": [{"index": idx, "name": str(col)} for idx, col in product_cols],
            "category_columns": [{"index": idx, "name": str(col)} for idx, col in category_cols],
            "numeric_columns": [{"index": idx, "name": str(col)} for idx, col in numeric_cols],
            "warehouse_stock_columns": [{"index": idx, "name": str(col)} for idx, col in possible_warehouse_stock_cols]
        }
        
        with open('inventory_structure.json', 'w', encoding='utf-8') as f:
            json.dump(structure, f, ensure_ascii=False, indent=2)
            
        print("\n✅ Структура сохранена в файл 'inventory_structure.json'")
        
        wb.close()
        
    except Exception as e:
        print(f"\n❌ Ошибка при анализе файла: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    file_path = Path("остатки на 08.07.2025.xlsx")
    if file_path.exists():
        analyze_inventory_file(file_path)
    else:
        print(f"❌ Файл не найден: {file_path}")