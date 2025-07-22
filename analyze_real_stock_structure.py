#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import openpyxl
import os
import sys

def analyze_stock_file_structure():
    """Анализ реальной структуры файла остатков"""
    
    file_path = "/mnt/f/Работа-Никита/Autosort_tovar/остатки на 08.07.2025.xlsx"
    
    if not os.path.exists(file_path):
        print(f"❌ Файл не найден: {file_path}")
        return
    
    print(f"📁 Анализ файла: {file_path}")
    print("=" * 80)
    
    try:
        # Сначала проверим с openpyxl
        print("🔍 АНАЛИЗ С OPENPYXL:")
        wb = openpyxl.load_workbook(file_path)
        print(f"📊 Листы в файле: {wb.sheetnames}")
        
        # Возьмем первый лист
        ws = wb.active
        sheet_name = wb.active.title
        print(f"📋 Активный лист: {sheet_name}")
        print(f"📏 Размеры листа: {ws.max_row} строк x {ws.max_column} колонок")
        
        print("\n" + "=" * 40)
        print("🔍 ПЕРВЫЕ 30 СТРОК И 15 КОЛОНОК:")
        print("=" * 40)
        
        # Читаем первые строки для понимания структуры
        for row_idx in range(1, min(31, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(16, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is None:
                    cell_value = ""
                else:
                    cell_value = str(cell_value).strip()
                row_data.append(cell_value)
            
            # Показываем только непустые строки или первые 10 строк
            if any(row_data) or row_idx <= 10:
                print(f"Строка {row_idx:2d}: {row_data}")
        
        print("\n" + "=" * 40)
        print("🔍 ПОИСК ЗАГОЛОВКОВ:")
        print("=" * 40)
        
        # Ищем строки с заголовками
        potential_headers = []
        for row_idx in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    row_data.append(str(cell_value).strip())
                else:
                    row_data.append("")
            
            # Проверяем, содержит ли строка ключевые слова
            row_text = " ".join(row_data).lower()
            if any(keyword in row_text for keyword in ["номенклатура", "склад", "наименование", "остаток", "количество"]):
                potential_headers.append((row_idx, row_data))
                print(f"Заголовок в строке {row_idx}: {row_data}")
        
        print("\n" + "=" * 40)
        print("🔍 АНАЛИЗ КОЛОНОК (первые 20):")
        print("=" * 40)
        
        # Анализируем все колонки для понимания их назначения
        for col_idx in range(1, min(21, ws.max_column + 1)):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"\nКолонка {col_letter} (индекс {col_idx}):")
            
            # Показываем первые 10 значений в колонке
            col_values = []
            for row_idx in range(1, min(11, ws.max_row + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    col_values.append(str(cell_value).strip()[:50])  # Обрезаем длинные значения
                else:
                    col_values.append("")
            
            print(f"  Значения: {col_values}")
        
        wb.close()
        
        print("\n" + "=" * 40)
        print("🔍 АНАЛИЗ С PANDAS:")
        print("=" * 40)
        
        # Теперь проверим с pandas
        try:
            # Читаем без заголовков
            df = pd.read_excel(file_path, header=None)
            print(f"📊 Размер DataFrame: {df.shape[0]} строк x {df.shape[1]} колонок")
            
            print("\n📋 Первые 15 строк:")
            for idx, row in df.head(15).iterrows():
                print(f"Строка {idx:2d}: {list(row.values[:10])}")  # Показываем первые 10 колонок
            
            print("\n📋 Колонки с данными (непустые):")
            for col_idx in range(min(15, df.shape[1])):
                non_empty = df.iloc[:, col_idx].dropna()
                if not non_empty.empty:
                    print(f"Колонка {col_idx}: {len(non_empty)} непустых значений")
                    print(f"  Первые значения: {list(non_empty.head(5).values)}")
        
        except Exception as e:
            print(f"❌ Ошибка pandas: {e}")
        
        print("\n" + "=" * 40)
        print("📝 РЕКОМЕНДАЦИИ ПО ОБРАБОТКЕ:")
        print("=" * 40)
        
        if potential_headers:
            header_row = potential_headers[0][0]
            print(f"✅ Найдена строка с заголовками: {header_row}")
            print(f"✅ Данные товаров начинаются со строки: {header_row + 1}")
            print(f"✅ Заголовки: {potential_headers[0][1]}")
        else:
            print("❌ Не удалось автоматически найти заголовки")
            print("ℹ️  Требуется ручной анализ структуры")
        
    except Exception as e:
        print(f"❌ Общая ошибка анализа: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_stock_file_structure()