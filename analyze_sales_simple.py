#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Простой анализ файла продаж без pandas
"""

from openpyxl import load_workbook
import re

def analyze_sales_file(file_path):
    """Детальный анализ файла продаж"""
    
    print(f"\n{'='*60}")
    print(f"АНАЛИЗ ФАЙЛА: {file_path}")
    print(f"{'='*60}\n")
    
    # Загружаем файл
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    print(f"📊 Размер: {ws.max_row} строк x {ws.max_column} колонок\n")
    
    # Показываем первые 15 строк
    print("🔍 ПЕРВЫЕ 15 СТРОК:")
    print("-" * 80)
    for row_idx in range(1, min(16, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(6, ws.max_column + 1)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_data.append(str(cell_value)[:20])
            else:
                row_data.append("---")
        print(f"Строка {row_idx:2d}: {' | '.join(row_data)}")
    
    # Ищем период в строке 4
    print(f"\n📅 СТРОКА 4 (период):")
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=4, column=col_idx).value
        if cell_value and 'период' in str(cell_value).lower():
            print(f"   Найдено: {cell_value}")
    
    # Анализируем строку 9 (заголовки)
    print(f"\n📋 СТРОКА 9 (заголовки):")
    headers = {}
    for col_idx in range(1, min(10, ws.max_column + 1)):
        header = ws.cell(row=9, column=col_idx).value
        if header:
            headers[col_idx] = str(header).strip()
            print(f"   Колонка {col_idx}: {header}")
    
    # Ищем колонку с выручкой
    revenue_col_idx = None
    for col_idx, header in headers.items():
        if 'выручка' in header.lower():
            revenue_col_idx = col_idx
            print(f"\n✅ Колонка 'Выручка' найдена: колонка {col_idx}")
            break
    
    if not revenue_col_idx:
        print("\n❌ Колонка 'Выручка' не найдена!")
        return None, 0
    
    # Анализируем выручку
    print(f"\n💰 АНАЛИЗ ВЫРУЧКИ (начиная со строки 10):")
    print("-" * 80)
    
    total_revenue = 0.0
    product_count = 0
    category_count = 0
    examples_shown = 0
    
    for row_idx in range(10, ws.max_row + 1):
        # Получаем номенклатуру (первая колонка)
        nomenclature = ws.cell(row=row_idx, column=1).value
        
        if not nomenclature:
            continue
        
        nomenclature_str = str(nomenclature)
        
        # Определяем тип строки
        # Категории обычно без отступов и без специфических признаков
        is_category = (not nomenclature_str.startswith(' ') and 
                      not nomenclature_str.startswith('\t') and
                      len(nomenclature_str) < 40 and
                      not any(sign in nomenclature_str.lower() for sign in ['*', '×', 'мм', 'см', '№']))
        
        # Получаем выручку
        revenue_value = ws.cell(row=row_idx, column=revenue_col_idx).value
        
        if revenue_value and isinstance(revenue_value, (int, float)):
            if is_category:
                category_count += 1
                if examples_shown < 3:
                    print(f"  КАТЕГОРИЯ: {nomenclature_str[:40]:<40} | Выручка: {revenue_value:>15,.2f}")
                    examples_shown += 1
            else:
                # Это товар
                total_revenue += revenue_value
                product_count += 1
                
                if product_count <= 5:
                    print(f"  Товар {product_count:3d}: {nomenclature_str[:40]:<40} | Выручка: {revenue_value:>15,.2f}")
    
    print(f"\n📊 РЕЗУЛЬТАТЫ:")
    print(f"   Найдено категорий: {category_count}")
    print(f"   Найдено товаров: {product_count}")
    print(f"   Общая выручка по товарам: {total_revenue:,.2f}")
    
    # Ищем итоговую строку
    print(f"\n🔍 ПОИСК ИТОГА В ФАЙЛЕ:")
    for row_idx in range(ws.max_row - 10, ws.max_row + 1):
        if row_idx > 0:
            nomenclature = ws.cell(row=row_idx, column=1).value
            if nomenclature and ('итого' in str(nomenclature).lower() or 'всего' in str(nomenclature).lower()):
                total_in_file = ws.cell(row=row_idx, column=revenue_col_idx).value
                print(f"   Найдена строка: '{nomenclature}'")
                print(f"   Итоговая выручка в файле: {total_in_file:,.2f}" if total_in_file else "   Значение не найдено")
                break
    
    wb.close()
    
    return total_revenue, product_count

# Анализируем файл
if __name__ == "__main__":
    file_path = "6_Склад_фурнитуры_Овощная_база_Магазин_продажи_01_07_2024_01_07.xlsx"
    
    try:
        total_revenue, product_count = analyze_sales_file(file_path)
        
        print(f"\n{'='*60}")
        print(f"ИТОГ: Выручка = {total_revenue:,.2f}, Товаров = {product_count}")
        print(f"По вашим данным должно быть: 27,080,953.00")
        print(f"{'='*60}")
        
    except FileNotFoundError:
        print(f"\n❌ Файл не найден: {file_path}")
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()