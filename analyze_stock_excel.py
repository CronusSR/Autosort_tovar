#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Анализ структуры Excel файла остатков для создания требований к JSON формату
"""

import sys
import os

def analyze_stock_excel_simple():
    """Простой анализ Excel файла остатков без pandas"""
    
    print("=== Анализ структуры файла остатков ===")
    
    filename = 'остатки на 08.07.2025.xlsx'
    
    if not os.path.exists(filename):
        print(f"❌ Файл {filename} не найден")
        return
    
    try:
        # Пробуем использовать openpyxl для чтения
        try:
            from openpyxl import load_workbook
            
            print(f"📊 Анализ файла: {filename}")
            
            # Открываем файл
            wb = load_workbook(filename, read_only=True)
            print(f"✅ Листы в файле: {wb.sheetnames}")
            
            # Берем первый лист
            ws = wb.active
            print(f"✅ Активный лист: {ws.title}")
            
            # Анализируем размеры
            print(f"✅ Размеры листа: {ws.max_row} строк, {ws.max_column} столбцов")
            
            # Читаем заголовки (первую строку)
            headers = []
            for col in range(1, min(ws.max_column + 1, 20)):  # Максимум 20 столбцов
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"Column_{col}")
            
            print(f"✅ Заголовки столбцов:")
            for i, header in enumerate(headers[:15]):  # Показываем первые 15
                print(f"   {i+1}. {header}")
            
            if len(headers) > 15:
                print(f"   ... и еще {len(headers) - 15} столбцов")
            
            # Анализируем первые несколько строк данных
            print(f"\n✅ Первые 3 строки данных:")
            for row in range(2, min(6, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(len(headers) + 1, 6)):  # Первые 5 столбцов
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data.append(str(cell_value)[:30])  # Обрезаем длинные значения
                    else:
                        row_data.append("—")
                print(f"   Строка {row}: {' | '.join(row_data)}")
            
            # Ищем столбец с наименованиями товаров
            print(f"\n✅ Поиск столбца с номенклатурой:")
            nomenclature_col = None
            for i, header in enumerate(headers):
                if any(keyword in header.lower() for keyword in ['наименование', 'номенклатура', 'товар']):
                    nomenclature_col = i + 1
                    print(f"   Найден столбец номенклатуры: {header} (колонка {nomenclature_col})")
                    break
            
            if not nomenclature_col:
                print("   ❌ Столбец номенклатуры не найден")
            
            # Анализируем столбцы со складами/остатками
            print(f"\n✅ Потенциальные столбцы складов:")
            warehouse_cols = []
            for i, header in enumerate(headers):
                # Пропускаем первый столбец (обычно номенклатура)
                if i == 0:
                    continue
                
                # Проверяем есть ли числовые данные в столбце
                has_numbers = False
                for row in range(2, min(10, ws.max_row + 1)):
                    cell_value = ws.cell(row=row, column=i + 1).value
                    if isinstance(cell_value, (int, float)) and cell_value != 0:
                        has_numbers = True
                        break
                
                if has_numbers:
                    warehouse_cols.append((i + 1, header))
                    print(f"   {len(warehouse_cols)}. {header} (колонка {i + 1})")
            
            print(f"\n✅ Найдено {len(warehouse_cols)} столбцов со складами")
            
            # Показываем примеры данных для первых складов
            if warehouse_cols and nomenclature_col:
                print(f"\n✅ Примеры остатков:")
                for row in range(2, min(7, ws.max_row + 1)):
                    product_name = ws.cell(row=row, column=nomenclature_col).value
                    if product_name:
                        print(f"   📦 {str(product_name)[:50]}...")
                        for col_num, col_name in warehouse_cols[:5]:  # Первые 5 складов
                            stock_value = ws.cell(row=row, column=col_num).value
                            if isinstance(stock_value, (int, float)) and stock_value > 0:
                                print(f"      - {col_name}: {stock_value}")
                        print()
            
            wb.close()
            
        except ImportError:
            print("❌ openpyxl не установлен, используем альтернативный метод")
            return analyze_with_basic_tools(filename)
            
    except Exception as e:
        print(f"❌ Ошибка анализа файла: {e}")
        import traceback
        traceback.print_exc()

def analyze_with_basic_tools(filename):
    """Альтернативный анализ без специальных библиотек"""
    
    print("📊 Альтернативный анализ файла...")
    
    # Получаем базовую информацию о файле
    file_size = os.path.getsize(filename)
    print(f"✅ Размер файла: {file_size} байт ({file_size/1024:.1f} KB)")
    
    print("""
    📋 Предполагаемая структура на основе типичных файлов остатков:
    
    Столбец 1: Наименование товара / Номенклатура
    Столбцы 2-N: Остатки по складам
    
    Типичные названия складов:
    - База Склад Фурнитура Комплект
    - Казыбаева Склад Фурнитура TRADE
    - склад фурнитура № 1
    - ТД Казыбаева ФУРНИТУРА магазин
    - Магазин фурнитуры
    - и другие...
    """)

if __name__ == "__main__":
    analyze_stock_excel_simple()