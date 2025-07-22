#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import os
import sys

def analyze_stock_file_simple():
    """Простой анализ структуры файла остатков без pandas"""
    
    file_path = "/mnt/f/Работа-Никита/Autosort_tovar/остатки на 08.07.2025.xlsx"
    
    if not os.path.exists(file_path):
        print(f"❌ Файл не найден: {file_path}")
        return
    
    print(f"📁 Анализ файла: остатки на 08.07.2025.xlsx")
    print("=" * 80)
    
    try:
        # Загружаем файл
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"📊 Листы в файле: {wb.sheetnames}")
        
        # Работаем с первым листом
        ws = wb.active
        sheet_name = wb.active.title
        print(f"📋 Активный лист: {sheet_name}")
        print(f"📏 Размеры листа: {ws.max_row} строк x {ws.max_column} колонок")
        
        print("\n" + "=" * 60)
        print("🔍 ДЕТАЛЬНЫЙ АНАЛИЗ ПЕРВЫХ 25 СТРОК:")
        print("=" * 60)
        
        # Читаем и анализируем первые строки
        warehouse_columns = []
        nomenclature_column = None
        header_row = None
        
        for row_idx in range(1, min(26, ws.max_row + 1)):
            print(f"\n--- СТРОКА {row_idx} ---")
            
            row_data = []
            for col_idx in range(1, min(21, ws.max_column + 1)):  # Первые 20 колонок
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                
                if cell_value is None:
                    cell_value = ""
                else:
                    cell_value = str(cell_value).strip()
                
                row_data.append(cell_value)
            
            # Выводим строку
            for i, value in enumerate(row_data, 1):
                if value:  # Показываем только непустые ячейки
                    col_letter = openpyxl.utils.get_column_letter(i)
                    print(f"  {col_letter}{row_idx}: '{value}'")
            
            # Анализируем содержимое строки на предмет заголовков
            row_text = " ".join(row_data).lower()
            keywords = ["номенклатура", "наименование", "склад", "остаток", "количество", "товар"]
            
            if any(keyword in row_text for keyword in keywords):
                print(f"  ⭐ ПОТЕНЦИАЛЬНАЯ СТРОКА ЗАГОЛОВКОВ!")
                header_row = row_idx
                
                # Ищем колонку с номенклатурой
                for i, value in enumerate(row_data, 1):
                    value_lower = value.lower()
                    if "номенклатура" in value_lower or "наименование" in value_lower:
                        nomenclature_column = i
                        print(f"  📝 Номенклатура в колонке {openpyxl.utils.get_column_letter(i)} (индекс {i})")
                    
                    # Ищем склады
                    if "склад" in value_lower and value_lower != "склад":
                        warehouse_columns.append((i, value))
                        print(f"  🏪 Склад найден в колонке {openpyxl.utils.get_column_letter(i)} (индекс {i}): '{value}'")
        
        print("\n" + "=" * 60)
        print("📊 СВОДКА АНАЛИЗА:")
        print("=" * 60)
        
        if header_row:
            print(f"✅ Строка заголовков: {header_row}")
            print(f"✅ Строка начала данных: {header_row + 1}")
        else:
            print("❌ Строка заголовков не найдена")
        
        if nomenclature_column:
            print(f"✅ Колонка номенклатуры: {openpyxl.utils.get_column_letter(nomenclature_column)} (индекс {nomenclature_column})")
        else:
            print("❌ Колонка номенклатуры не найдена")
        
        if warehouse_columns:
            print(f"✅ Найдено складов: {len(warehouse_columns)}")
            for col_idx, warehouse_name in warehouse_columns:
                print(f"   - {warehouse_name}: колонка {openpyxl.utils.get_column_letter(col_idx)} (индекс {col_idx})")
        else:
            print("❌ Склады не найдены")
        
        # Дополнительный анализ: поиск числовых данных
        print("\n" + "=" * 60)
        print("🔢 АНАЛИЗ ЧИСЛОВЫХ ДАННЫХ:")
        print("=" * 60)
        
        if header_row:
            data_start_row = header_row + 1
            print(f"Анализ данных начиная со строки {data_start_row}...")
            
            # Проверяем первые 5 строк данных
            for row_idx in range(data_start_row, min(data_start_row + 5, ws.max_row + 1)):
                print(f"\nСтрока данных {row_idx}:")
                
                # Показываем номенклатуру
                if nomenclature_column:
                    nomenclature = ws.cell(row=row_idx, column=nomenclature_column).value
                    print(f"  Номенклатура: '{nomenclature}'")
                
                # Показываем остатки по складам
                for col_idx, warehouse_name in warehouse_columns:
                    stock_value = ws.cell(row=row_idx, column=col_idx).value
                    print(f"  {warehouse_name}: {stock_value}")
        
        print("\n" + "=" * 60)
        print("🔧 РЕКОМЕНДАЦИИ ДЛЯ ИСПРАВЛЕНИЯ КОДА:")
        print("=" * 60)
        
        if header_row and nomenclature_column and warehouse_columns:
            print("✅ Структура файла понятна, можно исправить код!")
            print(f"📌 Использовать header_row = {header_row}")
            print(f"📌 Использовать nomenclature_column = {nomenclature_column - 1} (для pandas индексации)")
            print(f"📌 Склады:")
            for col_idx, warehouse_name in warehouse_columns:
                print(f"   '{warehouse_name}': колонка {col_idx - 1} (для pandas)")
        else:
            print("❌ Требуется дополнительный анализ структуры")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Ошибка анализа: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_stock_file_simple()