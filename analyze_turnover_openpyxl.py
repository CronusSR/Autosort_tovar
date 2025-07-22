#!/usr/bin/env python3
"""
Анализатор файла оборачиваемости с использованием openpyxl
"""

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import sys

def analyze_turnover_file():
    """Анализирует файл оборачиваемости и структуру ABC анализа"""
    
    file_path = "ОБОРАЧИВАЕМОСТЬ 10.07.2025.xlsx"
    
    try:
        # Открываем файл
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"📊 Листы в файле: {wb.sheetnames}")
        print("=" * 60)
        
        # Анализируем лист ОСТАТКИ
        if 'ОСТАТКИ' in wb.sheetnames:
            print("\n🔍 АНАЛИЗ ЛИСТА 'ОСТАТКИ':")
            ws_ostatok = wb['ОСТАТКИ']
            
            print(f"Размер листа: {ws_ostatok.max_row} строк, {ws_ostatok.max_column} колонок")
            
            # Анализируем заголовки
            print("\n📋 Заголовки колонок:")
            headers = []
            for col in range(1, min(ws_ostatok.max_column + 1, 31)):  # До колонки AE
                header = ws_ostatok.cell(row=1, column=col).value
                headers.append(f"{get_column_letter(col)}: {header}")
            
            for i, header in enumerate(headers):
                print(f"  {header}")
                if i > 0 and (i + 1) % 5 == 0:  # Разбиваем на группы по 5
                    print()
            
            # Анализируем конкретные колонки из формулы
            print(f"\n🔍 АНАЛИЗ КОЛОНОК ИЗ ФОРМУЛЫ SUMIFS:")
            
            # Колонка AD (30)
            ad_col = 30
            print(f"\nКолонка AD ({ad_col}):")
            print(f"  Заголовок: {ws_ostatok.cell(row=1, column=ad_col).value}")
            print("  Примеры значений:")
            for row in range(2, min(7, ws_ostatok.max_row + 1)):
                value = ws_ostatok.cell(row=row, column=ad_col).value
                print(f"    Строка {row}: {value}")
            
            # Колонка Q (17)
            q_col = 17
            print(f"\nКолонка Q ({q_col}):")
            print(f"  Заголовок: {ws_ostatok.cell(row=1, column=q_col).value}")
            print("  Примеры значений:")
            unique_q_values = set()
            for row in range(2, min(12, ws_ostatok.max_row + 1)):
                value = ws_ostatok.cell(row=row, column=q_col).value
                unique_q_values.add(value)
                if row <= 6:
                    print(f"    Строка {row}: {value}")
            print(f"  Уникальные значения (первые 10): {list(unique_q_values)[:10]}")
            
            # Колонка S (19)
            s_col = 19
            print(f"\nКолонка S ({s_col}):")
            print(f"  Заголовок: {ws_ostatok.cell(row=1, column=s_col).value}")
            print("  Примеры значений:")
            unique_s_values = set()
            for row in range(2, min(12, ws_ostatok.max_row + 1)):
                value = ws_ostatok.cell(row=row, column=s_col).value
                unique_s_values.add(value)
                if row <= 6:
                    print(f"    Строка {row}: {value}")
            print(f"  Уникальные значения (первые 10): {list(unique_s_values)[:10]}")
        
        # Анализируем лист ABC ПО СКЛАДАМ
        if 'ABC ПО СКЛАДАМ' in wb.sheetnames:
            print("\n\n🔍 АНАЛИЗ ЛИСТА 'ABC ПО СКЛАДАМ':")
            ws_abc = wb['ABC ПО СКЛАДАМ']
            
            print(f"Размер листа: {ws_abc.max_row} строк, {ws_abc.max_column} колонок")
            
            # Анализируем заголовки
            print("\n📋 Заголовки и структура:")
            for col in range(1, min(ws_abc.max_column + 1, 11)):
                header = ws_abc.cell(row=1, column=col).value
                print(f"  {get_column_letter(col)}: {header}")
            
            # Анализируем первые строки данных
            print(f"\n📋 Первые 10 строк данных:")
            for row in range(1, min(11, ws_abc.max_row + 1)):
                row_data = []
                for col in range(1, min(ws_abc.max_column + 1, 6)):
                    value = ws_abc.cell(row=row, column=col).value
                    row_data.append(str(value) if value is not None else "")
                print(f"  Строка {row}: {' | '.join(row_data)}")
            
            # Ищем категорию "Клей"
            print(f"\n🔍 ПОИСК КАТЕГОРИИ 'Клей':")
            found_klei = False
            for row in range(1, min(ws_abc.max_row + 1, 50)):
                for col in range(1, min(ws_abc.max_column + 1, 4)):
                    value = ws_abc.cell(row=row, column=col).value
                    if value and "Клей" in str(value):
                        print(f"  Найдено в {get_column_letter(col)}{row}: {value}")
                        # Выводим всю строку
                        row_data = []
                        for c in range(1, min(ws_abc.max_column + 1, 6)):
                            v = ws_abc.cell(row=row, column=c).value
                            row_data.append(str(v) if v is not None else "")
                        print(f"    Полная строка: {' | '.join(row_data)}")
                        found_klei = True
            
            if not found_klei:
                print("  Категория 'Клей' не найдена в первых 50 строках")
            
            # Проверяем значение в C3
            c3_value = ws_abc.cell(row=3, column=3).value
            print(f"\n📍 Значение C3: {c3_value}")
            
        # Проверяем наличие формул в ABC листе
        if 'ABC ПО СКЛАДАМ' in wb.sheetnames:
            print(f"\n🧮 ПОИСК ФОРМУЛ В ЛИСТЕ 'ABC ПО СКЛАДАМ':")
            ws_abc = wb['ABC ПО СКЛАДАМ']
            
            # Перезагружаем без data_only для получения формул
            wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
            ws_abc_formulas = wb_formulas['ABC ПО СКЛАДАМ']
            
            found_formulas = 0
            for row in range(1, min(ws_abc_formulas.max_row + 1, 20)):
                for col in range(1, min(ws_abc_formulas.max_column + 1, 10)):
                    cell = ws_abc_formulas.cell(row=row, column=col)
                    if cell.value and str(cell.value).startswith('='):
                        print(f"  {get_column_letter(col)}{row}: {cell.value}")
                        found_formulas += 1
                        if found_formulas >= 5:  # Ограничиваем вывод
                            break
                if found_formulas >= 5:
                    break
            
            if found_formulas == 0:
                print("  Формулы не найдены в первых 20 строках и 10 колонках")
        
    except FileNotFoundError:
        print(f"❌ Файл {file_path} не найден!")
        return
    except Exception as e:
        print(f"❌ Ошибка при анализе файла: {e}")
        import traceback
        traceback.print_exc()
        return
    
    print("\n" + "=" * 60)
    print("✅ Анализ завершен!")

if __name__ == "__main__":
    analyze_turnover_file()