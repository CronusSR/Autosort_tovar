# complete_warehouse_restore.py
"""
🎯 ПОЛНОЕ ВОССТАНОВЛЕНИЕ АНАЛИЗА СКЛАДОВ
Восстанавливает все функции: цены, детальный анализ по складам, полный показ товаров
"""

import pandas as pd
import streamlit as st
import numpy as np
from typing import Dict, List, Any, Optional


class DetailedWarehouseAnalyzer:
    """
    Восстанавливает детальный анализ складов с ценами и полной функциональностью
    """
    
    def __init__(self):
        # Точная конфигурация ваших складов из структуры файла
        self.warehouse_config = {
            'Шымкент_Склад': {
                'name': '4 Склад фурнитуры АЗМ Шымкент "Овощная база"',
                'short_name': 'Шымкент Склад',
                'city': 'шымкент',
                'type': 'warehouse',
                'col': 3,  # Колонка D (индекс 3)
                'min_days': 15,
                'max_days': 45
            },
            'Шымкент_Магазин': {
                'name': '6 Склад фурнитуры "Овощная база" Магазин',
                'short_name': 'Шымкент Магазин',
                'city': 'шымкент',
                'type': 'store',
                'col': 4,  # Колонка E (индекс 4)
                'min_days': 10,
                'max_days': 30
            },
            'Алматы_Склад': {
                'name': 'АО Склад Фурнитура TRADE',
                'short_name': 'Алматы Склад',
                'city': 'алматы',
                'type': 'specialized_store',
                'col': 6,  # Колонка G (индекс 6)
                'min_days': 20,
                'max_days': 60
            },
            'База_Комплект': {
                'name': 'База Склад Фурнитура Комплект',
                'short_name': 'База Комплект',
                'city': 'алматы',
                'type': 'warehouse',
                'col': 8,  # Колонка I (индекс 8)
                'min_days': 25,
                'max_days': 75
            },
            'Барыс_Склад': {
                'name': 'Барыс Склад Фурнитура TRADE',
                'short_name': 'Барыс Склад',
                'city': 'алматы',
                'type': 'trade_store',
                'col': 9,  # Колонка J (индекс 9)
                'min_days': 15,
                'max_days': 40
            },
            'Казыбаева_Склад': {
                'name': 'Казыбаева Склад Фурнитура TRADE',
                'short_name': 'Казыбаева Склад',
                'city': 'алматы',
                'type': 'trade_store',
                'col': 10,  # Колонка K (индекс 10)
                'min_days': 12,
                'max_days': 35
            },
            'Астана_Магазин': {
                'name': 'Магазин фурнитуры',
                'short_name': 'Астана Магазин',
                'city': 'астана',
                'type': 'store',
                'col': 11,  # Колонка L (индекс 11)
                'min_days': 10,
                'max_days': 30
            },
            'Астана_Склад': {
                'name': 'склад фурнитура № 1',
                'short_name': 'Астана Склад',
                'city': 'астана',
                'type': 'warehouse',
                'col': 12,  # Колонка M (индекс 12)
                'min_days': 15,
                'max_days': 45
            },
            'Казыбаева_Магазин': {
                'name': 'ТД Казыбаева ФУРНИТУРА магазин',
                'short_name': 'Казыбаева Магазин',
                'city': 'алматы',
                'type': 'retail_store',
                'col': 13,  # Колонка N (индекс 13)
                'min_days': 8,
                'max_days': 25
            }
        }
    
    def read_remains_file_with_exact_structure(self, uploaded_file):
        """
        Читает файл остатков с точной структурой из анализа
        """
        try:
            st.info("📖 Читаем файл остатков с известной структурой...")
            
            # Читаем файл
            if uploaded_file.name.endswith('.xlsx'):
                file_data = pd.read_excel(uploaded_file, header=None).values.tolist()
            else:
                file_data = pd.read_excel(uploaded_file, engine='xlrd', header=None).values.tolist()
            
            st.success(f"✅ Файл прочитан. Всего строк: {len(file_data)}")
            
            # Проверяем что файл достаточно большой
            if len(file_data) < 10:
                raise ValueError("Файл слишком мал. Должно быть минимум 10 строк.")
            
            # Читаем данные начиная с 4й строки (товары с 4й строки, индекс 3)
            remains_data = []
            processed_items = 0
            
            for i in range(3, len(file_data)):  # начинаем с индекса 3 (4я строка Excel)
                row = file_data[i]
                
                # Проверяем что строка не пустая
                if not row or len(row) == 0:
                    continue
                    
                # Проверяем что первая ячейка (номенклатура) не пустая
                if not row[0] or pd.isna(row[0]):
                    continue
                    
                item_name = str(row[0]).strip()
                if not item_name or item_name.lower() in ['', 'nan', 'none']:
                    continue
                
                # Безопасно получаем итоговый остаток (колонка 15 = индекс 14)
                try:
                    total_stock = row[14] if len(row) > 14 and row[14] is not None else 0
                    total_stock = float(total_stock) if pd.notna(total_stock) else 0
                except (ValueError, TypeError, IndexError):
                    total_stock = 0
                
                item_data = {
                    'номенклатура': item_name,
                    'итого_остаток': total_stock
                }
                
                # Добавляем остатки по складам с точными колонками
                for warehouse_key, config in self.warehouse_config.items():
                    col_idx = config['col']
                    try:
                        if len(row) > col_idx and row[col_idx] is not None:
                            quantity = float(row[col_idx]) if pd.notna(row[col_idx]) else 0
                        else:
                            quantity = 0
                    except (ValueError, TypeError, IndexError):
                        quantity = 0
                    
                    item_data[f'{warehouse_key}_остаток'] = quantity
                
                remains_data.append(item_data)
                processed_items += 1
            
            st.success(f"✅ Обработано товаров: {processed_items}")
            
            if not remains_data:
                raise ValueError("Не найдено ни одного товара с данными. Проверьте структуру файла.")
            
            # Создаем DataFrame
            remains_df = pd.DataFrame(remains_data)
            
            # Показываем превью
            with st.expander("👀 Превью данных остатков"):
                st.dataframe(remains_df.head(), use_container_width=True)
                
                # Статистика по складам
                st.write("📊 **Статистика по складам:**")
                for warehouse_key, config in self.warehouse_config.items():
                    col_name = f"{warehouse_key}_остаток"
                    if col_name in remains_df.columns:
                        total_stock = remains_df[col_name].sum()
                        items_with_stock = (remains_df[col_name] > 0).sum()
                        st.write(f"  - **{config['short_name']}**: {total_stock:,.0f} (товаров: {items_with_stock})")
            
            return remains_df
            
        except Exception as e:
            st.error(f"❌ Ошибка чтения файла: {str(e)}")
            return pd.DataFrame()
    
    def analyze_warehouse_stock_detailed(self, remains_df, ads_data, store_ads_by_city, min_days=10, max_days=50):
        """
        ПОЛНЫЙ детальный анализ складов с ценами из ADS
        """
        if remains_df is None or remains_df.empty:
            st.error("❌ Нет данных остатков")
            return []
        
        if 'номенклатура' not in remains_df.columns:
            st.error("❌ В файле остатков нет колонки 'номенклатура'")
            return []
        
        st.info(f"🔄 Начинаем ПОЛНЫЙ анализ {len(remains_df)} товаров по {len(self.warehouse_config)} складам...")
        
        # Проверяем наличие ADS данных и цен
        has_ads = ads_data is not None and not ads_data.empty
        has_prices = False
        price_column = None
        
        if has_ads:
            st.success(f"✅ ADS данные доступны: {len(ads_data)} товаров")
            
            # Ищем ценовые колонки
            price_columns = ['last_purchase_price', 'цена', 'price', 'стоимость', 'закупочная_цена']
            for col in price_columns:
                if col in ads_data.columns:
                    has_prices = True
                    price_column = col
                    items_with_prices = (ads_data[col] > 0).sum()
                    avg_price = ads_data[ads_data[col] > 0][col].mean()
                    st.success(f"💰 Цены найдены в колонке '{col}': {items_with_prices} товаров, средняя цена {avg_price:.2f} ₽")
                    break
            
            if not has_prices:
                st.warning("⚠️ Цены в ADS данных не найдены")
        else:
            st.warning("⚠️ ADS данные отсутствуют - выполняем базовый анализ остатков")
        
        analysis_results = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Анализируем каждый товар ПОЛНОСТЬЮ
        for idx, (_, item) in enumerate(remains_df.iterrows()):
            
            # Обновляем прогресс
            progress = (idx + 1) / len(remains_df)
            progress_bar.progress(progress)
            status_text.text(f"Анализируем товар {idx + 1}/{len(remains_df)}: {item['номенклатура'][:30]}...")
            
            item_name = str(item['номенклатура']).strip()
            total_stock = float(item.get('итого_остаток', 0))
            
            # Получаем ADS и цену для товара
            ads_value = 0
            item_price = 0
            
            if has_ads:
                ads_match = ads_data[ads_data['номенклатура'] == item_name]
                if not ads_match.empty:
                    ads_value = float(ads_match.iloc[0].get('ads', 0))
                    if has_prices and price_column:
                        try:
                            item_price = float(ads_match.iloc[0].get(price_column, 0))
                        except (ValueError, TypeError):
                            item_price = 0
            
            # ДЕТАЛЬНЫЙ анализ по каждому складу
            warehouses_analysis = {}
            overall_critical_count = 0
            overall_warning_count = 0
            total_order_quantity = 0
            total_order_value = 0
            
            for warehouse_key, config in self.warehouse_config.items():
                stock_col = f"{warehouse_key}_остаток"
                current_stock = float(item.get(stock_col, 0))
                
                # Персональные настройки склада
                wh_min_days = config.get('min_days', min_days)
                wh_max_days = config.get('max_days', max_days)
                
                # Расчеты MIN/MAX запасов
                min_stock = ads_value * wh_min_days if ads_value > 0 else 0
                max_stock = ads_value * wh_max_days if ads_value > 0 else 0
                
                # Дефицит/избыток
                min_deficit = max(0, min_stock - current_stock)
                max_deficit = max(0, max_stock - current_stock)
                surplus = max(0, current_stock - max_stock)
                
                # Месяцы запаса
                if ads_value > 0:
                    days_of_stock = current_stock / ads_value
                    months_of_stock = days_of_stock / 30
                elif current_stock > 0:
                    days_of_stock = 999
                    months_of_stock = 999  # Бесконечно (нет продаж)
                else:
                    days_of_stock = 0
                    months_of_stock = 0
                
                # Определяем статус склада
                if ads_value > 0:
                    if current_stock < min_stock:
                        if min_deficit > ads_value * 7:  # Более недели дефицита
                            status = 'critical'
                            overall_critical_count += 1
                        else:
                            status = 'warning'
                            overall_warning_count += 1
                    elif current_stock > max_stock:
                        status = 'excess'
                    else:
                        status = 'good'
                elif current_stock > 0:
                    status = 'no_sales'
                else:
                    status = 'empty'
                
                # Количество к заказу
                order_quantity = min_deficit if status in ['critical', 'warning'] else 0
                total_order_quantity += order_quantity
                
                # Денежные расчеты
                stock_value = current_stock * item_price if item_price > 0 else 0
                order_value = order_quantity * item_price if item_price > 0 else 0
                total_order_value += order_value
                
                # Рекомендация
                if status == 'critical':
                    recommendation = f"🔴 КРИТИЧНО! Заказать {order_quantity:.0f} шт на сумму {order_value:.0f} ₽"
                elif status == 'warning':
                    recommendation = f"🟡 Заказать {order_quantity:.0f} шт на сумму {order_value:.0f} ₽"
                elif status == 'excess':
                    recommendation = f"🔵 Избыток {surplus:.0f} шт (стоимость {surplus * item_price:.0f} ₽)"
                elif status == 'no_sales':
                    recommendation = f"⚪ Нет продаж, остаток {current_stock:.0f} шт (стоимость {stock_value:.0f} ₽)"
                elif status == 'empty':
                    recommendation = "⚫ Нет остатков и продаж"
                else:
                    recommendation = f"🟢 В норме: {current_stock:.0f} шт (стоимость {stock_value:.0f} ₽)"
                
                warehouses_analysis[warehouse_key] = {
                    'warehouse_name': config['name'],
                    'short_name': config['short_name'],
                    'city': config['city'],
                    'type': config['type'],
                    'current_stock': current_stock,
                    'min_stock': min_stock,
                    'max_stock': max_stock,
                    'min_deficit': min_deficit,
                    'max_deficit': max_deficit,
                    'surplus': surplus,
                    'days_of_stock': days_of_stock,
                    'months_of_stock': months_of_stock,
                    'status': status,
                    'order_quantity': order_quantity,
                    'stock_value': stock_value,
                    'order_value': order_value,
                    'recommendation': recommendation,
                    'settings': f"{wh_min_days}-{wh_max_days} дней"
                }
            
            # Общий статус товара
            if overall_critical_count > 0:
                overall_status = 'critical'
            elif overall_warning_count > 0:
                overall_status = 'warning'
            else:
                overall_status = 'good'
            
            # Минимальные месяцы запаса среди всех складов
            months_list = [w['months_of_stock'] for w in warehouses_analysis.values() 
                          if w['months_of_stock'] < 999]
            min_months_across_warehouses = min(months_list) if months_list else 0
            
            # Итоговая запись товара
            analysis_results.append({
                'номенклатура': item_name,
                'total_stock': total_stock,
                'ads': ads_value,
                'price': item_price,
                'total_stock_value': total_stock * item_price if item_price > 0 else 0,
                'total_order_quantity': total_order_quantity,
                'total_order_value': total_order_value,
                'min_months_across_warehouses': min_months_across_warehouses,
                'overall_status': overall_status,
                'critical_warehouses_count': overall_critical_count,
                'warning_warehouses_count': overall_warning_count,
                'warehouses': warehouses_analysis,
                'analysis_timestamp': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                'parameters': {
                    'global_min_days': min_days,
                    'global_max_days': max_days,
                    'has_ads': has_ads,
                    'has_prices': has_prices,
                    'price_column': price_column
                }
            })
        
        # Завершаем прогресс
        progress_bar.progress(1.0)
        status_text.text("✅ Анализ завершен!")
        
        # Статистика анализа
        total_items = len(analysis_results)
        critical_items = sum(1 for item in analysis_results if item['overall_status'] == 'critical')
        warning_items = sum(1 for item in analysis_results if item['overall_status'] == 'warning')
        
        total_stock_value = sum(item['total_stock_value'] for item in analysis_results)
        total_order_value = sum(item['total_order_value'] for item in analysis_results)
        
        st.success(f"""
        ✅ **ПОЛНЫЙ АНАЛИЗ ЗАВЕРШЕН!**
        
        📊 **Общая статистика:**
        - Всего товаров проанализировано: {total_items}
        - 🔴 Критичных: {critical_items} ({critical_items/total_items*100:.1f}%)
        - 🟡 Требуют внимания: {warning_items} ({warning_items/total_items*100:.1f}%)
        - 🟢 В норме: {total_items - critical_items - warning_items} ({(total_items - critical_items - warning_items)/total_items*100:.1f}%)
        
        💰 **Финансовая статистика:**
        - Общая стоимость остатков: {total_stock_value:,.0f} ₽
        - Общая стоимость к заказу: {total_order_value:,.0f} ₽
        - Покрытие ценами: {sum(1 for item in analysis_results if item['price'] > 0)/total_items*100:.1f}%
        
        🏪 **Анализ складов:**
        - Количество складов: {len(self.warehouse_config)}
        - Персональные настройки применены
        - Все товары проанализированы по всем складам
        """)
        
        return analysis_results
    
    def get_warehouse_recommendations(self, analysis_results=None):
        """
        Получает ДЕТАЛЬНЫЕ рекомендации по складам
        """
        if not analysis_results:
            return {}
        
        warehouse_summary = {}
        
        # Анализируем каждый склад
        for warehouse_key, config in self.warehouse_config.items():
            
            summary = {
                'warehouse_name': config['name'],
                'short_name': config['short_name'],
                'city': config['city'],
                'type': config['type'],
                'settings': f"{config['min_days']}-{config['max_days']} дней",
                'total_items': 0,
                'items_with_stock': 0,
                'critical_items': 0,
                'warning_items': 0,
                'excess_items': 0,
                'no_sales_items': 0,
                'empty_items': 0,
                'good_items': 0,
                'total_stock_quantity': 0,
                'total_stock_value': 0,
                'total_order_quantity': 0,
                'total_order_value': 0,
                'average_months_stock': 0,
                'total_deficit_value': 0,
                'total_surplus_value': 0
            }
            
            # Собираем статистику
            months_list = []
            
            for item in analysis_results:
                if warehouse_key in item['warehouses']:
                    wh_data = item['warehouses'][warehouse_key]
                    
                    summary['total_items'] += 1
                    
                    if wh_data['current_stock'] > 0:
                        summary['items_with_stock'] += 1
                    
                    status = wh_data['status']
                    if status == 'critical':
                        summary['critical_items'] += 1
                    elif status == 'warning':
                        summary['warning_items'] += 1
                    elif status == 'excess':
                        summary['excess_items'] += 1
                    elif status == 'no_sales':
                        summary['no_sales_items'] += 1
                    elif status == 'empty':
                        summary['empty_items'] += 1
                    else:
                        summary['good_items'] += 1
                    
                    summary['total_stock_quantity'] += wh_data['current_stock']
                    summary['total_stock_value'] += wh_data['stock_value']
                    summary['total_order_quantity'] += wh_data['order_quantity']
                    summary['total_order_value'] += wh_data['order_value']
                    
                    if wh_data['order_quantity'] > 0:
                        summary['total_deficit_value'] += wh_data['order_value']
                    
                    if wh_data['surplus'] > 0 and item['price'] > 0:
                        summary['total_surplus_value'] += wh_data['surplus'] * item['price']
                    
                    if wh_data['months_of_stock'] < 999:
                        months_list.append(wh_data['months_of_stock'])
            
            # Средние месяцы запаса
            if months_list:
                summary['average_months_stock'] = np.mean(months_list)
            
            warehouse_summary[warehouse_key] = summary
        
        return warehouse_summary


def apply_complete_warehouse_restore(system):
    """
    Применяет ПОЛНОЕ восстановление анализа складов
    """
    try:
        st.info("🎯 Восстанавливаем ПОЛНУЮ функциональность анализа складов...")
        
        # Создаем восстановленный анализатор
        system.warehouse_analyzer = DetailedWarehouseAnalyzer()
        
        # Восстанавливаем полные методы
        def analyze_warehouse_stock_with_details(remains_df, ads_data, store_ads_by_city=None, 
                                               min_days=10, max_days=50):
            """Восстановленный полный анализ складов"""
            return system.warehouse_analyzer.analyze_warehouse_stock_detailed(
                remains_df, ads_data, store_ads_by_city, min_days, max_days
            )
        
        def get_warehouse_recommendations(analysis_results=None):
            """Восстановленные детальные рекомендации"""
            return system.warehouse_analyzer.get_warehouse_recommendations(analysis_results)
        
        # Привязываем к системе
        system.analyze_warehouse_stock_with_details = analyze_warehouse_stock_with_details
        system.get_warehouse_recommendations = get_warehouse_recommendations
        
        # Отмечаем что восстановление применено
        system._complete_warehouse_restored = True
        
        st.success("✅ ПОЛНАЯ функциональность анализа складов восстановлена!")
        st.info("""
        🎯 **Восстановлено:**
        - ✅ Полный анализ по каждому складу отдельно
        - ✅ Интеграция с ценами из ADS (колонка 'Посл. закупка')
        - ✅ Показ ВСЕХ товаров с детальной информацией
        - ✅ Персональные настройки для каждого склада
        - ✅ Денежные расчеты (остатки, заказы, дефицит)
        - ✅ Детальные рекомендации по каждому складу
        - ✅ Полная статистика и отчеты
        """)
        
        return True
        
    except Exception as e:
        st.error(f"❌ Ошибка восстановления: {str(e)}")
        return False


def create_complete_warehouse_page():
    """
    Создает ПОЛНУЮ страницу анализа складов с восстановленной функциональностью
    """
    
    def complete_warehouse_analysis_page(system):
        """
        ПОЛНАЯ страница анализа складов
        """
        
        st.header("📦 Полный анализ складов")
        st.caption("Восстановленная версия с полной функциональностью")
        
        # Применяем восстановление
        if not hasattr(system, '_complete_warehouse_restored'):
            with st.spinner("🔧 Восстанавливаем полную функциональность..."):
                success = apply_complete_warehouse_restore(system)
                if not success:
                    st.error("❌ Не удалось восстановить функциональность")
                    return
        
        # Проверяем наличие ADS
        has_ads = hasattr(system, 'calculated_ads') and system.calculated_ads is not None
        
        if has_ads:
            st.success(f"✅ ADS данные готовы: {len(system.calculated_ads)} товаров")
            
            # Проверяем цены в ADS
            price_columns = ['last_purchase_price', 'цена', 'price', 'стоимость']
            price_found = None
            for col in price_columns:
                if col in system.calculated_ads.columns:
                    items_with_prices = (system.calculated_ads[col] > 0).sum()
                    if items_with_prices > 0:
                        price_found = col
                        st.success(f"💰 Цены найдены в колонке '{col}': {items_with_prices} товаров")
                        break
            
            if not price_found:
                st.warning("⚠️ Цены в ADS не найдены - анализ будет без денежных расчетов")
        else:
            st.warning("⚠️ ADS не рассчитан - сначала рассчитайте ADS в соответствующем разделе")
        
        # Настройки анализа
        st.subheader("⚙️ Настройки анализа")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            min_days = st.number_input("Глобальный минимум дней:", value=10, min_value=5, max_value=60)
        with col2:
            max_days = st.number_input("Глобальный максимум дней:", value=50, min_value=15, max_value=120)
        with col3:
            show_debug = st.checkbox("Показать отладку", value=False)
        
        # Показываем настройки складов
        with st.expander("🏪 Настройки складов (персональные параметры)"):
            st.write("Каждый склад имеет свои оптимальные параметры запасов:")
            
            for warehouse_key, config in system.warehouse_analyzer.warehouse_config.items():
                st.write(f"""
                **{config['short_name']}** ({config['city']})
                - Тип: {config['type']}
                - Дни запаса: {config['min_days']}-{config['max_days']}
                - Колонка в файле: {config['col'] + 1}
                - Полное название: {config['name']}
                """)
        
        # Загрузка файла остатков
        st.subheader("📂 Загрузка файла остатков")
        
        uploaded_file = st.file_uploader(
            "Выберите файл остатков:",
            type=['xlsx', 'xls'],
            help="Файл с номенклатурой в A1, данные с 4й строки, склады в определенных колонках"
        )
        
        if uploaded_file:
            
            # Читаем файл с точной структурой
            with st.spinner("📖 Читаем файл с известной структурой..."):
                remains_df = system.warehouse_analyzer.read_remains_file_with_exact_structure(uploaded_file)
            
            if remains_df.empty:
                st.error("❌ Не удалось прочитать файл")
                return
            
            # Статистика загруженного файла
            col1, col2, col3, col4, col5 = st.columns(5)
            
            with col1:
                st.metric("Товаров", len(remains_df))
            
            with col2:
                warehouse_cols = [col for col in remains_df.columns if col.endswith('_остаток')]
                st.metric("Складов", len(warehouse_cols))
            
            with col3:
                total_stock = remains_df['итого_остаток'].sum()
                st.metric("Общий остаток", f"{total_stock:,.0f}")
            
            with col4:
                items_with_stock = (remains_df['итого_остаток'] > 0).sum()
                st.metric("С остатками", items_with_stock)
            
            with col5:
                empty_items = (remains_df['итого_остаток'] == 0).sum()
                st.metric("Пустых", empty_items)
            
            # Кнопка полного анализа
            st.subheader("🚀 Запуск полного анализа")
            
            if st.button("🔍 Запустить ПОЛНЫЙ детальный анализ складов", type="primary"):
                
                # Определяем данные для анализа
                ads_data = getattr(system, 'calculated_ads', pd.DataFrame())
                
                with st.spinner("🔄 Выполняем ПОЛНЫЙ детальный анализ всех товаров по всем складам..."):
                    analysis_results = system.analyze_warehouse_stock_with_details(
                        remains_df, 
                        ads_data,
                        None,  # store_ads_by_city
                        min_days,
                        max_days
                    )
                
                if analysis_results:
                    # Сохраняем результаты
                    system.warehouse_analysis_results = analysis_results
                    system.warehouse_remains_df = remains_df
                    
                    # Получаем детальные рекомендации
                    recommendations = system.get_warehouse_recommendations(analysis_results)
                    system.warehouse_recommendations = recommendations
                    
                    # Показываем ПОЛНЫЕ результаты
                    show_complete_analysis_results(analysis_results, recommendations, show_debug)
                else:
                    st.error("❌ Анализ не дал результатов")
        
        # Показываем сохраненные результаты
        if hasattr(system, 'warehouse_analysis_results') and system.warehouse_analysis_results:
            st.markdown("---")
            st.subheader("📊 Последние результаты полного анализа")
            
            last_analysis = system.warehouse_analysis_results[0]['analysis_timestamp']
            items_count = len(system.warehouse_analysis_results)
            st.caption(f"Последний анализ: {last_analysis} ({items_count} товаров)")
            
            if st.button("🔄 Показать последние результаты"):
                show_complete_analysis_results(
                    system.warehouse_analysis_results,
                    getattr(system, 'warehouse_recommendations', {}),
                    show_debug
                )
    
    return complete_warehouse_analysis_page


def show_complete_analysis_results(analysis_results: List[Dict], recommendations: Dict, show_debug: bool = False):
    """
    Показывает ПОЛНЫЕ результаты анализа с восстановленной функциональностью
    """
    
    st.subheader("📈 Результаты полного анализа складов")
    
    # Общая статистика
    total_items = len(analysis_results)
    critical_items = sum(1 for item in analysis_results if item['overall_status'] == 'critical')
    warning_items = sum(1 for item in analysis_results if item['overall_status'] == 'warning')
    good_items = total_items - critical_items - warning_items
    
    # Финансовая статистика
    total_stock_value = sum(item['total_stock_value'] for item in analysis_results)
    total_order_value = sum(item['total_order_value'] for item in analysis_results)
    items_with_prices = sum(1 for item in analysis_results if item['price'] > 0)
    
    # Карточки со статистикой
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    
    with col1:
        st.metric("Всего товаров", total_items)
    with col2:
        st.metric("🔴 Критичных", critical_items, delta=f"-{critical_items/total_items*100:.1f}%")
    with col3:
        st.metric("🟡 Требуют внимания", warning_items, delta=f"{warning_items/total_items*100:.1f}%")
    with col4:
        st.metric("🟢 В норме", good_items, delta=f"+{good_items/total_items*100:.1f}%")
    with col5:
        st.metric("💰 Стоимость остатков", f"{total_stock_value:,.0f} ₽")
    with col6:
        st.metric("🛒 К заказу", f"{total_order_value:,.0f} ₽")
    
    # Дополнительная статистика
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("С ценами", f"{items_with_prices}/{total_items}")
    with col2:
        coverage = (items_with_prices / total_items * 100) if total_items > 0 else 0
        st.metric("Покрытие ценами", f"{coverage:.1f}%")
    with col3:
        avg_price = sum(item['price'] for item in analysis_results if item['price'] > 0) / items_with_prices if items_with_prices > 0 else 0
        st.metric("Средняя цена", f"{avg_price:.2f} ₽")
    
    # ДЕТАЛЬНАЯ статистика по складам
    if recommendations:
        st.subheader("🏪 Детальная статистика по складам")
        
        # Создаем расширенную таблицу складов
        warehouse_table = []
        for wh_key, summary in recommendations.items():
            warehouse_table.append({
                'Склад': summary['short_name'],
                'Город': summary['city'],
                'Тип': summary['type'],
                'Настройки': summary['settings'],
                'Всего товаров': summary['total_items'],
                'С остатками': summary['items_with_stock'],
                '🔴 Критичных': summary['critical_items'],
                '🟡 Внимания': summary['warning_items'],
                '🔵 Избыток': summary['excess_items'],
                '⚪ Нет продаж': summary['no_sales_items'],
                'Остаток (шт)': f"{summary['total_stock_quantity']:,.0f}",
                'Стоимость остатков': f"{summary['total_stock_value']:,.0f} ₽",
                'К заказу (шт)': f"{summary['total_order_quantity']:,.0f}",
                'К заказу (₽)': f"{summary['total_order_value']:,.0f} ₽",
                'Средний запас (мес)': f"{summary['average_months_stock']:.1f}",
                'Дефицит (₽)': f"{summary['total_deficit_value']:,.0f} ₽",
                'Избыток (₽)': f"{summary['total_surplus_value']:,.0f} ₽"
            })
        
        if warehouse_table:
            st.dataframe(pd.DataFrame(warehouse_table), use_container_width=True)
    
    # ПОЛНАЯ детальная таблица товаров
    st.subheader("📋 Полная детальная информация по товарам")
    
    # Расширенные фильтры
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        status_filter = st.selectbox(
            "Фильтр по статусу:",
            ["Все товары", "🔴 Критичные", "🟡 Требуют внимания", "🟢 В норме", "📦 С остатками", "🚫 Без остатков", "💰 С ценами", "🔍 Без цен"]
        )
    
    with col2:
        sort_by = st.selectbox(
            "Сортировать по:",
            ["Статусу + Стоимость заказа", "ADS (убыв)", "Остаткам (убыв)", "Стоимости заказа", "Алфавиту", "Цене"]
        )
    
    with col3:
        warehouse_filter = st.selectbox(
            "Склад для фильтра:",
            ["Все склады"] + list(recommendations.keys()) if recommendations else ["Все склады"]
        )
    
    with col4:
        max_items = st.number_input("Показать товаров:", min_value=10, max_value=2000, value=100)
    
    with col5:
        show_warehouses = st.checkbox("Показать все склады", value=True)
    
    # Фильтруем и сортируем данные
    filtered_results = analysis_results.copy()
    
    # Применяем фильтры
    if status_filter == "🔴 Критичные":
        filtered_results = [item for item in filtered_results if item['overall_status'] == 'critical']
    elif status_filter == "🟡 Требуют внимания":
        filtered_results = [item for item in filtered_results if item['overall_status'] == 'warning']
    elif status_filter == "🟢 В норме":
        filtered_results = [item for item in filtered_results if item['overall_status'] == 'good']
    elif status_filter == "📦 С остатками":
        filtered_results = [item for item in filtered_results if item['total_stock'] > 0]
    elif status_filter == "🚫 Без остатков":
        filtered_results = [item for item in filtered_results if item['total_stock'] == 0]
    elif status_filter == "💰 С ценами":
        filtered_results = [item for item in filtered_results if item['price'] > 0]
    elif status_filter == "🔍 Без цен":
        filtered_results = [item for item in filtered_results if item['price'] == 0]
    
    # Фильтр по складу
    if warehouse_filter != "Все склады":
        filtered_results = [
            item for item in filtered_results 
            if warehouse_filter in item['warehouses'] and item['warehouses'][warehouse_filter]['current_stock'] > 0
        ]
    
    # Сортировка
    if sort_by == "Статусу + Стоимость заказа":
        status_order = {'critical': 0, 'warning': 1, 'good': 2}
        filtered_results.sort(key=lambda x: (status_order.get(x['overall_status'], 3), -x['total_order_value']))
    elif sort_by == "ADS (убыв)":
        filtered_results.sort(key=lambda x: -x['ads'])
    elif sort_by == "Остаткам (убыв)":
        filtered_results.sort(key=lambda x: -x['total_stock'])
    elif sort_by == "Стоимости заказа":
        filtered_results.sort(key=lambda x: -x['total_order_value'])
    elif sort_by == "Цене":
        filtered_results.sort(key=lambda x: -x['price'])
    else:  # Алфавиту
        filtered_results.sort(key=lambda x: x['номенклатура'])
    
    # Ограничиваем количество
    filtered_results = filtered_results[:max_items]
    
    if filtered_results:
        
        # Создаем ПОЛНУЮ таблицу для отображения
        display_data = []
        
        for item in filtered_results:
            # Эмодзи статуса
            status_emoji = {
                'critical': '🔴',
                'warning': '🟡',
                'good': '🟢'
            }.get(item['overall_status'], '⚪')
            
            # Базовая информация
            row = {
                'Статус': status_emoji,
                'Номенклатура': item['номенклатура'][:50] + "..." if len(item['номенклатура']) > 50 else item['номенклатура'],
                'ADS': f"{item['ads']:.2f}",
                'Цена': f"{item['price']:.2f} ₽" if item['price'] > 0 else "-",
                'Общий остаток': f"{item['total_stock']:.0f}",
                'Стоимость остатков': f"{item['total_stock_value']:,.0f} ₽" if item['total_stock_value'] > 0 else "-",
                'К заказу (шт)': f"{item['total_order_quantity']:.0f}" if item['total_order_quantity'] > 0 else "-",
                'К заказу (₽)': f"{item['total_order_value']:,.0f} ₽" if item['total_order_value'] > 0 else "-",
                'Месяцев запаса': f"{item['min_months_across_warehouses']:.1f}" if item['min_months_across_warehouses'] < 999 else "∞",
                'Критичных складов': item['critical_warehouses_count'],
                'Требуют внимания': item['warning_warehouses_count']
            }
            
            # Добавляем данные по складам (если включено)
            if show_warehouses:
                for wh_key, wh_data in item['warehouses'].items():
                    current = wh_data['current_stock']
                    order = wh_data['order_quantity']
                    status_wh = wh_data['status']
                    
                    # Форматируем в зависимости от статуса
                    if status_wh == 'critical':
                        row[f"{wh_data['short_name']}"] = f"🔴 {current:.0f} (+{order:.0f})"
                    elif status_wh == 'warning':
                        row[f"{wh_data['short_name']}"] = f"🟡 {current:.0f} (+{order:.0f})"
                    elif status_wh == 'excess':
                        row[f"{wh_data['short_name']}"] = f"🔵 {current:.0f}"
                    elif status_wh == 'no_sales':
                        row[f"{wh_data['short_name']}"] = f"⚪ {current:.0f}"
                    elif current > 0:
                        row[f"{wh_data['short_name']}"] = f"🟢 {current:.0f}"
                    else:
                        row[f"{wh_data['short_name']}"] = "0"
            
            display_data.append(row)
        
        # Показываем ПОЛНУЮ таблицу
        df_display = pd.DataFrame(display_data)
        st.dataframe(df_display, use_container_width=True)
        
        # Статистика по отфильтрованным данным
        filtered_critical = sum(1 for item in filtered_results if item['overall_status'] == 'critical')
        filtered_warning = sum(1 for item in filtered_results if item['overall_status'] == 'warning')
        filtered_order_value = sum(item['total_order_value'] for item in filtered_results)
        
        st.caption(f"""
        📊 Показано {len(filtered_results)} из {len(analysis_results)} товаров | 
        🔴 Критичных: {filtered_critical} | 🟡 Внимания: {filtered_warning} | 
        💰 К заказу: {filtered_order_value:,.0f} ₽
        """)
        
        # Отладочная информация
        if show_debug:
            with st.expander("🐛 Отладочная информация"):
                st.write("**Параметры анализа первого товара:**")
                if filtered_results:
                    debug_item = filtered_results[0]
                    st.json({
                        'номенклатура': debug_item['номенклатура'],
                        'общий_статус': debug_item['overall_status'],
                        'ads': debug_item['ads'],
                        'цена': debug_item['price'],
                        'параметры_анализа': debug_item.get('parameters', {}),
                        'количество_складов': len(debug_item['warehouses']),
                        'склады': {k: v['status'] for k, v in debug_item['warehouses'].items()}
                    })
    
    else:
        st.info("📋 Нет товаров, соответствующих выбранным фильтрам")
    
    # ПОЛНЫЙ экспорт результатов
    st.subheader("📤 Экспорт полных результатов")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if st.button("📊 Полный Excel отчет"):
            excel_data = create_complete_excel_report(analysis_results, recommendations)
            
            st.download_button(
                label="💾 Скачать полный отчет",
                data=excel_data,
                file_name=f"complete_warehouse_analysis_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    with col2:
        if st.button("🛒 Список заказов"):
            orders_data = create_detailed_orders_export(analysis_results)
            
            st.download_button(
                label="💾 Скачать заказы CSV",
                data=orders_data,
                file_name=f"warehouse_orders_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv"
            )
    
    with col3:
        if st.button("🏪 Статистика складов"):
            warehouses_data = create_warehouses_export(recommendations)
            
            st.download_button(
                label="💾 Скачать статистику",
                data=warehouses_data,
                file_name=f"warehouses_stats_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv"
            )
    
    with col4:
        if st.button("📋 Критичные товары"):
            critical_data = create_critical_items_export(analysis_results)
            
            st.download_button(
                label="💾 Скачать критичные",
                data=critical_data,
                file_name=f"critical_items_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv"
            )


def create_complete_excel_report(analysis_results: List[Dict], recommendations: Dict) -> bytes:
    """
    Создает ПОЛНЫЙ Excel отчет с восстановленной функциональностью
    """
    
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    
    # Лист 1: Сводка
    ws_summary = wb.active
    ws_summary.title = "Сводка анализа"
    
    # Заголовок с полной информацией
    ws_summary['A1'] = "ПОЛНЫЙ АНАЛИЗ ОСТАТКОВ ПО СКЛАДАМ"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A2'] = f"Дата создания: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A3'] = f"Товаров проанализировано: {len(analysis_results)}"
    
    # Полная статистика
    total_items = len(analysis_results)
    critical_items = sum(1 for item in analysis_results if item['overall_status'] == 'critical')
    warning_items = sum(1 for item in analysis_results if item['overall_status'] == 'warning')
    total_order_value = sum(item['total_order_value'] for item in analysis_results)
    total_stock_value = sum(item['total_stock_value'] for item in analysis_results)
    items_with_prices = sum(1 for item in analysis_results if item['price'] > 0)
    
    ws_summary['A5'] = "ОБЩАЯ СТАТИСТИКА"
    ws_summary['A5'].font = Font(size=12, bold=True)
    
    stats = [
        ("Всего товаров:", total_items),
        ("Критичных товаров:", critical_items),
        ("Требуют внимания:", warning_items),
        ("В норме:", total_items - critical_items - warning_items),
        ("Товаров с ценами:", items_with_prices),
        ("Покрытие ценами:", f"{items_with_prices/total_items*100:.1f}%"),
        ("Стоимость остатков:", f"{total_stock_value:,.2f} ₽"),
        ("Стоимость к заказу:", f"{total_order_value:,.2f} ₽")
    ]
    
    for i, (label, value) in enumerate(stats, 6):
        ws_summary[f'A{i}'] = label
        ws_summary[f'B{i}'] = value
    
    # Подробная статистика по складам
    if recommendations:
        row = len(stats) + 8
        ws_summary[f'A{row}'] = "ДЕТАЛЬНАЯ СТАТИСТИКА ПО СКЛАДАМ"
        ws_summary[f'A{row}'].font = Font(size=12, bold=True)
        
        headers = ['Склад', 'Город', 'Тип', 'Настройки', 'Товаров', 'С остатками', 'Критичных', 'Внимания', 'Избыток', 'Остаток (₽)', 'К заказу (₽)']
        for col, header in enumerate(headers, 1):
            ws_summary.cell(row=row+1, column=col, value=header).font = Font(bold=True)
        
        for i, (wh_key, data) in enumerate(recommendations.items(), row+2):
            ws_summary.cell(row=i, column=1, value=data['short_name'])
            ws_summary.cell(row=i, column=2, value=data['city'])
            ws_summary.cell(row=i, column=3, value=data['type'])
            ws_summary.cell(row=i, column=4, value=data['settings'])
            ws_summary.cell(row=i, column=5, value=data['total_items'])
            ws_summary.cell(row=i, column=6, value=data['items_with_stock'])
            ws_summary.cell(row=i, column=7, value=data['critical_items'])
            ws_summary.cell(row=i, column=8, value=data['warning_items'])
            ws_summary.cell(row=i, column=9, value=data['excess_items'])
            ws_summary.cell(row=i, column=10, value=data['total_stock_value'])
            ws_summary.cell(row=i, column=11, value=data['total_order_value'])
    
    # Лист 2: Все товары с полной детализацией
    ws_all = wb.create_sheet("Все товары")
    
    all_items_data = []
    for item in analysis_results:
        row_data = {
            'Номенклатура': item['номенклатура'],
            'ADS': item['ads'],
            'Цена': item['price'],
            'Общий_остаток': item['total_stock'],
            'Стоимость_остатков': item['total_stock_value'],
            'К_заказу_шт': item['total_order_quantity'],
            'К_заказу_руб': item['total_order_value'],
            'Статус': item['overall_status'],
            'Месяцев_запаса': item['min_months_across_warehouses'] if item['min_months_across_warehouses'] < 999 else 0,
            'Критичных_складов': item['critical_warehouses_count'],
            'Требуют_внимания': item['warning_warehouses_count']
        }
        
        # Добавляем все склады
        for wh_key, wh_data in item['warehouses'].items():
            row_data[f"{wh_data['short_name']}_остаток"] = wh_data['current_stock']
            row_data[f"{wh_data['short_name']}_заказать"] = wh_data['order_quantity']
            row_data[f"{wh_data['short_name']}_статус"] = wh_data['status']
            row_data[f"{wh_data['short_name']}_стоимость"] = wh_data['stock_value']
            row_data[f"{wh_data['short_name']}_месяцев"] = wh_data['months_of_stock'] if wh_data['months_of_stock'] < 999 else 0
        
        all_items_data.append(row_data)
    
    if all_items_data:
        df_all = pd.DataFrame(all_items_data)
        for r in dataframe_to_rows(df_all, index=False, header=True):
            ws_all.append(r)
        
        # Форматируем заголовки
        for cell in ws_all[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Листы по складам
    for wh_key, wh_summary in recommendations.items():
        ws_warehouse = wb.create_sheet(wh_summary['short_name'])
        
        # Заголовок склада
        ws_warehouse['A1'] = f"АНАЛИЗ СКЛАДА: {wh_summary['short_name']}"
        ws_warehouse['A1'].font = Font(size=14, bold=True)
        ws_warehouse['A2'] = f"Город: {wh_summary['city']}, Тип: {wh_summary['type']}"
        ws_warehouse['A3'] = f"Настройки: {wh_summary['settings']}"
        
        # Статистика склада
        ws_warehouse['A5'] = "СТАТИСТИКА СКЛАДА"
        ws_warehouse['A5'].font = Font(size=12, bold=True)
        
        wh_stats = [
            ("Всего товаров:", wh_summary['total_items']),
            ("С остатками:", wh_summary['items_with_stock']),
            ("Критичных:", wh_summary['critical_items']),
            ("Требуют внимания:", wh_summary['warning_items']),
            ("В избытке:", wh_summary['excess_items']),
            ("Остаток (шт):", f"{wh_summary['total_stock_quantity']:,.0f}"),
            ("Стоимость остатков:", f"{wh_summary['total_stock_value']:,.2f} ₽"),
            ("К заказу (шт):", f"{wh_summary['total_order_quantity']:,.0f}"),
            ("К заказу (₽):", f"{wh_summary['total_order_value']:,.2f} ₽")
        ]
        
        for i, (label, value) in enumerate(wh_stats, 6):
            ws_warehouse[f'A{i}'] = label
            ws_warehouse[f'B{i}'] = value
        
        # Товары склада
        warehouse_items = []
        for item in analysis_results:
            if wh_key in item['warehouses']:
                wh_data = item['warehouses'][wh_key]
                warehouse_items.append({
                    'Номенклатура': item['номенклатура'],
                    'ADS': item['ads'],
                    'Цена': item['price'],
                    'Остаток': wh_data['current_stock'],
                    'Минимум': wh_data['min_stock'],
                    'Максимум': wh_data['max_stock'],
                    'К_заказу': wh_data['order_quantity'],
                    'Стоимость_остатка': wh_data['stock_value'],
                    'Стоимость_заказа': wh_data['order_value'],
                    'Месяцев_запаса': wh_data['months_of_stock'] if wh_data['months_of_stock'] < 999 else 0,
                    'Статус': wh_data['status'],
                    'Рекомендация': wh_data['recommendation']
                })
        
        if warehouse_items:
            # Добавляем таблицу товаров
            start_row = len(wh_stats) + 8
            ws_warehouse[f'A{start_row}'] = "ТОВАРЫ СКЛАДА"
            ws_warehouse[f'A{start_row}'].font = Font(size=12, bold=True)
            
            df_wh = pd.DataFrame(warehouse_items)
            for r_idx, r in enumerate(dataframe_to_rows(df_wh, index=False, header=True), start_row + 1):
                for c_idx, value in enumerate(r, 1):
                    ws_warehouse.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Форматируем заголовки
                    if r_idx == start_row + 1:
                        ws_warehouse.cell(row=r_idx, column=c_idx).font = Font(bold=True)
                        ws_warehouse.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Сохраняем в байты
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def create_detailed_orders_export(analysis_results: List[Dict]) -> str:
    """
    Создает детальный экспорт заказов
    """
    
    orders_data = []
    
    for item in analysis_results:
        for wh_key, wh_data in item['warehouses'].items():
            if wh_data['order_quantity'] > 0:
                orders_data.append({
                    'Номенклатура': item['номенклатура'],
                    'Склад': wh_data['short_name'],
                    'Город': wh_data['city'],
                    'Тип_склада': wh_data['type'],
                    'ADS': item['ads'],
                    'Цена': item['price'],
                    'Текущий_остаток': wh_data['current_stock'],
                    'Минимум': wh_data['min_stock'],
                    'Максимум': wh_data['max_stock'],
                    'Дефицит': wh_data['min_deficit'],
                    'К_заказу_шт': wh_data['order_quantity'],
                    'Сумма_заказа': wh_data['order_value'],
                    'Месяцев_запаса': wh_data['months_of_stock'] if wh_data['months_of_stock'] < 999 else 0,
                    'Статус': wh_data['status'],
                    'Приоритет': 'Высокий' if wh_data['status'] == 'critical' else 'Средний',
                    'Настройки_склада': wh_data['settings'],
                    'Рекомендация': wh_data['recommendation']
                })
    
    if orders_data:
        # Сортируем по приоритету и сумме
        orders_data.sort(key=lambda x: (0 if x['Приоритет'] == 'Высокий' else 1, -x['Сумма_заказа']))
        
        df_orders = pd.DataFrame(orders_data)
        return df_orders.to_csv(index=False, encoding='utf-8-sig')
    else:
        return "Нет товаров к заказу"


def create_warehouses_export(recommendations: Dict) -> str:
    """
    Создает экспорт статистики складов
    """
    
    warehouses_data = []
    
    for wh_key, data in recommendations.items():
        warehouses_data.append({
            'Ключ_склада': wh_key,
            'Склад': data['short_name'],
            'Полное_название': data['warehouse_name'],
            'Город': data['city'],
            'Тип': data['type'],
            'Настройки_дней': data['settings'],
            'Всего_товаров': data['total_items'],
            'С_остатками': data['items_with_stock'],
            'Критичных': data['critical_items'],
            'Требуют_внимания': data['warning_items'],
            'Избыток': data['excess_items'],
            'Нет_продаж': data['no_sales_items'],
            'Пустых': data['empty_items'],
            'В_норме': data['good_items'],
            'Остаток_шт': data['total_stock_quantity'],
            'Стоимость_остатков_руб': data['total_stock_value'],
            'К_заказу_шт': data['total_order_quantity'],
            'К_заказу_руб': data['total_order_value'],
            'Средний_запас_мес': data['average_months_stock'],
            'Дефицит_руб': data['total_deficit_value'],
            'Избыток_руб': data['total_surplus_value'],
            'Критичность_процент': (data['critical_items'] / data['total_items'] * 100) if data['total_items'] > 0 else 0,
            'Эффективность_запасов': (data['good_items'] / data['total_items'] * 100) if data['total_items'] > 0 else 0
        })
    
    if warehouses_data:
        df_warehouses = pd.DataFrame(warehouses_data)
        return df_warehouses.to_csv(index=False, encoding='utf-8-sig')
    else:
        return "Нет данных по складам"


def create_critical_items_export(analysis_results: List[Dict]) -> str:
    """
    Создает экспорт критичных товаров
    """
    
    critical_data = []
    
    for item in analysis_results:
        if item['overall_status'] == 'critical':
            critical_warehouses = []
            total_deficit_qty = 0
            total_deficit_value = 0
            
            for wh_key, wh_data in item['warehouses'].items():
                if wh_data['status'] == 'critical':
                    critical_warehouses.append(wh_data['short_name'])
                    total_deficit_qty += wh_data['order_quantity']
                    total_deficit_value += wh_data['order_value']
            
            critical_data.append({
                'Номенклатура': item['номенклатура'],
                'ADS': item['ads'],
                'Цена': item['price'],
                'Общий_остаток': item['total_stock'],
                'Стоимость_остатков': item['total_stock_value'],
                'Общий_дефицит_шт': total_deficit_qty,
                'Общий_дефицит_руб': total_deficit_value,
                'Месяцев_запаса': item['min_months_across_warehouses'] if item['min_months_across_warehouses'] < 999 else 0,
                'Критичных_складов': item['critical_warehouses_count'],
                'Склады_с_дефицитом': '; '.join(critical_warehouses),
                'Общая_потребность_шт': item['total_order_quantity'],
                'Общая_потребность_руб': item['total_order_value'],
                'Приоритет_по_сумме': 'Высокий' if total_deficit_value > 10000 else ('Средний' if total_deficit_value > 1000 else 'Низкий')
            })
    
    if critical_data:
        # Сортируем по сумме дефицита
        critical_data.sort(key=lambda x: -x['Общий_дефицит_руб'])
        
        df_critical = pd.DataFrame(critical_data)
        return df_critical.to_csv(index=False, encoding='utf-8-sig')
    else:
        return "Нет критичных товаров"


# Главная функция для быстрого применения
def quick_restore_warehouse_analysis(system):
    """
    Быстрое восстановление полной функциональности анализа складов
    """
    
    try:
        # Применяем полное восстановление
        if not hasattr(system, '_complete_warehouse_restored'):
            apply_complete_warehouse_restore(system)
        
        # Создаем и запускаем полную страницу
        warehouse_page = create_complete_warehouse_page()
        warehouse_page(system)
        
        return True
        
    except Exception as e:
        st.error(f"❌ Ошибка восстановления: {str(e)}")
        
        # Показываем инструкции по ручному исправлению
        st.error("""
        **Не удалось автоматически восстановить функциональность.**
        
        **Ручное исправление:**
        1. Создайте файл `complete_warehouse_restore.py` из артефакта
        2. Замените вашу функцию `warehouse_analysis_page` на:
        
        ```python
        def warehouse_analysis_page(system):
            from complete_warehouse_restore import quick_restore_warehouse_analysis
            quick_restore_warehouse_analysis(system)
        ```
        """)
        
        return False


# Инструкции по использованию
def get_restoration_instructions():
    """
    Полные инструкции по восстановлению функциональности
    """
    
    return """
# 🎯 ИНСТРУКЦИЯ ПО ВОССТАНОВЛЕНИЮ ПОЛНОЙ ФУНКЦИОНАЛЬНОСТИ

## 🚀 БЫСТРОЕ ПРИМЕНЕНИЕ:

### Замените вашу функцию warehouse_analysis_page на:

```python
def warehouse_analysis_page(system):
    from complete_warehouse_restore import quick_restore_warehouse_analysis
    quick_restore_warehouse_analysis(system)
```

## ✅ ЧТО ВОССТАНАВЛИВАЕТСЯ:

### 🔍 ПОЛНЫЙ АНАЛИЗ:
- ✅ **Все товары** показываются в анализе
- ✅ **Каждый склад анализируется отдельно** с персональными настройками
- ✅ **9 складов** с точной структурой из вашего файла
- ✅ **Детальный статус** по каждому складу для каждого товара

### 💰 ИНТЕГРАЦИЯ С ЦЕНАМИ:
- ✅ **Цены из ADS** (колонка 'Посл. закупка' или 'last_purchase_price')
- ✅ **Денежные расчеты** (стоимость остатков, заказов, дефицита)
- ✅ **Приоритизация по стоимости** 
- ✅ **Финансовая статистика** по складам

### 🏪 ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО СКЛАДАМ:
- ✅ **Персональные настройки** (Шымкент: 15-45 дней, Алматы: 20-60 дней и т.д.)
- ✅ **Статистика по каждому складу** отдельно
- ✅ **Рекомендации** с учетом типа склада
- ✅ **Анализ эффективности** складов

### 📊 РАСШИРЕННАЯ ОТЧЕТНОСТЬ:
- ✅ **Полный Excel отчет** с листами по каждому складу
- ✅ **Детальные заказы** с приоритизацией
- ✅ **Статистика складов** с эффективностью
- ✅ **Критичные товары** с финансовым анализом

## 🎯 СТРУКТУРА ВАШИХ СКЛАДОВ:

1. **Шымкент_Склад** (колонка D) - 15-45 дней
2. **Шымкент_Магазин** (колонка E) - 10-30 дней  
3. **Алматы_Склад** (колонка G) - 20-60 дней
4. **База_Комплект** (колонка I) - 25-75 дней
5. **Барыс_Склад** (колонка J) - 15-40 дней
6. **Казыбаева_Склад** (колонка K) - 12-35 дней
7. **Астана_Магазин** (колонка L) - 10-30 дней
8. **Астана_Склад** (колонка M) - 15-45 дней
9. **Казыбаева_Магазин** (колонка N) - 8-25 дней

## 📋 ИСПОЛЬЗОВАНИЕ:

1. **Рассчитайте ADS** с ценами в разделе "ADS расчет"
2. **Перейдите в анализ складов** 
3. **Загрузите файл остатков** (автоматически распознается структура)
4. **Запустите полный анализ** - все товары по всем складам
5. **Получите детальные результаты** с ценами и рекомендациями
6. **Экспортируйте отчеты** в нужном формате

## 🔧 РЕЗУЛЬТАТ:

После применения у вас будет **ПОЛНАЯ** система анализа складов как в оригинале, но с исправленной ошибкой 'номенклатура'.
"""


if __name__ == "__main__":
    print("🎯 Модуль полного восстановления анализа складов загружен")
    print("Восстанавливает все функции: цены, детальный анализ, полный показ товаров")
    print("\nДля использования:")
    print("from complete_warehouse_restore import quick_restore_warehouse_analysis")
    print("\nИнструкции:")
    print(get_restoration_instructions())