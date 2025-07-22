#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Простой анализ без pandas - используем openpyxl напрямую
try:
    from openpyxl import load_workbook
    
    print('=== АНАЛИЗ ФАЙЛА ОСТАТКОВ ===')
    
    # Загружаем файл
    wb = load_workbook('остатки на 08.07.2025.xlsx')
    ws = wb.active
    
    print(f'Активный лист: {ws.title}')
    print(f'Размер: {ws.max_row} строк, {ws.max_column} колонок')
    print()
    
    print('=== ЗАГОЛОВКИ (первая строка) ===')
    headers = []
    for col in range(1, min(ws.max_column + 1, 20)):  # Первые 20 колонок
        cell_value = ws.cell(row=1, column=col).value
        headers.append(cell_value)
        print(f'Колонка {col:2d}: {repr(cell_value)}')
    
    print('\n=== ПЕРВЫЕ 5 СТРОК ДАННЫХ ===')
    for row in range(1, min(6, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(6, ws.max_column + 1)):  # Первые 5 колонок
            cell_value = ws.cell(row=row, column=col).value
            row_data.append(str(cell_value)[:30])  # Ограничиваем длину
        print(f'Строка {row}: {row_data}')
    
    print('\n=== ПОИСК КОЛОНКИ С НАИМЕНОВАНИЯМИ ===')
    name_candidates = []
    
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            header_str = str(header).lower()
            if any(keyword in header_str for keyword in ['наименование', 'номенклатура', 'товар', 'name']):
                name_candidates.append((col, header))
                print(f'✅ Кандидат колонка {col}: {repr(header)}')
    
    if not name_candidates:
        print('❌ Не найдено явных заголовков с наименованиями')
        print('🔍 Проверяем содержимое первых колонок...')
        
        for col in range(1, min(6, ws.max_column + 1)):
            header = ws.cell(row=1, column=col).value
            print(f'\nКолонка {col}: {repr(header)}')
            
            # Проверяем несколько значений в колонке
            values = []
            for row in range(2, min(7, ws.max_row + 1)):
                val = ws.cell(row=row, column=col).value
                if val:
                    values.append(str(val)[:50])
            
            print(f'Примеры значений: {values[:3]}')
            
            # Проверяем, есть ли текстовые данные
            text_count = sum(1 for v in values if v and not v.replace('.', '').replace(',', '').isdigit())
            print(f'Текстовых значений: {text_count} из {len(values)}')
            
            if text_count > len(values) * 0.8 and len(values) > 0:
                print('✅ Вероятно, это колонка с наименованиями товаров')
    
    wb.close()
    
except ImportError:
    print('❌ openpyxl не установлен. Используем альтернативный способ...')
    
    # Попробуем через xlrd
    try:
        import xlrd
        print('Используем xlrd для анализа...')
        # ... код для xlrd
    except ImportError:
        print('❌ Нет доступных библиотек для чтения Excel')
        print('Попробуйте установить: pip install openpyxl')

except Exception as e:
    print(f'❌ Ошибка: {e}')
    import traceback
    traceback.print_exc()