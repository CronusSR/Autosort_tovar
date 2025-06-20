# final_warehouse_solution.py
"""
🎯 ФИНАЛЬНОЕ РЕШЕНИЕ АНАЛИЗА СКЛАДОВ
Точно настроено под ваши файлы и структуру данных
"""

import pandas as pd
import streamlit as st
import numpy as np
from typing import Dict, List, Any, Optional, Tuple


class YourWarehouseFileReader:
    """
    Ридер файлов точно под вашу структуру данных
    Основан на анализе файла 'остатки мини.xlsx'
    """
    
    def __init__(self):
        # Точная конфигурация ваших складов
        self.warehouse_mapping = {
            '4 Склад фурнитуры АЗМ Шымкент "Овощная база"': {
                'short_name': 'Шымкент_Склад',
                'city': 'шымкент',
                'type': 'warehouse',
                'min_days': 15,
                'max_days': 45
            },
            '6 Склад фурнитуры "Овощная база" Магазин': {
                'short_name': 'Шымкент_Магазин',
                'city': 'шымкент', 
                'type': 'store',
                'min_days': 10,
                'max_days': 30
            },
            'АО Склад Фурнитура TRADE': {
                'short_name': 'Алматы_Склад',
                'city': 'алматы',
                'type': 'specialized_store',
                'min_days': 20,
                'max_days': 60
            },
            'База Склад Фурнитура Комплект': {
                'short_name': 'База_Комплект',
                'city': 'алматы',
                'type': 'warehouse',
                'min_days': 25,
                'max_days': 75
            },
            'Барыс Склад Фурнитура TRADE': {
                'short_name': 'Барыс_Склад',
                'city': 'алматы',
                'type': 'trade_store',
                'min_days': 15,
                'max_days': 40
            },
            'Казыбаева Склад Фурнитура TRADE': {
                'short_name': 'Казыбаева_Склад',
                'city': 'алматы',
                'type': 'trade_store',
                'min_days': 12,
                'max_days': 35
            },
            'Магазин фурнитуры': {
                'short_name': 'Астана_Магазин',
                'city': 'астана',
                'type': 'store',
                'min_days': 10,
                'max_days': 30
            },
            'склад фурнитура № 1': {
                'short_name': 'Астана_Склад',
                'city': 'астана',
                'type': 'warehouse',
                'min_days': 15,
                'max_days': 45
            },
            'ТД Казыбаева ФУРНИТУРА магазин': {
                'short_name': 'Казыбаева_Магазин',
                'city': 'алматы',
                'type': 'retail_store',
                'min_days': 8,
                'max_days': 25
            }
        }
    
    def read_your_remains_file(self, uploaded_file) -> pd.DataFrame:
        """
        ТОЧНОЕ чтение вашего файла остатков
        Структура: Номенклатура в A1, склады с D1, данные с 4й строки
        """
        
        try:
            st.info("📖 Читаем ваш файл остатков (структура определена)...")
            
            # Читаем Excel файл
            if uploaded_file.name.endswith('.xlsx'):
                df_raw = pd.read_excel(uploaded_file, header=None)
            else:
                df_raw = pd.read_excel(uploaded_file, engine='xlrd', header=None)
            
            st.success(f"✅ Файл прочитан: {df_raw.shape[0]} строк, {df_raw.shape[1]} колонок")
            
            # Константы на основе анализа вашего файла
            NOMENCLATURE_ROW = 0  # строка 1 (индекс 0)
            NOMENCLATURE_COL = 0  # колонка A (индекс 0)
            DATA_START_ROW = 3    # строка 4 (индекс 3)
            
            # Извлекаем заголовки складов из первой строки
            header_row = df_raw.iloc[NOMENCLATURE_ROW]
            
            warehouse_columns = []
            for col_idx, cell_value in enumerate(header_row):
                if pd.notna(cell_value):
                    cell_str = str(cell_value).strip()
                    
                    # Проверяем, является ли это складом
                    if any(keyword in cell_str.lower() for keyword in ['склад', 'магазин', 'trade']):
                        
                        # Ищем конфигурацию склада
                        config = self.warehouse_mapping.get(cell_str)
                        if config:
                            warehouse_columns.append({
                                'col_index': col_idx,
                                'full_name': cell_str,
                                'short_name': config['short_name'],
                                'config': config
                            })
                            st.write(f"✅ Склад найден: {config['short_name']} (колонка {col_idx + 1})")
                        else:
                            # Создаем базовую конфигурацию для неизвестного склада
                            short_name = self._create_short_name_from_text(cell_str)
                            warehouse_columns.append({
                                'col_index': col_idx,
                                'full_name': cell_str,
                                'short_name': short_name,
                                'config': {
                                    'short_name': short_name,
                                    'city': 'неизвестно',
                                    'type': 'unknown',
                                    'min_days': 15,
                                    'max_days': 45
                                }
                            })
                            st.warning(f"⚠️ Неизвестный склад: {short_name} (колонка {col_idx + 1})")
            
            st.success(f"🏪 Всего складов найдено: {len(warehouse_columns)}")
            
            # Читаем данные товаров
            items_data = []
            processed_count = 0
            
            for row_idx in range(DATA_START_ROW, len(df_raw)):
                row = df_raw.iloc[row_idx]
                
                # Получаем номенклатуру
                nomenclature = row.iloc[NOMENCLATURE_COL] if NOMENCLATURE_COL < len(row) else None
                
                if pd.isna(nomenclature) or not str(nomenclature).strip():
                    continue
                
                nomenclature = str(nomenclature).strip()
                
                # Создаем запись товара
                item_data = {'номенклатура': nomenclature}
                
                # Собираем остатки по складам
                total_stock = 0
                
                for wh in warehouse_columns:
                    col_idx = wh['col_index']
                    stock_value = 0
                    
                    if col_idx < len(row):
                        cell_value = row.iloc[col_idx]
                        if pd.notna(cell_value):
                            try:
                                stock_value = float(cell_value)
                                total_stock += stock_value
                            except (ValueError, TypeError):
                                stock_value = 0
                    
                    # Добавляем остаток по складу
                    item_data[f"{wh['short_name']}_остаток"] = stock_value
                
                # Добавляем общий остаток
                item_data['итого_остаток'] = total_stock
                
                items_data.append(item_data)
                processed_count += 1
            
            if not items_data:
                raise ValueError("❌ Не найдено товаров с данными")
            
            # Создаем DataFrame
            result_df = pd.DataFrame(items_data)
            
            st.success(f"✅ Обработано товаров: {processed_count}")
            st.info(f"📊 Колонки: {list(result_df.columns)}")
            
            # Сохраняем конфигурацию складов для анализа
            self.detected_warehouses = warehouse_columns
            
            # Показываем превью
            with st.expander("👀 Превью обработанных данных"):
                st.dataframe(result_df.head(), use_container_width=True)
                
                # Статистика по складам
                st.write("📊 **Статистика по складам:**")
                for wh in warehouse_columns:
                    col_name = f"{wh['short_name']}_остаток"
                    if col_name in result_df.columns:
                        total_stock = result_df[col_name].sum()
                        items_with_stock = (result_df[col_name] > 0).sum()
                        st.write(f"  - **{wh['short_name']}**: {total_stock:,.0f} (товаров: {items_with_stock})")
            
            return result_df
            
        except Exception as e:
            st.error(f"❌ Ошибка чтения файла: {str(e)}")
            st.exception(e)
            return pd.DataFrame()
    
    def _create_short_name_from_text(self, full_name: str) -> str:
        """Создает короткое имя из полного названия склада"""
        
        name_lower = full_name.lower()
        
        # Определяем город
        if 'шымкент' in name_lower:
            city = 'Шымкент'
        elif 'алматы' in name_lower or 'trade' in name_lower:
            city = 'Алматы'
        elif 'астана' in name_lower:
            city = 'Астана'
        else:
            city = 'Город'
        
        # Определяем тип
        if 'магазин' in name_lower:
            type_name = 'Магазин'
        elif 'склад' in name_lower:
            type_name = 'Склад'
        else:
            type_name = 'Точка'
        
        return f"{city}_{type_name}"


class YourWarehouseAnalyzer:
    """
    Анализатор складов настроенный под ваши данные и бизнес-логику
    """
    
    def __init__(self, warehouse_reader: YourWarehouseFileReader):
        self.warehouse_reader = warehouse_reader
        self.last_analysis_results = None
        
    def analyze_your_warehouses(self, remains_df: pd.DataFrame, ads_data: pd.DataFrame,
                               global_min_days: int = 15, global_max_days: int = 45) -> List[Dict]:
        """
        ГЛАВНЫЙ метод анализа ваших складов с учетом всех особенностей
        """
        
        try:
            st.info(f"🔄 Запускаем анализ {len(remains_df)} товаров по {len(self.warehouse_reader.detected_warehouses)} складам...")
            
            # Проверяем наличие ADS данных
            if ads_data is None or ads_data.empty:
                st.warning("⚠️ Нет ADS данных - анализ будет ограничен")
                use_ads = False
            else:
                use_ads = True
                st.success(f"✅ ADS данные доступны: {len(ads_data)} товаров")
                
                # Проверяем наличие цен
                price_columns = ['last_purchase_price', 'цена', 'price', 'стоимость', 'закупочная_цена']
                price_column = None
                for col in price_columns:
                    if col in ads_data.columns:
                        price_column = col
                        break
                
                if price_column:
                    st.success(f"💰 Цены найдены в колонке '{price_column}'")
                else:
                    st.warning("⚠️ Цены в ADS данных не найдены")
            
            analysis_results = []
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Анализируем каждый товар
            for idx, (_, item) in enumerate(remains_df.iterrows()):
                
                # Обновляем прогресс
                progress = (idx + 1) / len(remains_df)
                progress_bar.progress(progress)
                status_text.text(f"Анализируем товар {idx + 1}/{len(remains_df)}: {item['номенклатура'][:30]}...")
                
                item_name = str(item['номенклатура']).strip()
                total_stock = float(item.get('итого_остаток', 0))
                
                # Получаем ADS и цену для товара
                ads_value = 0
                item_price = 0
                
                if use_ads:
                    ads_match = ads_data[ads_data['номенклатура'] == item_name]
                    if not ads_match.empty:
                        ads_value = float(ads_match.iloc[0].get('ads', 0))
                        if price_column:
                            try:
                                item_price = float(ads_match.iloc[0].get(price_column, 0))
                            except (ValueError, TypeError):
                                item_price = 0
                
                # Анализ по каждому складу
                warehouses_analysis = {}
                overall_critical_count = 0
                overall_warning_count = 0
                total_order_quantity = 0
                total_order_value = 0
                
                for wh in self.warehouse_reader.detected_warehouses:
                    wh_key = wh['short_name']
                    wh_config = wh['config']
                    
                    # Текущий остаток на складе
                    stock_col = f"{wh_key}_остаток"
                    current_stock = float(item.get(stock_col, 0))
                    
                    # Персональные настройки склада или глобальные
                    min_days = wh_config.get('min_days', global_min_days)
                    max_days = wh_config.get('max_days', global_max_days)
                    
                    # Расчеты MIN/MAX запасов
                    min_stock = ads_value * min_days if ads_value > 0 else 0
                    max_stock = ads_value * max_days if ads_value > 0 else 0
                    
                    # Дефицит/избыток
                    min_deficit = max(0, min_stock - current_stock)
                    max_deficit = max(0, max_stock - current_stock) 
                    surplus = max(0, current_stock - max_stock)
                    
                    # Месяцы запаса
                    if ads_value > 0:
                        days_of_stock = current_stock / ads_value
                        months_of_stock = days_of_stock / 30
                    elif current_stock > 0:
                        months_of_stock = 999  # Бесконечно (нет продаж)
                        days_of_stock = 999
                    else:
                        months_of_stock = 0
                        days_of_stock = 0
                    
                    # Определяем статус склада
                    if ads_value > 0:
                        if current_stock < min_stock:
                            if min_deficit > ads_value * 7:  # Более недели дефицита
                                status = 'critical'
                                overall_critical_count += 1
                            else:
                                status = 'warning'
                                overall_warning_count += 1
                        elif current_stock > max_stock:
                            status = 'excess'
                        else:
                            status = 'good'
                    elif current_stock > 0:
                        status = 'no_sales'
                    else:
                        status = 'empty'
                    
                    # Количество к заказу
                    order_quantity = min_deficit if status in ['critical', 'warning'] else 0
                    total_order_quantity += order_quantity
                    
                    # Денежные расчеты
                    stock_value = current_stock * item_price if item_price > 0 else 0
                    order_value = order_quantity * item_price if item_price > 0 else 0
                    total_order_value += order_value
                    
                    # Рекомендация
                    if status == 'critical':
                        recommendation = f"🔴 КРИТИЧНО! Заказать {order_quantity:.0f} шт. (дефицит {min_deficit:.0f})"
                    elif status == 'warning':
                        recommendation = f"🟡 Заказать {order_quantity:.0f} шт. (ниже минимума)"
                    elif status == 'excess':
                        recommendation = f"🔵 Избыток {surplus:.0f} шт. (выше максимума)"
                    elif status == 'no_sales':
                        recommendation = f"⚪ Нет продаж, остаток {current_stock:.0f} шт."
                    elif status == 'empty':
                        recommendation = "⚫ Нет остатков и продаж"
                    else:
                        recommendation = f"🟢 В норме ({current_stock:.0f} шт.)"
                    
                    warehouses_analysis[wh_key] = {
                        'warehouse_name': wh['full_name'],
                        'short_name': wh_key,
                        'city': wh_config.get('city', 'неизвестно'),
                        'type': wh_config.get('type', 'unknown'),
                        'current_stock': current_stock,
                        'min_stock': min_stock,
                        'max_stock': max_stock,
                        'min_deficit': min_deficit,
                        'max_deficit': max_deficit,
                        'surplus': surplus,
                        'days_of_stock': days_of_stock,
                        'months_of_stock': months_of_stock,
                        'status': status,
                        'order_quantity': order_quantity,
                        'stock_value': stock_value,
                        'order_value': order_value,
                        'recommendation': recommendation,
                        'settings': f"{min_days}-{max_days} дней"
                    }
                
                # Общий статус товара
                if overall_critical_count > 0:
                    overall_status = 'critical'
                elif overall_warning_count > 0:
                    overall_status = 'warning'
                else:
                    overall_status = 'good'
                
                # Минимальные месяцы запаса среди всех складов
                months_list = [w['months_of_stock'] for w in warehouses_analysis.values() 
                              if w['months_of_stock'] < 999]
                min_months_across_warehouses = min(months_list) if months_list else 0
                
                # Итоговая запись товара
                analysis_results.append({
                    'номенклатура': item_name,
                    'total_stock': total_stock,
                    'ads': ads_value,
                    'price': item_price,
                    'total_stock_value': total_stock * item_price if item_price > 0 else 0,
                    'total_order_quantity': total_order_quantity,
                    'total_order_value': total_order_value,
                    'min_months_across_warehouses': min_months_across_warehouses,
                    'overall_status': overall_status,
                    'critical_warehouses_count': overall_critical_count,
                    'warning_warehouses_count': overall_warning_count,
                    'warehouses': warehouses_analysis,
                    'analysis_timestamp': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'parameters': {
                        'global_min_days': global_min_days,
                        'global_max_days': global_max_days,
                        'has_ads': use_ads,
                        'has_prices': price_column is not None,
                        'price_column': price_column
                    }
                })
            
            # Завершаем прогресс
            progress_bar.progress(1.0)
            status_text.text("✅ Анализ завершен!")
            
            # Сохраняем результаты
            self.last_analysis_results = analysis_results
            
            # Статистика анализа
            total_items = len(analysis_results)
            critical_items = sum(1 for item in analysis_results if item['overall_status'] == 'critical')
            warning_items = sum(1 for item in analysis_results if item['overall_status'] == 'warning')
            
            st.success(f"""
            ✅ **Анализ завершен успешно!**
            
            📊 **Общая статистика:**
            - Всего товаров: {total_items}
            - 🔴 Критичных: {critical_items} ({critical_items/total_items*100:.1f}%)
            - 🟡 Требуют внимания: {warning_items} ({warning_items/total_items*100:.1f}%)
            - 🟢 В норме: {total_items - critical_items - warning_items} ({(total_items - critical_items - warning_items)/total_items*100:.1f}%)
            
            💰 **Финансовая статистика:**
            - Общая стоимость остатков: {sum(item['total_stock_value'] for item in analysis_results):,.0f} ₽
            - Общая стоимость к заказу: {sum(item['total_order_value'] for item in analysis_results):,.0f} ₽
            """)
            
            return analysis_results
            
        except Exception as e:
            st.error(f"❌ Ошибка анализа: {str(e)}")
            st.exception(e)
            return []
    
    def get_warehouse_summary(self, analysis_results: List[Dict] = None) -> Dict:
        """
        Получает сводку по складам
        """
        
        if not analysis_results:
            analysis_results = self.last_analysis_results
        
        if not analysis_results:
            return {}
        
        warehouse_summary = {}
        
        # Собираем все уникальные склады
        all_warehouses = set()
        for item in analysis_results:
            all_warehouses.update(item['warehouses'].keys())
        
        # Анализируем каждый склад
        for wh_key in all_warehouses:
            
            # Базовая информация о складе
            sample_wh = None
            for item in analysis_results:
                if wh_key in item['warehouses']:
                    sample_wh = item['warehouses'][wh_key]
                    break
            
            if not sample_wh:
                continue
            
            summary = {
                'warehouse_name': sample_wh['warehouse_name'],
                'short_name': sample_wh['short_name'],
                'city': sample_wh['city'],
                'type': sample_wh['type'],
                'total_items': 0,
                'items_with_stock': 0,
                'critical_items': 0,
                'warning_items': 0,
                'excess_items': 0,
                'no_sales_items': 0,
                'empty_items': 0,
                'good_items': 0,
                'total_stock_quantity': 0,
                'total_stock_value': 0,
                'total_order_quantity': 0,
                'total_order_value': 0,
                'average_months_stock': 0
            }
            
            # Собираем статистику
            months_list = []
            
            for item in analysis_results:
                if wh_key in item['warehouses']:
                    wh_data = item['warehouses'][wh_key]
                    
                    summary['total_items'] += 1
                    
                    if wh_data['current_stock'] > 0:
                        summary['items_with_stock'] += 1
                    
                    status = wh_data['status']
                    if status == 'critical':
                        summary['critical_items'] += 1
                    elif status == 'warning':
                        summary['warning_items'] += 1
                    elif status == 'excess':
                        summary['excess_items'] += 1
                    elif status == 'no_sales':
                        summary['no_sales_items'] += 1
                    elif status == 'empty':
                        summary['empty_items'] += 1
                    else:
                        summary['good_items'] += 1
                    
                    summary['total_stock_quantity'] += wh_data['current_stock']
                    summary['total_stock_value'] += wh_data['stock_value']
                    summary['total_order_quantity'] += wh_data['order_quantity']
                    summary['total_order_value'] += wh_data['order_value']
                    
                    if wh_data['months_of_stock'] < 999:
                        months_list.append(wh_data['months_of_stock'])
            
            # Средние месяцы запаса
            if months_list:
                summary['average_months_stock'] = np.mean(months_list)
            
            warehouse_summary[wh_key] = summary
        
        return warehouse_summary


def create_your_complete_warehouse_page():
    """
    ПОЛНАЯ страница анализа складов для вашей системы
    """
    
    def your_warehouse_analysis_page(system):
        """
        Ваша персональная страница анализа складов
        """
        
        st.header("📦 Анализ складов - Ваша версия")
        st.caption("Настроено под структуру ваших файлов и бизнес-процессы")
        
        # Проверяем наличие ADS
        if not hasattr(system, 'calculated_ads') or system.calculated_ads is None:
            st.warning("⚠️ Для полноценного анализа сначала рассчитайте ADS в разделе 'ADS расчет'")
            
            # Показываем что можно сделать без ADS
            with st.expander("ℹ️ Что можно делать без ADS"):
                st.info("""
                **Без ADS данных доступно:**
                - ✅ Загрузка и обработка файла остатков
                - ✅ Просмотр текущих остатков по складам
                - ✅ Базовая статистика по складам
                - ✅ Экспорт данных остатков
                
                **Для полного анализа нужен ADS:**
                - 🔴 Расчет минимальных и максимальных запасов
                - 🔴 Определение дефицита/избытка
                - 🔴 Рекомендации по заказам
                - 🔴 Анализ месяцев запаса
                """)
        else:
            st.success(f"✅ ADS данные готовы: {len(system.calculated_ads)} товаров")
        
        # Инициализируем компоненты
        if not hasattr(system, 'warehouse_file_reader'):
            system.warehouse_file_reader = YourWarehouseFileReader()
        
        if not hasattr(system, 'warehouse_analyzer'):
            system.warehouse_analyzer = YourWarehouseAnalyzer(system.warehouse_file_reader)
        
        # Настройки анализа
        st.subheader("⚙️ Настройки анализа")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            global_min_days = st.number_input(
                "Глобальный минимум дней:", 
                min_value=5, max_value=90, value=15,
                help="Используется если у склада нет персональных настроек"
            )
        with col2:
            global_max_days = st.number_input(
                "Глобальный максимум дней:", 
                min_value=15, max_value=180, value=45,
                help="Используется если у склада нет персональных настроек"
            )
        with col3:
            show_debug = st.checkbox("Показать отладку", value=False)
        
        # Показываем настройки складов
        with st.expander("🏪 Персональные настройки складов"):
            st.write("Каждый склад имеет свои оптимальные параметры запасов:")
            
            for full_name, config in system.warehouse_file_reader.warehouse_mapping.items():
                st.write(f"""
                **{config['short_name']}** ({config['city']})
                - Тип: {config['type']}
                - Дни запаса: {config['min_days']}-{config['max_days']}
                - Полное название: {full_name}
                """)
        
        # Загрузка файла остатков
        st.subheader("📂 Загрузка файла остатков")
        
        uploaded_file = st.file_uploader(
            "Выберите файл остатков:",
            type=['xlsx', 'xls'],
            help="Ваш файл с номенклатурой в A1 и складами начиная с D1"
        )
        
        if uploaded_file:
            
            # Читаем файл
            with st.spinner("📖 Обрабатываем ваш файл..."):
                remains_df = system.warehouse_file_reader.read_your_remains_file(uploaded_file)
            
            if remains_df.empty:
                st.error("❌ Не удалось обработать файл")
                return
            
            # Статистика загруженного файла
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Товаров загружено", len(remains_df))
            
            with col2:
                warehouse_cols = [col for col in remains_df.columns if col.endswith('_остаток')]
                st.metric("Складов найдено", len(warehouse_cols))
            
            with col3:
                total_stock = remains_df['итого_остаток'].sum()
                st.metric("Общий остаток", f"{total_stock:,.0f}")
            
            with col4:
                items_with_stock = (remains_df['итого_остаток'] > 0).sum()
                st.metric("С остатками", items_with_stock)
            
            # Кнопка анализа
            st.subheader("🚀 Запуск анализа")
            
            if st.button("🔍 Запустить полный анализ складов", type="primary"):
                
                # Определяем какие данные использовать для ADS
                ads_data = getattr(system, 'calculated_ads', None)
                
                with st.spinner("🔄 Выполняем детальный анализ..."):
                    analysis_results = system.warehouse_analyzer.analyze_your_warehouses(
                        remains_df, 
                        ads_data,
                        global_min_days,
                        global_max_days
                    )
                

                    
                if analysis_results:
                    # Сохраняем результаты
                    system.warehouse_analysis_results = analysis_results
                    system.warehouse_remains_df = remains_df
                    
                    # Получаем сводку по складам
                    warehouse_summary = system.warehouse_analyzer.get_warehouse_summary(analysis_results)
                    system.warehouse_summary = warehouse_summary
                    
                    st.success("✅ Анализ завершен! Результаты сохранены в системе.")
                    
                    # Показываем результаты
                    show_your_analysis_results(analysis_results, warehouse_summary, show_debug)
                
                else:
                    st.error("❌ Анализ не дал результатов")
        
        # Показываем сохраненные результаты
        if hasattr(system, 'warehouse_analysis_results') and system.warehouse_analysis_results:
            st.markdown("---")
            st.subheader("📊 Последние результаты анализа")
            
            last_analysis = system.warehouse_analysis_results[0]['analysis_timestamp']
            st.caption(f"Последний анализ: {last_analysis}")
            
            if st.button("🔄 Показать последние результаты"):
                show_your_analysis_results(
                    system.warehouse_analysis_results,
                    getattr(system, 'warehouse_summary', {}),
                    show_debug
                )
    
    return your_warehouse_analysis_page


def show_your_analysis_results(analysis_results: List[Dict], warehouse_summary: Dict, show_debug: bool = False):
    """
    Показывает результаты анализа в удобном формате
    """
    
    st.subheader("📈 Результаты анализа складов")
    
    # Общая статистика
    total_items = len(analysis_results)
    critical_items = sum(1 for item in analysis_results if item['overall_status'] == 'critical')
    warning_items = sum(1 for item in analysis_results if item['overall_status'] == 'warning')
    good_items = total_items - critical_items - warning_items
    
    # Финансовая статистика
    total_stock_value = sum(item['total_stock_value'] for item in analysis_results)
    total_order_value = sum(item['total_order_value'] for item in analysis_results)
    
    # Карточки со статистикой
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.metric("Всего товаров", total_items)
    with col2:
        st.metric("🔴 Критичных", critical_items, delta=f"-{critical_items/total_items*100:.1f}%")
    with col3:
        st.metric("🟡 Требуют внимания", warning_items, delta=f"{warning_items/total_items*100:.1f}%")
    with col4:
        st.metric("🟢 В норме", good_items, delta=f"+{good_items/total_items*100:.1f}%")
    with col5:
        st.metric("💰 К заказу", f"{total_order_value:,.0f} ₽")
    
    # Статистика по складам
    if warehouse_summary:
        st.subheader("🏪 Статистика по складам")
        
        # Создаем таблицу складов
        warehouse_table = []
        for wh_key, summary in warehouse_summary.items():
            warehouse_table.append({
                'Склад': summary['short_name'],
                'Город': summary['city'],
                'Всего товаров': summary['total_items'],
                '🔴 Критичных': summary['critical_items'],
                '🟡 Требуют внимания': summary['warning_items'],
                '🔵 Избыток': summary['excess_items'],
                'Остаток (шт)': f"{summary['total_stock_quantity']:,.0f}",
                'Стоимость остатков': f"{summary['total_stock_value']:,.0f} ₽",
                'К заказу (шт)': f"{summary['total_order_quantity']:,.0f}",
                'К заказу (₽)': f"{summary['total_order_value']:,.0f} ₽",
                'Средний запас (мес)': f"{summary['average_months_stock']:.1f}"
            })
        
        if warehouse_table:
            st.dataframe(pd.DataFrame(warehouse_table), use_container_width=True)
    
    # Детальная таблица товаров
    st.subheader("📋 Детальная информация по товарам")
    
    # Фильтры
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        status_filter = st.selectbox(
            "Фильтр по статусу:",
            ["Все товары", "🔴 Только критичные", "🟡 Требуют внимания", "🟢 В норме", "📦 С остатками", "🚫 Без остатков"]
        )
    
    with col2:
        sort_by = st.selectbox(
            "Сортировать по:",
            ["Статусу", "ADS (убыв)", "Остаткам (убыв)", "Алфавиту", "Стоимости заказа"]
        )
    
    with col3:
        warehouse_filter = st.selectbox(
            "Склад:",
            ["Все склады"] + list(warehouse_summary.keys()) if warehouse_summary else ["Все склады"]
        )
    
    with col4:
        max_items = st.number_input("Показать товаров:", min_value=10, max_value=1000, value=50)
    
    # Фильтруем и сортируем данные
    filtered_results = analysis_results.copy()
    
    # Фильтр по статусу
    if status_filter == "🔴 Только критичные":
        filtered_results = [item for item in filtered_results if item['overall_status'] == 'critical']
    elif status_filter == "🟡 Требуют внимания":
        filtered_results = [item for item in filtered_results if item['overall_status'] == 'warning']
    elif status_filter == "🟢 В норме":
        filtered_results = [item for item in filtered_results if item['overall_status'] == 'good']
    elif status_filter == "📦 С остатками":
        filtered_results = [item for item in filtered_results if item['total_stock'] > 0]
    elif status_filter == "🚫 Без остатков":
        filtered_results = [item for item in filtered_results if item['total_stock'] == 0]
    
    # Фильтр по складу
    if warehouse_filter != "Все склады":
        filtered_results = [
            item for item in filtered_results 
            if warehouse_filter in item['warehouses'] and item['warehouses'][warehouse_filter]['current_stock'] > 0
        ]
    
    # Сортировка
    if sort_by == "Статусу":
        status_order = {'critical': 0, 'warning': 1, 'good': 2}
        filtered_results.sort(key=lambda x: (status_order.get(x['overall_status'], 3), -x['total_order_value']))
    elif sort_by == "ADS (убыв)":
        filtered_results.sort(key=lambda x: -x['ads'])
    elif sort_by == "Остаткам (убыв)":
        filtered_results.sort(key=lambda x: -x['total_stock'])
    elif sort_by == "Стоимости заказа":
        filtered_results.sort(key=lambda x: -x['total_order_value'])
    else:  # Алфавиту
        filtered_results.sort(key=lambda x: x['номенклатура'])
    
    # Ограничиваем количество
    filtered_results = filtered_results[:max_items]
    
    if filtered_results:
        
        # Создаем таблицу для отображения
        display_data = []
        
        for item in filtered_results:
            # Эмодзи статуса
            status_emoji = {
                'critical': '🔴',
                'warning': '🟡',
                'good': '🟢'
            }.get(item['overall_status'], '⚪')
            
            # Базовая строка
            row = {
                'Статус': status_emoji,
                'Номенклатура': item['номенклатура'][:40] + "..." if len(item['номенклатура']) > 40 else item['номенклатура'],
                'ADS': f"{item['ads']:.2f}",
                'Цена': f"{item['price']:.2f} ₽" if item['price'] > 0 else "-",
                'Общий остаток': f"{item['total_stock']:.0f}",
                'К заказу (шт)': f"{item['total_order_quantity']:.0f}" if item['total_order_quantity'] > 0 else "-",
                'К заказу (₽)': f"{item['total_order_value']:,.0f} ₽" if item['total_order_value'] > 0 else "-",
                'Месяцев запаса': f"{item['min_months_across_warehouses']:.1f}" if item['min_months_across_warehouses'] < 999 else "∞"
            }
            
            # Добавляем данные по складам
            for wh_key, wh_data in item['warehouses'].items():
                current = wh_data['current_stock']
                order = wh_data['order_quantity']
                
                if order > 0:
                    row[f"{wh_data['short_name']}"] = f"{current:.0f} ⚠️ +{order:.0f}"
                elif current > 0:
                    row[f"{wh_data['short_name']}"] = f"{current:.0f}"
                else:
                    row[f"{wh_data['short_name']}"] = "0"
            
            display_data.append(row)
        
        # Показываем таблицу
        df_display = pd.DataFrame(display_data)
        st.dataframe(df_display, use_container_width=True)
        
        # Статистика по отфильтрованным данным
        st.caption(f"Показано {len(filtered_results)} из {len(analysis_results)} товаров")
        
        # Детальная информация при отладке
        if show_debug:
            with st.expander("🐛 Отладочная информация"):
                st.write("**Параметры первого товара:**")
                if filtered_results:
                    st.json(filtered_results[0])
    
    else:
        st.info("📋 Нет товаров, соответствующих выбранным фильтрам")
    
    # Экспорт результатов
    st.subheader("📤 Экспорт результатов")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("📊 Экспорт всех данных в Excel"):
            excel_data = create_your_excel_report(analysis_results, warehouse_summary)
            
            st.download_button(
                label="💾 Скачать полный отчет Excel",
                data=excel_data,
                file_name=f"warehouse_analysis_full_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    with col2:
        if st.button("📋 Экспорт товаров к заказу"):
            orders_data = create_orders_csv(analysis_results)
            
            st.download_button(
                label="💾 Скачать список заказов CSV",
                data=orders_data,
                file_name=f"orders_list_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv"
            )
    
    with col3:
        if st.button("🏪 Экспорт статистики складов"):
            warehouses_data = create_warehouses_csv(warehouse_summary)
            
            st.download_button(
                label="💾 Скачать статистику складов CSV",
                data=warehouses_data,
                file_name=f"warehouses_stats_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv"
            )


def create_your_excel_report(analysis_results: List[Dict], warehouse_summary: Dict) -> bytes:
    """
    Создает полный Excel отчет по анализу складов
    """
    
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    
    # Лист 1: Сводка
    ws_summary = wb.active
    ws_summary.title = "Сводка анализа"
    
    # Заголовок
    ws_summary['A1'] = "АНАЛИЗ ОСТАТКОВ ПО СКЛАДАМ"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A2'] = f"Дата создания: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Общая статистика
    total_items = len(analysis_results)
    critical_items = sum(1 for item in analysis_results if item['overall_status'] == 'critical')
    warning_items = sum(1 for item in analysis_results if item['overall_status'] == 'warning')
    total_order_value = sum(item['total_order_value'] for item in analysis_results)
    total_stock_value = sum(item['total_stock_value'] for item in analysis_results)
    
    ws_summary['A4'] = "ОБЩАЯ СТАТИСТИКА"
    ws_summary['A4'].font = Font(size=12, bold=True)
    
    stats = [
        ("Всего товаров:", total_items),
        ("Критичных товаров:", critical_items),
        ("Требуют внимания:", warning_items),
        ("В норме:", total_items - critical_items - warning_items),
        ("Стоимость остатков:", f"{total_stock_value:,.2f} ₽"),
        ("Стоимость к заказу:", f"{total_order_value:,.2f} ₽")
    ]
    
    for i, (label, value) in enumerate(stats, 5):
        ws_summary[f'A{i}'] = label
        ws_summary[f'B{i}'] = value
    
    # Статистика по складам
    if warehouse_summary:
        row = len(stats) + 7
        ws_summary[f'A{row}'] = "СТАТИСТИКА ПО СКЛАДАМ"
        ws_summary[f'A{row}'].font = Font(size=12, bold=True)
        
        headers = ['Склад', 'Город', 'Товаров', 'Критичных', 'Требуют внимания', 'Остаток (₽)', 'К заказу (₽)']
        for col, header in enumerate(headers, 1):
            ws_summary.cell(row=row+1, column=col, value=header).font = Font(bold=True)
        
        for i, (wh_key, data) in enumerate(warehouse_summary.items(), row+2):
            ws_summary.cell(row=i, column=1, value=data['short_name'])
            ws_summary.cell(row=i, column=2, value=data['city'])
            ws_summary.cell(row=i, column=3, value=data['total_items'])
            ws_summary.cell(row=i, column=4, value=data['critical_items'])
            ws_summary.cell(row=i, column=5, value=data['warning_items'])
            ws_summary.cell(row=i, column=6, value=data['total_stock_value'])
            ws_summary.cell(row=i, column=7, value=data['total_order_value'])
    
    # Лист 2: Все товары
    ws_all = wb.create_sheet("Все товары")
    
    all_items_data = []
    for item in analysis_results:
        row_data = {
            'Номенклатура': item['номенклатура'],
            'ADS': item['ads'],
            'Цена': item['price'],
            'Общий остаток': item['total_stock'],
            'Стоимость остатков': item['total_stock_value'],
            'К заказу (шт)': item['total_order_quantity'],
            'К заказу (₽)': item['total_order_value'],
            'Статус': item['overall_status'],
            'Месяцев запаса': item['min_months_across_warehouses'] if item['min_months_across_warehouses'] < 999 else 0
        }
        
        # Добавляем данные по складам
        for wh_key, wh_data in item['warehouses'].items():
            row_data[f"{wh_data['short_name']}_остаток"] = wh_data['current_stock']
            row_data[f"{wh_data['short_name']}_заказать"] = wh_data['order_quantity']
            row_data[f"{wh_data['short_name']}_статус"] = wh_data['status']
        
        all_items_data.append(row_data)
    
    if all_items_data:
        df_all = pd.DataFrame(all_items_data)
        for r in dataframe_to_rows(df_all, index=False, header=True):
            ws_all.append(r)
        
        # Форматируем заголовки
        for cell in ws_all[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Лист 3: Критичные товары
    ws_critical = wb.create_sheet("Критичные товары")
    
    critical_data = []
    for item in analysis_results:
        if item['overall_status'] == 'critical':
            for wh_key, wh_data in item['warehouses'].items():
                if wh_data['status'] == 'critical':
                    critical_data.append({
                        'Номенклатура': item['номенклатура'],
                        'Склад': wh_data['short_name'],
                        'Город': wh_data['city'],
                        'ADS': item['ads'],
                        'Цена': item['price'],
                        'Текущий остаток': wh_data['current_stock'],
                        'Минимум': wh_data['min_stock'],
                        'Дефицит': wh_data['min_deficit'],
                        'К заказу': wh_data['order_quantity'],
                        'Сумма заказа': wh_data['order_value'],
                        'Дней запаса': wh_data['days_of_stock'],
                        'Рекомендация': wh_data['recommendation']
                    })
    
    if critical_data:
        df_critical = pd.DataFrame(critical_data)
        for r in dataframe_to_rows(df_critical, index=False, header=True):
            ws_critical.append(r)
        
        # Красное форматирование для критичных
        for cell in ws_critical[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Лист 4: Заказы по складам
    ws_orders = wb.create_sheet("Заказы по складам")
    
    for wh_key, wh_summary in warehouse_summary.items():
        if wh_summary['total_order_quantity'] > 0:
            
            # Заголовок склада
            current_row = ws_orders.max_row + 2
            ws_orders.cell(row=current_row, column=1, value=f"СКЛАД: {wh_summary['short_name']} ({wh_summary['city']})")
            ws_orders.cell(row=current_row, column=1).font = Font(size=12, bold=True)
            
            # Заголовки колонок
            headers = ['Номенклатура', 'ADS', 'Цена', 'Остаток', 'Минимум', 'К заказу', 'Сумма']
            for col, header in enumerate(headers, 1):
                ws_orders.cell(row=current_row+1, column=col, value=header).font = Font(bold=True)
            
            # Товары для заказа
            row_num = current_row + 2
            total_order_sum = 0
            
            for item in analysis_results:
                if wh_key in item['warehouses']:
                    wh_data = item['warehouses'][wh_key]
                    if wh_data['order_quantity'] > 0:
                        ws_orders.cell(row=row_num, column=1, value=item['номенклатура'])
                        ws_orders.cell(row=row_num, column=2, value=item['ads'])
                        ws_orders.cell(row=row_num, column=3, value=item['price'])
                        ws_orders.cell(row=row_num, column=4, value=wh_data['current_stock'])
                        ws_orders.cell(row=row_num, column=5, value=wh_data['min_stock'])
                        ws_orders.cell(row=row_num, column=6, value=wh_data['order_quantity'])
                        ws_orders.cell(row=row_num, column=7, value=wh_data['order_value'])
                        
                        total_order_sum += wh_data['order_value']
                        row_num += 1
            
            # Итого по складу
            ws_orders.cell(row=row_num, column=6, value="ИТОГО:")
            ws_orders.cell(row=row_num, column=7, value=total_order_sum).font = Font(bold=True)
    
    # Сохраняем в байты
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def create_orders_csv(analysis_results: List[Dict]) -> str:
    """
    Создает CSV файл с товарами к заказу
    """
    
    orders_data = []
    
    for item in analysis_results:
        for wh_key, wh_data in item['warehouses'].items():
            if wh_data['order_quantity'] > 0:
                orders_data.append({
                    'Номенклатура': item['номенклатура'],
                    'Склад': wh_data['short_name'],
                    'Город': wh_data['city'],
                    'ADS': item['ads'],
                    'Цена': item['price'],
                    'Текущий_остаток': wh_data['current_stock'],
                    'Минимум': wh_data['min_stock'],
                    'К_заказу_шт': wh_data['order_quantity'],
                    'Сумма_заказа': wh_data['order_value'],
                    'Статус': wh_data['status'],
                    'Приоритет': 'Высокий' if wh_data['status'] == 'critical' else 'Средний'
                })
    
    if orders_data:
        df_orders = pd.DataFrame(orders_data)
        return df_orders.to_csv(index=False, encoding='utf-8-sig')
    else:
        return "Нет товаров к заказу"


def create_warehouses_csv(warehouse_summary: Dict) -> str:
    """
    Создает CSV файл со статистикой складов
    """
    
    warehouses_data = []
    
    for wh_key, data in warehouse_summary.items():
        warehouses_data.append({
            'Склад': data['short_name'],
            'Полное_название': data['warehouse_name'],
            'Город': data['city'],
            'Тип': data['type'],
            'Всего_товаров': data['total_items'],
            'С_остатками': data['items_with_stock'],
            'Критичных': data['critical_items'],
            'Требуют_внимания': data['warning_items'],
            'Избыток': data['excess_items'],
            'Нет_продаж': data['no_sales_items'],
            'В_норме': data['good_items'],
            'Остаток_шт': data['total_stock_quantity'],
            'Стоимость_остатков': data['total_stock_value'],
            'К_заказу_шт': data['total_order_quantity'],
            'К_заказу_руб': data['total_order_value'],
            'Средний_запас_мес': data['average_months_stock']
        })
    
    if warehouses_data:
        df_warehouses = pd.DataFrame(warehouses_data)
        return df_warehouses.to_csv(index=False, encoding='utf-8-sig')
    else:
        return "Нет данных по складам"


# Главная функция для интеграции
def apply_your_complete_solution(system):
    """
    Применяет ваше полное решение к системе
    """
    
    try:
        st.info("🎯 Применяем ваше персональное решение анализа складов...")
        
        # Добавляем компоненты
        system.warehouse_file_reader = YourWarehouseFileReader()
        system.warehouse_analyzer = YourWarehouseAnalyzer(system.warehouse_file_reader)
        
        # Добавляем методы
        def analyze_warehouse_stock_with_details(remains_df, ads_data, store_ads_by_city=None, 
                                               min_days=15, max_days=45):
            return system.warehouse_analyzer.analyze_your_warehouses(
                remains_df, ads_data, min_days, max_days
            )
        
        def get_warehouse_recommendations(analysis_results=None):
            return system.warehouse_analyzer.get_warehouse_summary(analysis_results)
        
        system.analyze_warehouse_stock_with_details = analyze_warehouse_stock_with_details
        system.get_warehouse_recommendations = get_warehouse_recommendations
        
        # Отмечаем что решение применено
        system._your_complete_solution_applied = True
        
        st.success("✅ Ваше персональное решение успешно применено!")
        return True
        
    except Exception as e:
        st.error(f"❌ Ошибка применения решения: {str(e)}")
        return False


# Инструкции по использованию
def get_usage_instructions():
    """
    Инструкции по использованию вашего решения
    """
    
    return """
# 🎯 ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ ВАШЕГО РЕШЕНИЯ

## 🚀 БЫСТРЫЙ СТАРТ:

### 1. Сохраните код в файл `final_warehouse_solution.py`

### 2. Добавьте в ваш основной файл:

```python
# В начале файла
from final_warehouse_solution import (
    apply_your_complete_solution,
    create_your_complete_warehouse_page
)

# В функции где обрабатываются страницы:
elif page == "📦 Анализ складов":
    # Применяем ваше решение
    if not hasattr(system, '_your_complete_solution_applied'):
        apply_your_complete_solution(system)
    
    # Создаем и запускаем страницу
    warehouse_page = create_your_complete_warehouse_page()
    warehouse_page(system)
```

## 📊 СТРУКТУРА ВАШИХ ФАЙЛОВ:

✅ **Поддерживаемая структура (основана на анализе ваших файлов):**
- Номенклатура в A1
- Склады начиная с D1, E1, G1 и т.д.
- Дополнительные заголовки в строках 2-3
- Данные товаров с 4й строки

✅ **Автоматически распознает ваши склады:**
1. Шымкент_Склад (15-45 дней)
2. Шымкент_Магазин (10-30 дней)  
3. Алматы_Склад (20-60 дней)
4. База_Комплект (25-75 дней)
5. Барыс_Склад (15-40 дней)
6. Казыбаева_Склад (12-35 дней)
7. Астана_Магазин (10-30 дней)
8. Астана_Склад (15-45 дней)
9. Казыбаева_Магазин (8-25 дней)

## 🎯 ИСПОЛЬЗОВАНИЕ:

1. **Рассчитайте ADS** в разделе "ADS расчет" (с ценами)
2. **Перейдите в "Анализ складов"**
3. **Загрузите файл остатков** (ваш формат автоматически распознается)
4. **Настройте параметры** анализа
5. **Запустите анализ** - получите полные результаты
6. **Экспортируйте отчеты** в Excel/CSV

## ✅ РЕЗУЛЬТАТ:

- 🔍 **Детальный анализ** каждого товара по каждому складу
- 📊 **Персональные настройки** для каждого склада
## ✅ РЕЗУЛЬТАТ:

- 🔍 **Детальный анализ** каждого товара по каждому складу
- 📊 **Персональные настройки** для каждого склада
- 💰 **Денежные расчеты** (стоимость остатков, заказов)
- 🎯 **Приоритизация** по критичности
- 📈 **Статистика по складам** и городам
- 📤 **Полные отчеты** Excel/CSV
- 🔄 **Автоматическое распознавание** структуры файлов

## 🔧 ОТЛАДКА:

Включите "Показать отладку" для просмотра:
- Процесса чтения файла
- Найденных складов и их настроек
- Промежуточных расчетов
- Подробной статистики

## 📞 ПОДДЕРЖКА:

Если что-то не работает:
1. Проверьте структуру файла остатков
2. Убедитесь что ADS рассчитан
3. Включите режим отладки
4. Проверьте логи в консоли

## 🎉 ГОТОВО!

Ваше решение полностью настроено под ваши файлы и бизнес-процессы.
"""


if __name__ == "__main__":
    print("🎯 Финальное решение анализа складов загружено")
    print("Точно настроено под ваши файлы и структуру данных")
    print("\nДля использования:")
    print("from final_warehouse_solution import apply_your_complete_solution, create_your_complete_warehouse_page")
    print("\nИнструкции:")
    print(get_usage_instructions())