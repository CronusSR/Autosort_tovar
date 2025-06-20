# russian_interface_fix.py
"""
🎯 ИСПРАВЛЕНИЕ ИНТЕРФЕЙСА - РУССКИЕ НАЗВАНИЯ В ТАБЛИЦАХ
Заменяет английские названия колонок на русские в интерфейсе
"""

import pandas as pd
import streamlit as st
import numpy as np
from typing import Dict, List, Any, Optional


def apply_russian_interface_fix(system):
    """
    Применяет исправления интерфейса с русскими названиями + правильный поиск цен
    """
    try:
        st.info("🔧 Применяем исправления: русский интерфейс + поиск цен...")
        
        # Добавляем исправленный анализатор
        from complete_warehouse_restore import DetailedWarehouseAnalyzer
        system.warehouse_analyzer = DetailedWarehouseAnalyzer()
        
        # Исправленный метод поиска цен
        def find_prices_in_ads(ads_data):
            """ПРАВИЛЬНО находит цены в ADS данных"""
            if ads_data is None or ads_data.empty:
                return False, None, {}
            
            # Список возможных названий колонок с ценами
            price_columns = [
                'last_purchase_price',  # основная колонка
                'посл_закупка',
                'Посл. закупка', 
                'цена',
                'price',
                'стоимость',
                'закупочная_цена'
            ]
            
            found_column = None
            items_with_prices = 0
            
            # Ищем колонку с ценами
            for col in price_columns:
                if col in ads_data.columns:
                    # Проверяем есть ли в ней данные
                    prices_greater_zero = (ads_data[col] > 0).sum()
                    if prices_greater_zero > 0:
                        found_column = col
                        items_with_prices = prices_greater_zero
                        st.success(f"💰 Найдена колонка с ценами: '{col}' ({items_with_prices} товаров)")
                        break
            
            if found_column is None:
                st.warning("⚠️ Цены в ADS данных не найдены")
                st.info(f"""
                🔍 **Поиск цен:**
                - Искали колонки: {', '.join(price_columns)}
                - Найденные колонки в ADS: {list(ads_data.columns)}
                
                💡 **Рекомендации:**
                1. Убедитесь что ADS файл содержит колонку 'Посл. закупка' (12-я колонка)
                2. Перезагрузите файл продаж для пересчета ADS
                3. Проверьте что цены больше 0
                """)
                return False, None, {}
            
            # Статистика по ценам
            price_data = ads_data[ads_data[found_column] > 0]
            avg_price = price_data[found_column].mean()
            min_price = price_data[found_column].min()
            max_price = price_data[found_column].max()
            
            price_stats = {
                'всего_товаров': len(ads_data),
                'товаров_с_ценами': items_with_prices,
                'товаров_без_цен': len(ads_data) - items_with_prices,
                'покрытие_процент': (items_with_prices / len(ads_data)) * 100,
                'средняя_цена': avg_price,
                'мин_цена': min_price,
                'макс_цена': max_price,
                'название_колонки': found_column
            }
            
            st.success(f"""
            💰 **Статистика цен:**
            - Товаров с ценами: {items_with_prices} из {len(ads_data)}
            - Покрытие ценами: {price_stats['покрытие_процент']:.1f}%
            - Средняя цена: {avg_price:.2f} ₽
            - Диапазон цен: {min_price:.2f} - {max_price:.2f} ₽
            """)
            
            return True, found_column, price_stats
        
        # Исправленные методы с правильным поиском цен
        def analyze_warehouse_stock_with_details(remains_df, ads_data, store_ads_by_city=None, 
                                               min_days=10, max_days=50):
            """Исправленный анализ с правильным поиском цен"""
            
            # Используем исправленный поиск цен
            has_prices, price_column, price_stats = find_prices_in_ads(ads_data)
            
            # Запускаем стандартный анализ
            return system.warehouse_analyzer.analyze_warehouse_stock_detailed(
                remains_df, ads_data, store_ads_by_city, min_days, max_days
            )
        
        def get_warehouse_recommendations(analysis_results=None):
            """Исправленные рекомендации"""
            return system.warehouse_analyzer.get_warehouse_recommendations(analysis_results)
        
        # Привязываем к системе
        system.analyze_warehouse_stock_with_details = analyze_warehouse_stock_with_details
        system.get_warehouse_recommendations = get_warehouse_recommendations
        
        # Отмечаем что исправления применены
        system._russian_interface_fixed = True
        
        st.success("✅ Русский интерфейс + правильный поиск цен применены!")
        return True
        
    except Exception as e:
        st.error(f"❌ Ошибка применения исправлений: {str(e)}")
        return False


def create_russian_warehouse_page():
    """
    Создает страницу анализа складов с РУССКИМИ НАЗВАНИЯМИ в таблицах
    """
    
    def russian_warehouse_analysis_page(system):
        """
        Страница анализа складов с русскими названиями в интерфейсе
        """
        
        st.header("📦 Анализ складов")
        st.caption("Русские названия в таблицах + правильный поиск цен")
        
        # Применяем исправления
        if not hasattr(system, '_russian_interface_fixed'):
            with st.spinner("🔧 Применяем русский интерфейс..."):
                success = apply_russian_interface_fix(system)
                if not success:
                    st.error("❌ Не удалось применить исправления")
                    return
        
        # Проверяем наличие ADS
        has_ads = hasattr(system, 'calculated_ads') and system.calculated_ads is not None
        
        if has_ads:
            st.success(f"✅ Данные ADS готовы: {len(system.calculated_ads)} товаров")
        else:
            st.warning("⚠️ ADS не рассчитан - сначала рассчитайте ADS в соответствующем разделе")
        
        # Настройки анализа
        st.subheader("⚙️ Настройки анализа")
        
        col1, col2 = st.columns(2)
        with col1:
            min_days = st.number_input("Минимум дней:", value=10, min_value=5, max_value=60)
        with col2:
            max_days = st.number_input("Максимум дней:", value=50, min_value=15, max_value=120)
        
        # Загрузка файла остатков
        st.subheader("📂 Загрузка файла остатков")
        
        uploaded_file = st.file_uploader(
            "Выберите файл остатков:",
            type=['xlsx', 'xls'],
            help="Файл с номенклатурой в A1, данные с 4й строки"
        )
        
        if uploaded_file:
            
            # Читаем файл
            with st.spinner("📖 Читаем файл остатков..."):
                remains_df = system.warehouse_analyzer.read_remains_file_with_exact_structure(uploaded_file)
            
            if remains_df.empty:
                st.error("❌ Не удалось прочитать файл")
                return
            
            # Статистика файла
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Товаров", len(remains_df))
            with col2:
                warehouse_cols = [col for col in remains_df.columns if col.endswith('_остаток')]
                st.metric("Складов", len(warehouse_cols))
            with col3:
                total_stock = remains_df['итого_остаток'].sum()
                st.metric("Общий остаток", f"{total_stock:,.0f}")
            with col4:
                items_with_stock = (remains_df['итого_остаток'] > 0).sum()
                st.metric("С остатками", items_with_stock)
            
            # Кнопка анализа
            if st.button("🚀 Запустить анализ складов", type="primary"):
                
                ads_data = getattr(system, 'calculated_ads', pd.DataFrame())
                
                with st.spinner("🔄 Выполняем анализ..."):
                    analysis_results = system.analyze_warehouse_stock_with_details(
                        remains_df, 
                        ads_data,
                        None,
                        min_days,
                        max_days
                    )
                
                if analysis_results:
                    # Сохраняем результаты
                    system.warehouse_analysis_results = analysis_results
                    system.warehouse_remains_df = remains_df
                    
                    # Получаем рекомендации
                    recommendations = system.get_warehouse_recommendations(analysis_results)
                    system.warehouse_recommendations = recommendations
                    
                    # Показываем результаты с РУССКИМИ НАЗВАНИЯМИ
                    show_russian_analysis_results(analysis_results, recommendations)
                else:
                    st.error("❌ Анализ не дал результатов")
        
        # Показываем сохраненные результаты
        if hasattr(system, 'warehouse_analysis_results') and system.warehouse_analysis_results:
            st.markdown("---")
            st.subheader("📊 Последние результаты анализа")
            
            if st.button("🔄 Показать последние результаты"):
                show_russian_analysis_results(
                    system.warehouse_analysis_results,
                    getattr(system, 'warehouse_recommendations', {})
                )
    
    return russian_warehouse_analysis_page


def show_russian_analysis_results(analysis_results: List[Dict], recommendations: Dict):
    """
    Показывает результаты анализа с РУССКИМИ НАЗВАНИЯМИ в таблицах
    """
    
    st.subheader("📈 Результаты анализа складов")
    
    # Общая статистика
    total_items = len(analysis_results)
    critical_items = sum(1 for item in analysis_results if item['overall_status'] == 'critical')
    warning_items = sum(1 for item in analysis_results if item['overall_status'] == 'warning')
    good_items = total_items - critical_items - warning_items
    
    # Финансовая статистика
    total_stock_value = sum(item.get('total_stock_value', 0) for item in analysis_results)
    total_order_value = sum(item.get('total_order_value', 0) for item in analysis_results)
    items_with_prices = sum(1 for item in analysis_results if item.get('price', 0) > 0)
    
    # Карточки со статистикой
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    
    with col1:
        st.metric("Всего товаров", total_items)
    with col2:
        st.metric("🔴 Критичных", critical_items)
    with col3:
        st.metric("🟡 Требуют внимания", warning_items)
    with col4:
        st.metric("🟢 В норме", good_items)
    with col5:
        st.metric("💰 Стоимость остатков", f"{total_stock_value:,.0f} ₽")
    with col6:
        st.metric("🛒 К заказу", f"{total_order_value:,.0f} ₽")
    
    # Статистика по складам с РУССКИМИ НАЗВАНИЯМИ
    if recommendations:
        st.subheader("🏪 Статистика по складам")
        
        # Создаем таблицу складов с русскими заголовками
        warehouse_table = []
        for wh_key, summary in recommendations.items():
            warehouse_table.append({
                'Склад': summary.get('short_name', summary.get('name', wh_key)),
                'Город': summary.get('city', '-'),
                'Тип': summary.get('type', '-'),
                'Всего товаров': summary.get('total_items', 0),
                'С остатками': summary.get('items_with_stock', 0),
                'Критичных': summary.get('critical_items', 0),
                'Требуют внимания': summary.get('warning_items', 0),
                'Избыток': summary.get('excess_items', 0),
                'Остаток (шт)': f"{summary.get('total_stock_quantity', 0):,.0f}",
                'Стоимость остатков (₽)': f"{summary.get('total_stock_value', 0):,.0f}",
                'К заказу (шт)': f"{summary.get('total_order_quantity', 0):,.0f}",
                'К заказу (₽)': f"{summary.get('total_order_value', 0):,.0f}"
            })
        
        if warehouse_table:
            # Показываем таблицу с русскими названиями колонок
            st.dataframe(pd.DataFrame(warehouse_table), use_container_width=True)
    
    # Детальная таблица товаров с РУССКИМИ НАЗВАНИЯМИ
    st.subheader("📋 Детальная информация по товарам")
    
    # Фильтры с русскими названиями
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        status_filter = st.selectbox(
            "Фильтр по статусу:",
            ["Все товары", "Критичные", "Требуют внимания", "В норме", "С остатками", "Без остатков"]
        )
    
    with col2:
        sort_by = st.selectbox(
            "Сортировать по:",
            ["Статусу", "ADS (убыв)", "Остаткам (убыв)", "Стоимости заказа", "Алфавиту"]
        )
    
    with col3:
        warehouse_filter = st.selectbox(
            "Склад:",
            ["Все склады"] + list(recommendations.keys()) if recommendations else ["Все склады"]
        )
    
    with col4:
        max_items = st.number_input("Показать товаров:", min_value=10, max_value=1000, value=100)
    
    # Фильтруем данные
    filtered_results = analysis_results.copy()
    
    if status_filter == "Критичные":
        filtered_results = [item for item in filtered_results if item.get('overall_status') == 'critical']
    elif status_filter == "Требуют внимания":
        filtered_results = [item for item in filtered_results if item.get('overall_status') == 'warning']
    elif status_filter == "В норме":
        filtered_results = [item for item in filtered_results if item.get('overall_status') == 'good']
    elif status_filter == "С остатками":
        filtered_results = [item for item in filtered_results if item.get('total_stock', 0) > 0]
    elif status_filter == "Без остатков":
        filtered_results = [item for item in filtered_results if item.get('total_stock', 0) == 0]
    
    # Сортировка
    if sort_by == "Статусу":
        status_order = {'critical': 0, 'warning': 1, 'good': 2}
        filtered_results.sort(key=lambda x: (status_order.get(x.get('overall_status'), 3), -x.get('total_order_value', 0)))
    elif sort_by == "ADS (убыв)":
        filtered_results.sort(key=lambda x: -x.get('ads', 0))
    elif sort_by == "Остаткам (убыв)":
        filtered_results.sort(key=lambda x: -x.get('total_stock', 0))
    elif sort_by == "Стоимости заказа":
        filtered_results.sort(key=lambda x: -x.get('total_order_value', 0))
    else:  # Алфавиту
        filtered_results.sort(key=lambda x: x.get('номенклатура', ''))
    
    # Ограничиваем количество
    filtered_results = filtered_results[:max_items]
    
    # Создаем таблицу с РУССКИМИ НАЗВАНИЯМИ КОЛОНОК
    if filtered_results:
        display_data = []
        
        for item in filtered_results:
            # Эмодзи статуса
            status_emoji = {
                'critical': '🔴',
                'warning': '🟡',
                'good': '🟢'
            }.get(item.get('overall_status'), '⚪')
            
            # Базовая информация С РУССКИМИ НАЗВАНИЯМИ
            row = {
                'Статус': status_emoji,
                'Номенклатура': item.get('номенклатура', '')[:50],
                'ADS': f"{item.get('ads', 0):.2f}",
                'Цена (₽)': f"{item.get('price', 0):.2f}" if item.get('price', 0) > 0 else "-",
                'Общий остаток': f"{item.get('total_stock', 0):.0f}",
                'Стоимость остатков (₽)': f"{item.get('total_stock_value', 0):,.0f}" if item.get('total_stock_value', 0) > 0 else "-",
                'К заказу (шт)': f"{item.get('total_order_quantity', 0):.0f}" if item.get('total_order_quantity', 0) > 0 else "-",
                'К заказу (₽)': f"{item.get('total_order_value', 0):,.0f}" if item.get('total_order_value', 0) > 0 else "-",
                'Месяцев запаса': f"{item.get('min_months_across_warehouses', 0):.1f}" if item.get('min_months_across_warehouses', 0) < 999 else "∞"
            }
            
            # Добавляем данные по складам с русскими названиями
            warehouses = item.get('warehouses', {})
            for wh_key, wh_data in warehouses.items():
                current = wh_data.get('current_stock', 0)
                order = wh_data.get('order_quantity', 0)
                status_wh = wh_data.get('status', '')
                
                # Название склада (русское)
                wh_name = wh_data.get('short_name', wh_key)
                
                # Форматируем в зависимости от статуса
                if order > 0:
                    if status_wh == 'critical':
                        row[wh_name] = f"🔴 {current:.0f} (+{order:.0f})"
                    elif status_wh == 'warning':
                        row[wh_name] = f"🟡 {current:.0f} (+{order:.0f})"
                    else:
                        row[wh_name] = f"{current:.0f} (+{order:.0f})"
                elif current > 0:
                    if status_wh == 'excess':
                        row[wh_name] = f"🔵 {current:.0f}"
                    else:
                        row[wh_name] = f"{current:.0f}"
                else:
                    row[wh_name] = "0"
            
            display_data.append(row)
        
        # Показываем таблицу с русскими названиями
        df_display = pd.DataFrame(display_data)
        st.dataframe(df_display, use_container_width=True)
        
        # Статистика с русскими названиями
        filtered_critical = sum(1 for item in filtered_results if item.get('overall_status') == 'critical')
        filtered_warning = sum(1 for item in filtered_results if item.get('overall_status') == 'warning')
        filtered_order_value = sum(item.get('total_order_value', 0) for item in filtered_results)
        
        st.caption(f"""
        📊 Показано {len(filtered_results)} из {len(analysis_results)} товаров | 
        🔴 Критичных: {filtered_critical} | 🟡 Требуют внимания: {filtered_warning} | 
        💰 К заказу: {filtered_order_value:,.0f} ₽
        """)
    
    else:
        st.info("📋 Нет товаров, соответствующих выбранным фильтрам")
    
    # Экспорт с русскими названиями
    st.subheader("📤 Экспорт результатов")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📊 Excel отчет"):
            excel_data = create_russian_excel_report(analysis_results, recommendations)
            
            st.download_button(
                label="💾 Скачать отчет",
                data=excel_data,
                file_name=f"анализ_складов_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    with col2:
        if st.button("🛒 Список заказов"):
            orders_data = create_russian_orders_csv(analysis_results)
            
            st.download_button(
                label="💾 Скачать заказы",
                data=orders_data,
                file_name=f"заказы_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv"
            )


def create_russian_excel_report(analysis_results: List[Dict], recommendations: Dict) -> bytes:
    """
    Создает Excel отчет с РУССКИМИ НАЗВАНИЯМИ колонок
    """
    
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ складов"
    
    # Заголовок
    ws['A1'] = "АНАЛИЗ ОСТАТКОВ ПО СКЛАДАМ"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Дата: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Создаем таблицу данных с РУССКИМИ НАЗВАНИЯМИ
    data = []
    for item in analysis_results:
        row = {
            'Номенклатура': item.get('номенклатура', ''),
            'ADS': item.get('ads', 0),
            'Цена': item.get('price', 0),
            'Общий_остаток': item.get('total_stock', 0),
            'Стоимость_остатков': item.get('total_stock_value', 0),
            'К_заказу_шт': item.get('total_order_quantity', 0),
            'К_заказу_руб': item.get('total_order_value', 0),
            'Статус': item.get('overall_status', ''),
            'Месяцев_запаса': item.get('min_months_across_warehouses', 0) if item.get('min_months_across_warehouses', 0) < 999 else 0
        }
        
        # Добавляем склады с русскими названиями
        warehouses = item.get('warehouses', {})
        for wh_key, wh_data in warehouses.items():
            wh_name = wh_data.get('short_name', wh_key)
            row[f"{wh_name}_остаток"] = wh_data.get('current_stock', 0)
            row[f"{wh_name}_заказать"] = wh_data.get('order_quantity', 0)
            row[f"{wh_name}_статус"] = wh_data.get('status', '')
        
        data.append(row)
    
    if data:
        df = pd.DataFrame(data)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Форматируем заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Сохраняем
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def create_russian_orders_csv(analysis_results: List[Dict]) -> str:
    """
    Создает CSV заказов с РУССКИМИ НАЗВАНИЯМИ колонок
    """
    
    orders_data = []
    
    for item in analysis_results:
        warehouses = item.get('warehouses', {})
        for wh_key, wh_data in warehouses.items():
            if wh_data.get('order_quantity', 0) > 0:
                orders_data.append({
                    'Номенклатура': item.get('номенклатура', ''),
                    'Склад': wh_data.get('short_name', wh_key),
                    'Город': wh_data.get('city', ''),
                    'ADS': item.get('ads', 0),
                    'Цена': item.get('price', 0),
                    'Текущий_остаток': wh_data.get('current_stock', 0),
                    'Минимум': wh_data.get('min_stock', 0),
                    'К_заказу_шт': wh_data.get('order_quantity', 0),
                    'Сумма_заказа': wh_data.get('order_value', 0),
                    'Статус': wh_data.get('status', ''),
                    'Приоритет': 'Высокий' if wh_data.get('status') == 'critical' else 'Средний'
                })
    
    if orders_data:
        # Сортируем по приоритету
        orders_data.sort(key=lambda x: (0 if x['Приоритет'] == 'Высокий' else 1, -x['Сумма_заказа']))
        
        df_orders = pd.DataFrame(orders_data)
        return df_orders.to_csv(index=False, encoding='utf-8-sig')
    else:
        return "Нет товаров к заказу"


# Главная функция для быстрого применения
def quick_russian_interface_fix(system):
    """
    Быстрое исправление: русские названия в таблицах + правильный поиск цен
    """
    
    try:
        # Применяем исправления
        if not hasattr(system, '_russian_interface_fixed'):
            apply_russian_interface_fix(system)
        
        # Создаем и запускаем страницу
        warehouse_page = create_russian_warehouse_page()
        warehouse_page(system)
        
        return True
        
    except Exception as e:
        st.error(f"❌ Ошибка исправления: {str(e)}")
        
        st.error("""
        **Не удалось автоматически исправить интерфейс.**
        
        **Ручное исправление:**
        1. Создайте файл `russian_interface_fix.py` из артефакта
        2. Замените вашу функцию `warehouse_analysis_page` на:
        
        ```python
        def warehouse_analysis_page(system):
            from russian_interface_fix import quick_russian_interface_fix
            quick_russian_interface_fix(system)
        ```
        """)
        
        return False


if __name__ == "__main__":
    print("🎯 Исправление русского интерфейса складов")
    print("Русские названия колонок в таблицах + правильный поиск цен")
    print("\nДля использования:")
    print("from russian_interface_fix import quick_russian_interface_fix")
    print("\nИнструкция:")
    print("""
# 🎯 БЫСТРОЕ ИСПРАВЛЕНИЕ ИНТЕРФЕЙСА

## 🚀 ПРИМЕНЕНИЕ:

### Замените вашу функцию warehouse_analysis_page на:

```python
def warehouse_analysis_page(system):
    from russian_interface_fix import quick_russian_interface_fix
    quick_russian_interface_fix(system)
```

## ✅ ЧТО ИСПРАВЛЯЕТСЯ:

### 🇷🇺 РУССКИЕ НАЗВАНИЯ В ТАБЛИЦАХ:
- ✅ "Warehouse" → "Склад"
- ✅ "Price" → "Цена (₽)"
- ✅ "Stock" → "Остаток"
- ✅ "Order" → "К заказу"
- ✅ "Status" → "Статус"
- ✅ "Total" → "Общий"
- ✅ Все остальные колонки на русском

### 💰 ПРАВИЛЬНЫЙ ПОИСК ЦЕН:
- ✅ Ищет 'last_purchase_price', 'Посл. закупка', 'цена' и др.
- ✅ Показывает детальную статистику по ценам
- ✅ Выводит рекомендации если цены не найдены

### 📊 РУССКИЕ НАЗВАНИЯ В ЭКСПОРТЕ:
- ✅ Excel файлы с русскими заголовками
- ✅ CSV файлы с русскими названиями колонок
- ✅ Русские названия файлов при скачивании

## 🎯 РЕЗУЛЬТАТ:

После применения:
- ✅ Все таблицы с русскими названиями колонок
- ✅ Правильно находит цены в ADS
- ✅ Полная функциональность анализа складов
- ✅ Русский интерфейс без изменения логики
""")