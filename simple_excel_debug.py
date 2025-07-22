#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Простой анализатор Excel файла без pandas
"""

import openpyxl
import sys
import os

def analyze_excel_file():
    """Анализ Excel файла"""
    
    filename = 'остатки на 08.07.2025.xlsx'
    
    if not os.path.exists(filename):
        print(f"❌ Файл {filename} не найден")
        return
    
    try:
        # Открываем файл
        workbook = openpyxl.load_workbook(filename, data_only=True)
        sheet = workbook.active
        
        print("=== РЕАЛЬНЫЙ АНАЛИЗ ФАЙЛА ОСТАТКОВ ===")
        print(f"Файл: {filename}")
        print(f"Размер: {sheet.max_row} строк, {sheet.max_column} колонок")
        print()
        
        print("=== ПЕРВЫЕ 20 СТРОК ===")
        for row_num in range(1, min(21, sheet.max_row + 1)):
            print(f"Строка {row_num}:")
            row_data = []
            for col_num in range(1, min(15, sheet.max_column + 1)):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    row_data.append(f"Col{col_num}: {cell.value}")
            
            if row_data:
                print(f"  {' | '.join(row_data)}")
            else:
                print("  (пустая строка)")
            print()
        
        print("\n=== ПОИСК НОМЕНКЛАТУРЫ ===")
        for row_num in range(1, min(30, sheet.max_row + 1)):
            for col_num in range(1, min(10, sheet.max_column + 1)):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value and 'номенклатура' in str(cell.value).lower():
                    print(f"НАЙДЕНА 'номенклатура' в строке {row_num}, колонке {col_num}: {cell.value}")
        
        print("\n=== АНАЛИЗ ЗАГОЛОВКОВ СКЛАДОВ ===")
        # Проверяем строки 5-10 на наличие названий складов
        for row_num in range(5, min(11, sheet.max_row + 1)):
            print(f"Строка {row_num}:")
            for col_num in range(1, min(12, sheet.max_column + 1)):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value:
                    value_str = str(cell.value)
                    if any(word in value_str.lower() for word in ['склад', 'магазин', 'хаб', 'казыбаева', 'астана', 'шымкент']):
                        print(f"  Col{col_num}: {value_str}")
        
        print("\n=== ПОИСК ДАННЫХ ТОВАРОВ ===")
        # Ищем строки с данными товаров
        for row_num in range(8, min(30, sheet.max_row + 1)):
            first_cell = sheet.cell(row=row_num, column=1)
            if first_cell.value and len(str(first_cell.value)) > 5:
                # Проверяем есть ли числа в других колонках
                numbers = []
                for col_num in range(2, min(10, sheet.max_column + 1)):
                    cell = sheet.cell(row=row_num, column=col_num)
                    if isinstance(cell.value, (int, float)) and cell.value > 0:
                        numbers.append(f"Col{col_num}:{cell.value}")
                
                if len(numbers) >= 1:
                    print(f"Строка {row_num}: {str(first_cell.value)[:40]}...")
                    print(f"  Остатки: {' | '.join(numbers)}")
                    break  # Показываем только первый пример
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_excel_file()