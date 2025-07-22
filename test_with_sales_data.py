"""
Тест системы с реальными данными продаж
"""

import openpyxl
import json
from collections import defaultdict

def analyze_sales_structure():
    """Анализируем структуру файла продаж"""
    print("📊 Анализ структуры файла продаж...")
    
    try:
        # Открываем Excel файл
        wb = openpyxl.load_workbook('общ_продажи_по_всем_складам_с_01_07_2024_01_07_2025_гг.xlsx')
        sheet = wb.active
        
        print(f"✅ Файл продаж загружен")
        print(f"   Размер: {sheet.max_row} строк, {sheet.max_column} колонок")
        
        # Анализируем заголовки
        headers = []
        for col in range(1, min(sheet.max_column + 1, 20)):  # Первые 20 колонок
            cell_value = sheet.cell(row=3, column=col).value  # Строка 3 как в других файлах
            if cell_value:
                headers.append(str(cell_value).strip())
        
        print(f"   Найдено заголовков: {len(headers)}")
        
        # Ищем колонки с артикулами и продажами
        article_cols = [i for i, h in enumerate(headers) if 'артикул' in h.lower()]
        sales_cols = [i for i, h in enumerate(headers) if 'прод' in h.lower() or 'сумм' in h.lower()]
        
        print(f"   Колонки с артикулами: {[headers[i] for i in article_cols]}")
        print(f"   Колонки с продажами: {len(sales_cols)}")
        
        if sales_cols:
            print("   Примеры колонок продаж:")
            for i in sales_cols[:5]:
                print(f"     - {headers[i]}")
        
        # Анализируем данные
        print(f"\n📦 Анализ данных продаж...")
        
        sales_data = defaultdict(lambda: defaultdict(float))
        processed_rows = 0
        
        for row in range(4, min(sheet.max_row + 1, 1000)):  # Первые 1000 строк
            article_cell = sheet.cell(row=row, column=article_cols[0] + 1) if article_cols else None
            
            if article_cell and article_cell.value:
                article = str(article_cell.value).strip()
                
                # Собираем продажи по складам
                for col_idx in sales_cols:
                    sales_cell = sheet.cell(row=row, column=col_idx + 1)
                    if sales_cell.value and isinstance(sales_cell.value, (int, float)):
                        warehouse_name = headers[col_idx]
                        sales_data[article][warehouse_name] += float(sales_cell.value)
                
                processed_rows += 1
        
        print(f"   Обработано строк: {processed_rows}")
        print(f"   Уникальных артикулов: {len(sales_data)}")
        
        return sales_data, headers
        
    except Exception as e:
        print(f"❌ Ошибка анализа продаж: {e}")
        return None, None

def test_integration_with_stock():
    """Тест интеграции данных продаж и остатков"""
    print("\n🔗 Тест интеграции остатков и продаж...")
    
    # Загружаем остатки
    try:
        with open('2025-06-30 (4).json', 'r', encoding='utf-8-sig') as f:
            stock_data = json.load(f)
        print("✅ Остатки загружены")
    except Exception as e:
        print(f"❌ Ошибка загрузки остатков: {e}")
        return False
    
    # Загружаем продажи
    sales_data, headers = analyze_sales_structure()
    if not sales_data:
        return False
    
    # Анализируем пересечения
    stock_articles = set()
    for warehouse in stock_data['ОстаткиПоСкладам']:
        for item in warehouse['Остатки']:
            stock_articles.add(item['Артикул'])
    
    sales_articles = set(sales_data.keys())
    
    intersection = stock_articles & sales_articles
    
    print(f"📊 Статистика пересечений:")
    print(f"   Артикулов в остатках: {len(stock_articles):,}")
    print(f"   Артикулов в продажах: {len(sales_articles):,}")
    print(f"   Пересечение: {len(intersection):,}")
    print(f"   Покрытие: {len(intersection)/len(stock_articles)*100:.1f}%")
    
    # Анализируем топ товары по продажам
    if intersection:
        print(f"\n🏆 Топ-10 товаров по продажам (из пересечения):")
        
        article_total_sales = {}
        for article in intersection:
            total_sales = sum(sales_data[article].values())
            if total_sales > 0:
                article_total_sales[article] = total_sales
        
        top_sales = sorted(article_total_sales.items(), key=lambda x: x[1], reverse=True)[:10]
        
        for i, (article, total_sales) in enumerate(top_sales, 1):
            # Найдем название товара из остатков
            item_name = "Неизвестно"
            for warehouse in stock_data['ОстаткиПоСкладам']:
                for item in warehouse['Остатки']:
                    if item['Артикул'] == article:
                        item_name = item['Номенклатура']
                        break
                if item_name != "Неизвестно":
                    break
            
            print(f"   {i:2d}. {article}")
            print(f"       {item_name[:60]}")
            print(f"       Продажи: {total_sales:,.0f}")
    
    # Генерируем простые рекомендации
    print(f"\n🚚 Генерация рекомендаций...")
    
    recommendations = []
    
    # Ищем товары с дисбалансом остатков vs продаж
    for article in list(intersection)[:500]:  # Первые 500 для теста
        # Собираем данные по складам для этого товара
        stock_by_warehouse = {}
        for warehouse in stock_data['ОстаткиПоСкладам']:
            wh_name = warehouse['Склад']
            for item in warehouse['Остатки']:
                if item['Артикул'] == article:
                    stock_by_warehouse[wh_name] = {
                        'qty': item['Количество'],
                        'cost': item['Стоимость'],
                        'name': item['Номенклатура']
                    }
        
        # Собираем продажи по складам
        sales_by_warehouse = sales_data.get(article, {})
        
        # Анализируем дисбаланс
        if len(stock_by_warehouse) > 1:  # Товар на нескольких складах
            for wh1, stock1 in stock_by_warehouse.items():
                for wh2, stock2 in stock_by_warehouse.items():
                    if wh1 != wh2:
                        # Ищем соответствующие продажи
                        sales1 = 0
                        sales2 = 0
                        
                        for sales_col, sales_val in sales_by_warehouse.items():
                            if any(word in sales_col.lower() for word in wh1.lower().split()):
                                sales1 += sales_val
                            elif any(word in sales_col.lower() for word in wh2.lower().split()):
                                sales2 += sales_val
                        
                        # Если есть дисбаланс: много остатков + мало продаж vs мало остатков + много продаж
                        total_stock = stock1['qty'] + stock2['qty']
                        total_sales = sales1 + sales2
                        
                        if (total_sales > 0 and 
                            stock1['qty'] > total_stock * 0.7 and sales1 < total_sales * 0.3 and
                            stock2['qty'] < total_stock * 0.3 and sales2 > total_sales * 0.7):
                            
                            move_qty = min(stock1['qty'] * 0.3, total_stock * 0.2)
                            
                            if move_qty >= 10:  # Минимум 10 единиц
                                recommendations.append({
                                    'article': article,
                                    'name': stock1['name'],
                                    'from': wh1,
                                    'to': wh2,
                                    'qty': int(move_qty),
                                    'from_stock': stock1['qty'],
                                    'to_stock': stock2['qty'],
                                    'from_sales': sales1,
                                    'to_sales': sales2,
                                    'priority': 'Высокий' if sales2 > sales1 * 3 else 'Средний'
                                })
    
    print(f"   Сгенерировано рекомендаций: {len(recommendations)}")
    
    # Показываем топ-10 рекомендаций
    recommendations.sort(key=lambda x: x['to_sales'], reverse=True)
    
    for i, rec in enumerate(recommendations[:10], 1):
        print(f"   {i:2d}. {rec['article']} - {rec['priority']}")
        print(f"       {rec['name'][:50]}")
        print(f"       {rec['from'][:30]} -> {rec['to'][:30]}")
        print(f"       К перемещению: {rec['qty']}")
        print(f"       Остатки: {rec['from_stock']} -> {rec['to_stock']}")
        print(f"       Продажи: {rec['from_sales']:.0f} -> {rec['to_sales']:.0f}")
    
    print(f"\n✅ Интеграция протестирована успешно!")
    return True

if __name__ == "__main__":
    test_integration_with_stock()