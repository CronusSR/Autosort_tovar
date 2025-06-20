# warehouse_complete_fix.py
"""
🔧 ПОЛНОЕ ИСПРАВЛЕНИЕ АНАЛИЗА СКЛАДОВ
Исправляет все ошибки связанные с анализом остатков по складам
"""

import pandas as pd
import streamlit as st
import json
from typing import Dict, List, Any, Optional, Tuple


class WarehouseFileReader:
    """
    Класс для правильного чтения файлов остатков
    """
    
    def __init__(self):
        self.debug_mode = True
        
    def read_remains_file_smart(self, uploaded_file) -> pd.DataFrame:
        """
        ИСПРАВЛЕННОЕ чтение файла остатков с правильной обработкой структуры
        """
        try:
            if self.debug_mode:
                st.info("🔍 Начинаем умное чтение файла остатков...")
            
            # Читаем файл
            if uploaded_file.name.endswith('.xlsx'):
                raw_data = pd.read_excel(uploaded_file, header=None)
            else:
                raw_data = pd.read_excel(uploaded_file, engine='xlrd', header=None)
            
            if self.debug_mode:
                st.write(f"📊 Размер файла: {raw_data.shape[0]} строк, {raw_data.shape[1]} колонок")
            
            # Ищем строку с номенклатурой
            nomenclature_row = None
            for i, row in raw_data.iterrows():
                if any(cell and 'номенклатура' in str(cell).lower() for cell in row if pd.notna(cell)):
                    nomenclature_row = i
                    break
            
            if nomenclature_row is None:
                raise ValueError("❌ Не найдена строка с заголовком 'Номенклатура'")
            
            if self.debug_mode:
                st.success(f"✅ Найдена номенклатура в строке {nomenclature_row + 1}")
            
            # Извлекаем заголовки складов
            header_row = raw_data.iloc[nomenclature_row]
            warehouse_columns = []
            nomenclature_col = None
            
            for j, cell in enumerate(header_row):
                if pd.notna(cell):
                    cell_str = str(cell).strip()
                    if 'номенклатура' in cell_str.lower():
                        nomenclature_col = j
                    elif 'склад' in cell_str.lower() or 'магазин' in cell_str.lower():
                        warehouse_columns.append({
                            'index': j,
                            'name': cell_str,
                            'short_name': self._create_short_name(cell_str)
                        })
            
            if nomenclature_col is None:
                raise ValueError("❌ Не найдена колонка номенклатуры")
            
            if self.debug_mode:
                st.write(f"📍 Номенклатура в колонке {nomenclature_col}")
                st.write(f"🏪 Найдено складов: {len(warehouse_columns)}")
                for wh in warehouse_columns:
                    st.write(f"  - {wh['name']} (колонка {wh['index']})")
            
            # Ищем строку начала данных (после заголовков)
            data_start_row = nomenclature_row + 1
            
            # Проверяем есть ли дополнительные строки заголовков
            for i in range(nomenclature_row + 1, min(nomenclature_row + 4, len(raw_data))):
                row = raw_data.iloc[i]
                if any(cell and ('количество' in str(cell).lower() or 'остаток' in str(cell).lower()) 
                       for cell in row if pd.notna(cell)):
                    data_start_row = i + 1
                    break
            
            if self.debug_mode:
                st.write(f"📋 Данные начинаются с строки {data_start_row + 1}")
            
            # Читаем данные товаров
            items_data = []
            processed_count = 0
            
            for i in range(data_start_row, len(raw_data)):
                row = raw_data.iloc[i]
                
                # Проверяем что есть номенклатура
                nomenclature = row.iloc[nomenclature_col] if nomenclature_col < len(row) else None
                if pd.isna(nomenclature) or not str(nomenclature).strip():
                    continue
                
                nomenclature = str(nomenclature).strip()
                
                # Создаем запись товара
                item_data = {'номенклатура': nomenclature}
                
                # Добавляем остатки по складам
                total_stock = 0
                for wh in warehouse_columns:
                    col_idx = wh['index']
                    stock_value = 0
                    
                    if col_idx < len(row):
                        cell_value = row.iloc[col_idx]
                        if pd.notna(cell_value):
                            try:
                                stock_value = float(cell_value)
                                total_stock += stock_value
                            except (ValueError, TypeError):
                                stock_value = 0
                    
                    # Используем короткое имя склада для колонки
                    column_name = f"{wh['short_name']}_остаток"
                    item_data[column_name] = stock_value
                
                # Добавляем итоговый остаток
                item_data['итого_остаток'] = total_stock
                
                items_data.append(item_data)
                processed_count += 1
            
            if not items_data:
                raise ValueError("❌ Не найдено товаров с данными")
            
            # Создаем DataFrame
            result_df = pd.DataFrame(items_data)
            
            if self.debug_mode:
                st.success(f"✅ Обработано товаров: {processed_count}")
                st.write(f"📊 Колонки в результате: {list(result_df.columns)}")
                
                # Показываем превью
                with st.expander("👀 Превью обработанных данных"):
                    st.dataframe(result_df.head(5))
            
            return result_df
            
        except Exception as e:
            st.error(f"❌ Ошибка чтения файла: {str(e)}")
            if self.debug_mode:
                st.exception(e)
            return pd.DataFrame()
    
    def _create_short_name(self, full_name: str) -> str:
        """Создает короткое имя склада"""
        name = full_name.lower()
        
        if 'шымкент' in name and 'склад' in name:
            return 'Шымкент_Склад'
        elif 'шымкент' in name and 'магазин' in name:
            return 'Шымкент_Магазин'
        elif 'астана' in name and 'склад' in name:
            return 'Астана_Склад'
        elif 'астана' in name and 'магазин' in name:
            return 'Астана_Магазин'
        elif 'алматы' in name or 'ао склад' in name:
            return 'Алматы_Склад'
        else:
            # Создаем имя из первых значимых слов
            words = [w for w in full_name.split() if len(w) > 2]
            return '_'.join(words[:2]) if words else 'Неизвестный_Склад'


class WarehouseAnalyzer:
    """
    ИСПРАВЛЕННЫЙ класс для анализа складов
    """
    
    def __init__(self):
        self.warehouse_config = {
            'Шымкент_Склад': {
                'name': '4 Склад фурнитуры АЗМ Шымкент "Овощная база"',
                'short_name': 'Шымкент Склад',
                'city': 'шымкент',
                'type': 'warehouse',
                'min_days': 15,
                'max_days': 45
            },
            'Шымкент_Магазин': {
                'name': '6 Склад фурнитуры "Овощная база" Магазин',
                'short_name': 'Шымкент Магазин',
                'city': 'шымкент',
                'type': 'store',
                'min_days': 10,
                'max_days': 30
            },
            'Алматы_Склад': {
                'name': 'АО Склад Фурнитура TRADE',
                'short_name': 'Алматы Склад',
                'city': 'алматы',
                'type': 'specialized_store',
                'min_days': 20,
                'max_days': 60
            },
            'Астана_Склад': {
                'name': 'склад фурнитура № 1',
                'short_name': 'Астана Склад',
                'city': 'астана',
                'type': 'warehouse',
                'min_days': 15,
                'max_days': 45
            },
            'Астана_Магазин': {
                'name': 'Магазин фурнитуры',
                'short_name': 'Астана Магазин',
                'city': 'астана',
                'type': 'store',
                'min_days': 10,
                'max_days': 30
            }
        }
    
    def analyze_warehouse_stock_detailed(self, remains_df: pd.DataFrame, ads_data: pd.DataFrame, 
                                        store_ads_by_city: Dict = None, min_days: int = 10, 
                                        max_days: int = 50) -> List[Dict]:
        """
        ИСПРАВЛЕННЫЙ детальный анализ складов
        """
        try:
            if remains_df.empty:
                st.error("❌ Файл остатков пуст")
                return []
            
            if 'номенклатура' not in remains_df.columns:
                st.error("❌ В файле остатков нет колонки 'номенклатура'")
                return []
            
            st.info(f"🔄 Анализируем {len(remains_df)} товаров...")
            
            analysis_results = []
            processed_count = 0
            
            # Определяем какие склады есть в данных
            available_warehouses = []
            for col in remains_df.columns:
                if col.endswith('_остаток'):
                    warehouse_key = col.replace('_остаток', '')
                    if warehouse_key in self.warehouse_config:
                        available_warehouses.append(warehouse_key)
            
            st.write(f"🏪 Найдены склады: {', '.join([self.warehouse_config[w]['short_name'] for w in available_warehouses])}")
            
            # Проверяем наличие цен в ADS
            has_prices = False
            price_column = None
            if ads_data is not None and not ads_data.empty:
                price_columns = ['last_purchase_price', 'цена', 'price', 'стоимость']
                for col in price_columns:
                    if col in ads_data.columns:
                        has_prices = True
                        price_column = col
                        break
            
            if has_prices:
                st.success(f"💰 Цены найдены в колонке '{price_column}'")
            else:
                st.warning("⚠️ Цены не найдены в ADS данных")
            
            # Анализируем каждый товар
            for idx, item in remains_df.iterrows():
                item_name = str(item['номенклатура']).strip()
                
                if not item_name or item_name.lower() in ['nan', 'none', '']:
                    continue
                
                # Получаем ADS и цену
                ads_value = 0
                item_price = 0
                
                if ads_data is not None and not ads_data.empty:
                    ads_match = ads_data[ads_data['номенклатура'] == item_name]
                    if not ads_match.empty:
                        ads_value = float(ads_match.iloc[0].get('ads', 0))
                        if has_prices and price_column:
                            try:
                                item_price = float(ads_match.iloc[0].get(price_column, 0))
                            except (ValueError, TypeError):
                                item_price = 0
                
                # Общие характеристики товара
                total_stock = float(item.get('итого_остаток', 0))
                
                # Анализ по складам
                warehouses_analysis = {}
                
                for warehouse_key in available_warehouses:
                    config = self.warehouse_config[warehouse_key]
                    stock_col = f"{warehouse_key}_остаток"
                    current_stock = float(item.get(stock_col, 0))
                    
                    # Персональные настройки для склада
                    wh_min_days = config.get('min_days', min_days)
                    wh_max_days = config.get('max_days', max_days)
                    
                    # Расчеты
                    min_stock = ads_value * wh_min_days if ads_value > 0 else 0
                    max_stock = ads_value * wh_max_days if ads_value > 0 else 0
                    
                    # Дефицит/избыток
                    min_deficit = max(0, min_stock - current_stock)
                    max_deficit = max(0, max_stock - current_stock)
                    surplus = max(0, current_stock - max_stock)
                    
                    # Месяцы запаса
                    months_of_stock = 0
                    if ads_value > 0:
                        days_of_stock = current_stock / ads_value
                        months_of_stock = days_of_stock / 30
                    elif current_stock > 0:
                        months_of_stock = 999  # Бесконечно (нет продаж)
                    
                    # Статус
                    if current_stock < min_stock and ads_value > 0:
                        status = 'critical' if min_deficit > ads_value * 5 else 'warning'
                    elif current_stock > max_stock:
                        status = 'excess'
                    elif ads_value == 0 and current_stock > 0:
                        status = 'no_sales'
                    else:
                        status = 'good'
                    
                    # Рекомендация к заказу
                    order_quantity = min_deficit if min_deficit > 0 else 0
                    
                    # Денежные расчеты
                    stock_value = current_stock * item_price if item_price > 0 else 0
                    order_value = order_quantity * item_price if item_price > 0 else 0
                    
                    warehouses_analysis[warehouse_key] = {
                        'warehouse_name': config['name'],
                        'short_name': config['short_name'],
                        'current_stock': current_stock,
                        'min_stock': min_stock,
                        'max_stock': max_stock,
                        'min_deficit': min_deficit,
                        'max_deficit': max_deficit,
                        'surplus': surplus,
                        'months_of_stock': months_of_stock,
                        'status': status,
                        'order_quantity': order_quantity,
                        'stock_value': stock_value,
                        'order_value': order_value,
                        'days_settings': f"{wh_min_days}-{wh_max_days}"
                    }
                
                # Общий статус товара
                critical_warehouses = sum(1 for w in warehouses_analysis.values() if w['status'] == 'critical')
                warning_warehouses = sum(1 for w in warehouses_analysis.values() if w['status'] == 'warning')
                
                if critical_warehouses > 0:
                    overall_status = 'critical'
                elif warning_warehouses > 0:
                    overall_status = 'warning'
                else:
                    overall_status = 'good'
                
                # Минимальные месяцы среди всех складов
                months_list = [w['months_of_stock'] for w in warehouses_analysis.values() 
                              if w['months_of_stock'] < 999]
                min_months = min(months_list) if months_list else 0
                
                analysis_results.append({
                    'номенклатура': item_name,
                    'total_stock': total_stock,
                    'ads': ads_value,
                    'price': item_price,
                    'total_stock_value': total_stock * item_price if item_price > 0 else 0,
                    'min_months_across_warehouses': min_months,
                    'overall_status': overall_status,
                    'critical_warehouses_count': critical_warehouses,
                    'warning_warehouses_count': warning_warehouses,
                    'warehouses': warehouses_analysis,
                    'analysis_parameters': {
                        'global_min_days': min_days,
                        'global_max_days': max_days,
                        'has_prices': has_prices,
                        'price_column': price_column
                    }
                })
                
                processed_count += 1
            
            st.success(f"✅ Анализ завершен! Обработано {processed_count} товаров")
            return analysis_results
            
        except Exception as e:
            st.error(f"❌ Ошибка анализа: {str(e)}")
            st.exception(e)
            return []
    
    def get_warehouse_recommendations(self, analysis_results: List[Dict] = None) -> Dict:
        """
        Получение рекомендаций по складам
        """
        if not analysis_results:
            return {}
        
        # Агрегируем данные по складам
        warehouse_summary = {}
        
        for warehouse_key, config in self.warehouse_config.items():
            warehouse_summary[warehouse_key] = {
                'name': config['short_name'],
                'city': config['city'],
                'total_items': 0,
                'critical_items': 0,
                'warning_items': 0,
                'excess_items': 0,
                'no_sales_items': 0,
                'total_order_quantity': 0,
                'total_order_value': 0,
                'total_stock_value': 0
            }
        
        # Обрабатываем результаты анализа
        for item in analysis_results:
            for warehouse_key, wh_data in item['warehouses'].items():
                if warehouse_key in warehouse_summary:
                    summary = warehouse_summary[warehouse_key]
                    summary['total_items'] += 1
                    
                    if wh_data['status'] == 'critical':
                        summary['critical_items'] += 1
                    elif wh_data['status'] == 'warning':
                        summary['warning_items'] += 1
                    elif wh_data['status'] == 'excess':
                        summary['excess_items'] += 1
                    elif wh_data['status'] == 'no_sales':
                        summary['no_sales_items'] += 1
                    
                    summary['total_order_quantity'] += wh_data.get('order_quantity', 0)
                    summary['total_order_value'] += wh_data.get('order_value', 0)
                    summary['total_stock_value'] += wh_data.get('stock_value', 0)
        
        return warehouse_summary


def apply_warehouse_complete_fix(system):
    """
    Применяет полное исправление к системе анализа складов
    """
    try:
        # Добавляем исправленные компоненты
        system.warehouse_file_reader = WarehouseFileReader()
        system.warehouse_analyzer = WarehouseAnalyzer()
        
        # Добавляем исправленные методы
        def analyze_warehouse_stock_with_details(remains_df, ads_data, store_ads_by_city=None, 
                                               min_days=10, max_days=50):
            """Исправленный метод анализа складов"""
            return system.warehouse_analyzer.analyze_warehouse_stock_detailed(
                remains_df, ads_data, store_ads_by_city, min_days, max_days
            )
        
        def get_warehouse_recommendations(analysis_results=None):
            """Исправленный метод получения рекомендаций"""
            return system.warehouse_analyzer.get_warehouse_recommendations(analysis_results)
        
        # Привязываем методы к системе
        system.analyze_warehouse_stock_with_details = analyze_warehouse_stock_with_details
        system.get_warehouse_recommendations = get_warehouse_recommendations
        
        # Отмечаем что исправления применены
        system._warehouse_fix_applied = True
        
        return True
        
    except Exception as e:
        st.error(f"❌ Ошибка применения исправлений: {str(e)}")
        return False


def create_fixed_warehouse_analysis_page():
    """
    ПОЛНОСТЬЮ ИСПРАВЛЕННАЯ страница анализа складов
    """
    
    def warehouse_analysis_page_fixed(system):
        """
        Исправленная страница анализа остатков по складам
        """
        
        st.header("📦 Анализ остатков по складам")
        
        # Применяем исправления
        if not hasattr(system, '_warehouse_fix_applied'):
            with st.spinner("🔧 Применяем исправления системы..."):
                success = apply_warehouse_complete_fix(system)
                if success:
                    st.success("✅ Система анализа складов обновлена!")
                else:
                    st.error("❌ Не удалось применить исправления")
                    return
        
        # Проверяем наличие ADS данных
        if not hasattr(system, 'calculated_ads') or system.calculated_ads is None:
            st.warning("⚠️ Сначала рассчитайте ADS в разделе 'ADS расчет'")
            return
        
        st.success(f"✅ ADS данные готовы: {len(system.calculated_ads)} товаров")
        
        # Настройки анализа
        st.subheader("⚙️ Параметры анализа")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            min_days = st.number_input("Минимальные дни:", min_value=5, max_value=60, value=15)
        with col2:
            max_days = st.number_input("Максимальные дни:", min_value=20, max_value=120, value=45)
        with col3:
            debug_mode = st.checkbox("Режим отладки", value=True)
        
        # Загрузка файла остатков
        st.subheader("📂 Загрузка файла остатков")
        
        uploaded_file = st.file_uploader(
            "Выберите файл остатков:",
            type=['xlsx', 'xls'],
            help="Файл должен содержать колонку 'Номенклатура' и колонки остатков по складам"
        )
        
        if uploaded_file:
            # Используем исправленный ридер файлов
            system.warehouse_file_reader.debug_mode = debug_mode
            
            with st.spinner("📖 Читаем и обрабатываем файл остатков..."):
                remains_df = system.warehouse_file_reader.read_remains_file_smart(uploaded_file)
            
            if remains_df.empty:
                st.error("❌ Не удалось прочитать файл остатков")
                return
            
            # Показываем статистику файла
            st.success(f"✅ Файл успешно обработан!")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Товаров", len(remains_df))
            with col2:
                warehouse_cols = [col for col in remains_df.columns if col.endswith('_остаток')]
                st.metric("Складов", len(warehouse_cols))
            with col3:
                total_stock = remains_df['итого_остаток'].sum()
                st.metric("Общий остаток", f"{total_stock:,.0f}")
            with col4:
                non_zero_items = (remains_df['итого_остаток'] > 0).sum()
                st.metric("С остатками", non_zero_items)
            
            # Показываем найденные склады
            with st.expander("🏪 Найденные склады"):
                warehouse_cols = [col for col in remains_df.columns if col.endswith('_остаток')]
                for col in warehouse_cols:
                    warehouse_key = col.replace('_остаток', '')
                    config = system.warehouse_analyzer.warehouse_config.get(warehouse_key, {})
                    stock_sum = remains_df[col].sum()
                    items_with_stock = (remains_df[col] > 0).sum()
                    
                    st.write(f"**{config.get('short_name', warehouse_key)}**")
                    st.write(f"  - Общий остаток: {stock_sum:,.0f}")
                    st.write(f"  - Товаров с остатками: {items_with_stock}")
                    st.write(f"  - Настройки: {config.get('min_days', 15)}-{config.get('max_days', 45)} дней")
            
            # Кнопка запуска анализа
            st.subheader("🚀 Запуск анализа")
            
            if st.button("🔍 Запустить детальный анализ складов", type="primary"):
                
                with st.spinner("🔄 Выполняем детальный анализ..."):
                    
                    # Запускаем исправленный анализ
                    analysis_results = system.analyze_warehouse_stock_with_details(
                        remains_df, 
                        system.calculated_ads, 
                        None,  # store_ads_by_city - пока не используется
                        min_days, 
                        max_days
                    )
                    
                    if not analysis_results:
                        st.error("❌ Анализ не дал результатов")
                        return
                    
                    # Сохраняем результаты в системе
                    system.warehouse_analysis_results = analysis_results
                    
                    # Получаем рекомендации
                    recommendations = system.get_warehouse_recommendations(analysis_results)
                    system.warehouse_recommendations = recommendations
                    
                    st.success(f"✅ Анализ завершен! Обработано {len(analysis_results)} товаров")
                
                # Показываем результаты
                show_warehouse_analysis_results(analysis_results, recommendations)
        
        # Если есть сохраненные результаты, показываем их
        if hasattr(system, 'warehouse_analysis_results') and system.warehouse_analysis_results:
            st.markdown("---")
            st.subheader("📊 Последние результаты анализа")
            
            if st.button("🔄 Показать результаты"):
                show_warehouse_analysis_results(
                    system.warehouse_analysis_results, 
                    getattr(system, 'warehouse_recommendations', {})
                )
    
    return warehouse_analysis_page_fixed


def show_warehouse_analysis_results(analysis_results: List[Dict], recommendations: Dict):
    """
    Показывает результаты анализа складов
    """
    
    st.subheader("📈 Результаты анализа складов")
    
    # Общая статистика
    total_items = len(analysis_results)
    critical_items = sum(1 for item in analysis_results if item['overall_status'] == 'critical')
    warning_items = sum(1 for item in analysis_results if item['overall_status'] == 'warning')
    good_items = sum(1 for item in analysis_results if item['overall_status'] == 'good')
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Всего товаров", total_items)
    with col2:
        st.metric("Критичных", critical_items, delta=f"-{critical_items/total_items*100:.1f}%")
    with col3:
        st.metric("Требуют внимания", warning_items, delta=f"{warning_items/total_items*100:.1f}%")
    with col4:
        st.metric("В норме", good_items, delta=f"+{good_items/total_items*100:.1f}%")
    
    # Статистика по складам
    if recommendations:
        st.subheader("🏪 Статистика по складам")
        
        warehouse_df = pd.DataFrame.from_dict(recommendations, orient='index')
        warehouse_df.reset_index(inplace=True)
        warehouse_df.rename(columns={'index': 'warehouse_key'}, inplace=True)
        
        # Форматируем для отображения
        display_df = warehouse_df.copy()
        display_df['total_order_value'] = display_df['total_order_value'].apply(lambda x: f"{x:,.0f} ₽")
        display_df['total_stock_value'] = display_df['total_stock_value'].apply(lambda x: f"{x:,.0f} ₽")
        
        st.dataframe(
            display_df[['name', 'city', 'critical_items', 'warning_items', 'total_order_value', 'total_stock_value']],
            column_config={
                'name': 'Склад',
                'city': 'Город',
                'critical_items': 'Критичных',
                'warning_items': 'Требуют внимания',
                'total_order_value': 'К заказу (₽)',
                'total_stock_value': 'Стоимость остатков'
            },
            use_container_width=True
        )
    
    # Детальная таблица товаров
    st.subheader("📋 Детальная информация по товарам")
    
    # Фильтры
    col1, col2, col3 = st.columns(3)
    with col1:
        status_filter = st.selectbox(
            "Фильтр по статусу:",
            ["Все", "Критичные", "Требуют внимания", "В норме"],
            index=0
        )
    with col2:
        sort_by = st.selectbox(
            "Сортировать по:",
            ["Алфавиту", "Критичности", "ADS", "Остаткам"],
            index=1
        )
    with col3:
        show_prices = st.checkbox("Показать цены", value=True)
    
    # Фильтруем данные
    filtered_results = analysis_results.copy()
    
    if status_filter == "Критичные":
        filtered_results = [item for item in filtered_results if item['overall_status'] == 'critical']
    elif status_filter == "Требуют внимания":
        filtered_results = [item for item in filtered_results if item['overall_status'] == 'warning']
    elif status_filter == "В норме":
        filtered_results = [item for item in filtered_results if item['overall_status'] == 'good']
    
    # Сортируем
    if sort_by == "Критичности":
        status_order = {'critical': 0, 'warning': 1, 'good': 2}
        filtered_results.sort(key=lambda x: (status_order.get(x['overall_status'], 3), -x['ads']))
    elif sort_by == "ADS":
        filtered_results.sort(key=lambda x: -x['ads'])
    elif sort_by == "Остаткам":
        filtered_results.sort(key=lambda x: -x['total_stock'])
    else:  # Алфавиту
        filtered_results.sort(key=lambda x: x['номенклатура'])
    
    # Создаем таблицу для отображения
    display_data = []
    
    for item in filtered_results:
        # Определяем статус эмодзи
        status_emoji = {
            'critical': '🔴',
            'warning': '🟡', 
            'good': '🟢'
        }.get(item['overall_status'], '⚪')
        
        # Базовая информация
        row_data = {
            'Статус': status_emoji,
            'Номенклатура': item['номенклатура'][:50] + "..." if len(item['номенклатура']) > 50 else item['номенклатура'],
            'ADS': f"{item['ads']:.2f}",
            'Общий остаток': f"{item['total_stock']:.0f}",
            'Месяцев запаса': f"{item['min_months_across_warehouses']:.1f}"
        }
        
        # Добавляем цены если нужно
        if show_prices and item['price'] > 0:
            row_data['Цена'] = f"{item['price']:.2f} ₽"
            row_data['Стоимость остатков'] = f"{item['total_stock_value']:,.0f} ₽"
        
        # Добавляем информацию по складам
        for warehouse_key, wh_data in item['warehouses'].items():
            warehouse_name = wh_data['short_name']
            current_stock = wh_data['current_stock']
            order_qty = wh_data['order_quantity']
            
            if order_qty > 0:
                row_data[f"{warehouse_name} (остаток)"] = f"{current_stock:.0f} ❗ +{order_qty:.0f}"
            else:
                row_data[f"{warehouse_name} (остаток)"] = f"{current_stock:.0f}"
        
        display_data.append(row_data)
    
    if display_data:
        st.dataframe(pd.DataFrame(display_data), use_container_width=True)
        
        # Экспорт результатов
        st.subheader("📤 Экспорт результатов")
        
        if st.button("📊 Создать Excel отчет"):
            excel_data = create_warehouse_excel_report(analysis_results, recommendations)
            
            st.download_button(
                label="💾 Скачать детальный отчет",
                data=excel_data,
                file_name=f"warehouse_analysis_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("📋 Нет данных для отображения с выбранными фильтрами")


def create_warehouse_excel_report(analysis_results: List[Dict], recommendations: Dict) -> bytes:
    """
    Создает Excel отчет по анализу складов
    """
    
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Создаем workbook
    wb = Workbook()
    
    # Лист 1: Сводка
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    
    # Заголовок
    ws_summary['A1'] = "АНАЛИЗ ОСТАТКОВ ПО СКЛАДАМ"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A2'] = f"Дата: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Общая статистика
    ws_summary['A4'] = "ОБЩАЯ СТАТИСТИКА"
    ws_summary['A4'].font = Font(size=12, bold=True)
    
    total_items = len(analysis_results)
    critical_items = sum(1 for item in analysis_results if item['overall_status'] == 'critical')
    warning_items = sum(1 for item in analysis_results if item['overall_status'] == 'warning')
    
    ws_summary['A5'] = "Всего товаров:"
    ws_summary['B5'] = total_items
    ws_summary['A6'] = "Критичных:"
    ws_summary['B6'] = critical_items
    ws_summary['A7'] = "Требуют внимания:"
    ws_summary['B7'] = warning_items
    
    # Статистика по складам
    if recommendations:
        ws_summary['A9'] = "СТАТИСТИКА ПО СКЛАДАМ"
        ws_summary['A9'].font = Font(size=12, bold=True)
        
        headers = ['Склад', 'Город', 'Критичных', 'Требуют внимания', 'К заказу (₽)', 'Стоимость остатков (₽)']
        for col, header in enumerate(headers, 1):
            ws_summary.cell(row=10, column=col, value=header).font = Font(bold=True)
        
        for row, (warehouse_key, data) in enumerate(recommendations.items(), 11):
            ws_summary.cell(row=row, column=1, value=data['name'])
            ws_summary.cell(row=row, column=2, value=data['city'])
            ws_summary.cell(row=row, column=3, value=data['critical_items'])
            ws_summary.cell(row=row, column=4, value=data['warning_items'])
            ws_summary.cell(row=row, column=5, value=data['total_order_value'])
            ws_summary.cell(row=row, column=6, value=data['total_stock_value'])
    
    # Лист 2: Детальные данные
    ws_details = wb.create_sheet("Детальные данные")
    
    # Создаем детальную таблицу
    detail_data = []
    for item in analysis_results:
        base_row = {
            'Номенклатура': item['номенклатура'],
            'ADS': item['ads'],
            'Цена': item['price'],
            'Общий остаток': item['total_stock'],
            'Стоимость остатков': item['total_stock_value'],
            'Статус': item['overall_status'],
            'Месяцев запаса': item['min_months_across_warehouses']
        }
        
        # Добавляем данные по складам
        for warehouse_key, wh_data in item['warehouses'].items():
            base_row[f"{wh_data['short_name']}_остаток"] = wh_data['current_stock']
            base_row[f"{wh_data['short_name']}_заказать"] = wh_data['order_quantity']
            base_row[f"{wh_data['short_name']}_статус"] = wh_data['status']
        
        detail_data.append(base_row)
    
    # Преобразуем в DataFrame и добавляем в Excel
    detail_df = pd.DataFrame(detail_data)
    
    for r in dataframe_to_rows(detail_df, index=False, header=True):
        ws_details.append(r)
    
    # Форматирование заголовков
    for cell in ws_details[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Лист 3: Товары к заказу
    ws_orders = wb.create_sheet("К заказу")
    
    # Собираем товары которые нужно заказать
    order_data = []
    for item in analysis_results:
        for warehouse_key, wh_data in item['warehouses'].items():
            if wh_data['order_quantity'] > 0:
                order_data.append({
                    'Номенклатура': item['номенклатура'],
                    'Склад': wh_data['short_name'],
                    'Текущий остаток': wh_data['current_stock'],
                    'Минимум': wh_data['min_stock'],
                    'К заказу': wh_data['order_quantity'],
                    'Цена': item['price'],
                    'Сумма заказа': wh_data['order_value'],
                    'ADS': item['ads']
                })
    
    if order_data:
        order_df = pd.DataFrame(order_data)
        for r in dataframe_to_rows(order_df, index=False, header=True):
            ws_orders.append(r)
        
        # Форматирование
        for cell in ws_orders[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Сохраняем в байты
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


# Инструкции по интеграции
def get_integration_instructions():
    """
    Возвращает инструкции по интеграции исправления
    """
    
    return """
# 🛠️ ИНСТРУКЦИЯ ПО ПРИМЕНЕНИЮ ИСПРАВЛЕНИЙ АНАЛИЗА СКЛАДОВ

## 📋 Что исправляет этот код:

1. **❌ Ошибка 'номенклатура'** - неправильное чтение структуры файла остатков
2. **❌ Отсутствующие методы** - analyze_warehouse_stock_with_details и другие
3. **❌ Неправильная обработка файлов** - сложная структура Excel файлов
4. **❌ Отсутствие цен** - интеграция с ценами из ADS
5. **❌ Плохое отображение результатов** - улучшенная визуализация

## 🚀 БЫСТРАЯ ИНТЕГРАЦИЯ:

### Шаг 1: Создайте файл `warehouse_complete_fix.py`
Скопируйте весь код из артефакта

### Шаг 2: Добавьте в ваш основной streamlit файл

```python
# В начале файла добавьте импорт:
from warehouse_complete_fix import (
    apply_warehouse_complete_fix,
    create_fixed_warehouse_analysis_page
)

# В функции где определяются страницы добавьте:
elif page == "📦 Анализ складов":
    warehouse_analysis_page_fixed = create_fixed_warehouse_analysis_page()
    warehouse_analysis_page_fixed(system)
```

### Шаг 3: ИЛИ замените существующую функцию

Если у вас уже есть `warehouse_analysis_page`, замените её на:

```python
def warehouse_analysis_page(system):
    # Применяем исправления
    if not hasattr(system, '_warehouse_fix_applied'):
        from warehouse_complete_fix import apply_warehouse_complete_fix
        apply_warehouse_complete_fix(system)
    
    # Вызываем исправленную версию
    from warehouse_complete_fix import create_fixed_warehouse_analysis_page
    fixed_page = create_fixed_warehouse_analysis_page()
    fixed_page(system)
```

## ✅ РЕЗУЛЬТАТ ПОСЛЕ ПРИМЕНЕНИЯ:

1. **Правильное чтение файлов остатков** любой структуры
2. **Автоматическое определение складов** и их настроек
3. **Интеграция с ценами** из ADS данных
4. **Детальный анализ** по каждому складу
5. **Красивое отображение результатов** с фильтрами
6. **Excel экспорт** с детальными отчетами
7. **Рекомендации по заказам** в денежном выражении

## 📊 СТРУКТУРА ПОДДЕРЖИВАЕМЫХ ФАЙЛОВ:

Код автоматически определяет:
- Строку с заголовком "Номенклатура"
- Колонки складов (содержащие "склад" или "магазин")
- Строки с данными товаров
- Дополнительные заголовки ("Количество", "Остаток")

## 🎯 ИСПОЛЬЗОВАНИЕ:

1. Рассчитайте ADS в разделе "ADS расчет"
2. Перейдите в "Анализ складов"
3. Загрузите файл остатков (любой структуры)
4. Настройте параметры анализа
5. Запустите анализ
6. Просмотрите результаты и экспортируйте отчет

## 🔧 ОТЛАДКА:

Включите "Режим отладки" для просмотра:
- Процесса чтения файла
- Найденных заголовков и складов
- Обработанных данных
- Подробностей анализа
"""


if __name__ == "__main__":
    print("🔧 Модуль исправлений анализа складов загружен")
    print(get_integration_instructions())